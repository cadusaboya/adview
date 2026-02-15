import logging
import decimal
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q, Sum, F, Count, Prefetch
from django.db.models.functions import Coalesce
from decimal import Decimal
from .mixins import CompanyScopedViewSetMixin
from ..models import Payment, ContaBancaria, Custodia, Transfer, Allocation, Receita, Despesa
from ..serializers import PaymentSerializer, ContaBancariaSerializer, CustodiaSerializer, TransferSerializer, AllocationSerializer
from ..pagination import DynamicPageSizePagination
from datetime import datetime, date

logger = logging.getLogger(__name__)


class PaymentViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """
    API endpoint para registrar pagamentos neutros (entrada/saída de caixa).
    As alocações para Receitas/Despesas/Passivos são feitas via Allocation.
    """
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'conta_bancaria'
        ).prefetch_related(
            Prefetch(
                'allocations',
                queryset=Allocation.objects.select_related(
                    'receita__cliente',
                    'despesa__responsavel',
                    'custodia__cliente',
                    'custodia__funcionario',
                    'transfer__from_bank',
                    'transfer__to_bank'
                )
            )
        )
        params = self.request.query_params

        # Filtros por data
        start_date = params.get('start_date')
        end_date = params.get('end_date')

        if start_date:
            queryset = queryset.filter(data_pagamento__gte=start_date)
        if end_date:
            queryset = queryset.filter(data_pagamento__lte=end_date)

        # Filtro por tipo (Entrada/Saída)
        tipo = params.get('tipo')  # 'E' | 'S'
        if tipo:
            queryset = queryset.filter(tipo=tipo)

        # Filtro por conta bancária
        conta_bancaria_id = params.get('conta_bancaria_id')
        if conta_bancaria_id:
            queryset = queryset.filter(conta_bancaria_id=conta_bancaria_id)

        # Filtro por situação da receita/despesa
        situacao = params.get('situacao')  # 'P' | 'A' | 'V'
        if situacao:
            # Filtra payments que têm allocations com receitas ou despesas na situação especificada
            queryset = queryset.filter(
                Q(allocations__receita__situacao=situacao) |
                Q(allocations__despesa__situacao=situacao)
            ).distinct()

        # Busca global
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(valor__icontains=search) |
                Q(observacao__icontains=search) |
                Q(data_pagamento__icontains=search) |
                # Filtro por nome do banco
                Q(conta_bancaria__nome__icontains=search) |
                # Filtro por entidades vinculadas
                Q(allocations__receita__nome__icontains=search) |
                Q(allocations__despesa__nome__icontains=search) |
                Q(allocations__custodia__nome__icontains=search) |
                Q(allocations__transfer__from_bank__nome__icontains=search) |
                Q(allocations__transfer__to_bank__nome__icontains=search)
            ).distinct()

        return queryset.order_by('-data_pagamento', '-id')

    def perform_create(self, serializer):
        from django.db.models import F
        from django.db import transaction

        with transaction.atomic():
            payment = serializer.save(company=self.request.user.company)

            # Atualiza saldo da conta bancária usando F() para operação atômica
            # select_for_update() garante que não haja race condition
            if payment.tipo == 'E':
                # Entrada de dinheiro (+)
                ContaBancaria.objects.select_for_update().filter(
                    pk=payment.conta_bancaria.pk
                ).update(saldo_atual=F('saldo_atual') + payment.valor)
            else:
                # Saída de dinheiro (-)
                ContaBancaria.objects.select_for_update().filter(
                    pk=payment.conta_bancaria.pk
                ).update(saldo_atual=F('saldo_atual') - payment.valor)

    def perform_update(self, serializer):
        from django.db.models import F
        from django.db import transaction

        with transaction.atomic():
            # Guarda informações antigas antes de atualizar
            old_payment = Payment.objects.select_for_update().get(pk=serializer.instance.pk)
            old_valor = old_payment.valor
            old_tipo = old_payment.tipo
            old_conta = old_payment.conta_bancaria

            payment = serializer.save()

            # Reverte a operação antiga usando F() para operação atômica
            if old_tipo == 'E':
                ContaBancaria.objects.select_for_update().filter(pk=old_conta.pk).update(
                    saldo_atual=F('saldo_atual') - old_valor
                )
            else:
                ContaBancaria.objects.select_for_update().filter(pk=old_conta.pk).update(
                    saldo_atual=F('saldo_atual') + old_valor
                )

            # Aplica a operação nova usando F() para operação atômica
            if payment.tipo == 'E':
                ContaBancaria.objects.select_for_update().filter(
                    pk=payment.conta_bancaria.pk
                ).update(saldo_atual=F('saldo_atual') + payment.valor)
            else:
                ContaBancaria.objects.select_for_update().filter(
                    pk=payment.conta_bancaria.pk
                ).update(saldo_atual=F('saldo_atual') - payment.valor)

            # Atualiza status de todas as contas alocadas
            for allocation in payment.allocations.all():
                if allocation.receita:
                    allocation.receita.atualizar_status()
                elif allocation.despesa:
                    allocation.despesa.atualizar_status()
                elif allocation.custodia:
                    allocation.custodia.atualizar_status()

    @action(detail=False, methods=['post'], url_path='import-extrato')
    def import_extrato(self, request):
        """
        Importa pagamentos a partir de um arquivo XLSX de extrato bancário do BTG.
        Espera:
        - file: arquivo XLSX
        - conta_bancaria_id: ID da conta bancária
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Arquivo não fornecido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if 'conta_bancaria_id' not in request.data:
            return Response(
                {'error': 'conta_bancaria_id não fornecido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES['file']
        conta_bancaria_id = request.data['conta_bancaria_id']

        # Verifica se a conta bancária existe e pertence ao usuário
        try:
            conta_bancaria = ContaBancaria.objects.get(
                id=conta_bancaria_id,
                company=request.user.company
            )
        except ContaBancaria.DoesNotExist:
            return Response(
                {'error': 'Conta bancária não encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            from openpyxl import load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
            from django.db.models import F

            # Carrega o workbook
            try:
                wb = load_workbook(file, data_only=True)
            except InvalidFileException:
                return Response(
                    {'error': 'Arquivo inválido. Por favor, envie um arquivo XLSX válido.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            ws = wb.active

            # Procura a linha de cabeçalho
            # Formato BTG: "Data de lançamento | Descrição do lançamento | Entradas / Saídas (R$) | Saldo (R$)"
            header_row = None
            date_col = None
            value_col = None
            desc_col = None

            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
                if row and any(row):
                    # Procura por palavras-chave nas células
                    for col_idx, cell_value in enumerate(row):
                        if cell_value and isinstance(cell_value, str):
                            cell_lower = cell_value.lower().strip()
                            # Remove acentos para comparação
                            import unicodedata
                            cell_normalized = ''.join(
                                c for c in unicodedata.normalize('NFD', cell_lower)
                                if unicodedata.category(c) != 'Mn'
                            )

                            # Procura coluna de data (Data de lançamento, Data, etc.)
                            if date_col is None and 'data' in cell_normalized:
                                date_col = col_idx

                            # Procura coluna de valor (Entradas/Saídas, Valor, etc.)
                            # Ignora coluna "Saldo"
                            if value_col is None and 'saldo' not in cell_normalized:
                                if any(kw in cell_normalized for kw in ['entrada', 'saida', 'valor', 'movimentacao']):
                                    value_col = col_idx

                            # Procura coluna de descrição (mas não a coluna de data)
                            # Prioriza "Descrição do lançamento" (formato BTG) ou "Histórico"
                            if desc_col is None and 'data' not in cell_normalized:
                                # Verifica primeiro se é exatamente "descrição do lançamento" ou similar
                                if 'lancamento' in cell_normalized and 'descri' in cell_normalized:
                                    desc_col = col_idx
                                # Caso contrário, aceita qualquer coluna com descrição ou histórico
                                elif any(kw in cell_normalized for kw in ['descri', 'historico']):
                                    desc_col = col_idx

                    # Se encontrou data E valor, considera essa linha como cabeçalho
                    if date_col is not None and value_col is not None:
                        header_row = idx
                        break

            if header_row is None or date_col is None:
                return Response(
                    {'error': 'Não foi possível identificar a coluna de data no extrato.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if value_col is None:
                return Response(
                    {'error': 'Não foi possível identificar a coluna de valor (Entradas/Saídas) no extrato.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Processa as linhas de dados
            created_count = 0
            skipped_count = 0  # Pagamentos duplicados ignorados
            errors = []
            potential_duplicates = []  # Lista de duplicatas potenciais para confirmação do usuário
            payments_to_create = []  # Lista de pagamentos a serem criados (apenas na segunda passagem)
            total_entradas = Decimal('0.00')
            total_saidas = Decimal('0.00')

            # Verifica se há lista de linhas confirmadas para importar (segunda passagem)
            force_import_lines = request.data.get('force_import_lines', [])
            confirmed = request.data.get('confirmed', 'false').lower() == 'true'

            if isinstance(force_import_lines, str):
                import json
                try:
                    force_import_lines = json.loads(force_import_lines)
                except json.JSONDecodeError as e:
                    # Se não conseguir parsear o JSON, assume lista vazia
                    force_import_lines = []
                    print(f"Aviso: Erro ao parsear force_import_lines JSON: {e}")

            # Se confirmed=true, significa que o usuário já viu o diálogo e escolheu
            # Neste caso, apenas importar as linhas que estão em force_import_lines
            # e pular as duplicatas potenciais que não foram selecionadas

            # Recarrega a conta para garantir que temos o saldo mais recente
            conta_bancaria.refresh_from_db()
            saldo_inicial = conta_bancaria.saldo_atual

            for idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                if not row or not any(row):
                    continue

                try:
                    # Extrai data
                    date_value = row[date_col] if date_col < len(row) else None
                    if not date_value:
                        continue

                    # Converte data
                    if isinstance(date_value, datetime):
                        data_pagamento = date_value.date()
                    elif isinstance(date_value, date):
                        data_pagamento = date_value
                    else:
                        # Tenta parsear string
                        from datetime import datetime as dt
                        date_str = str(date_value).strip()

                        # Tenta diferentes formatos
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y']:
                            try:
                                data_pagamento = dt.strptime(date_str, fmt).date()
                                break
                            except ValueError:
                                continue
                        else:
                            errors.append(f'Linha {idx}: formato de data inválido: {date_value}')
                            continue

                    # Extrai valor
                    value_raw = row[value_col] if value_col < len(row) else None
                    if value_raw is None or value_raw == '' or value_raw == 0:
                        continue

                    # Converte valor
                    if isinstance(value_raw, (int, float, Decimal)):
                        # Converte para string primeiro para evitar problemas de precisão de float
                        valor_num = Decimal(str(value_raw))
                    else:
                        # Remove caracteres não numéricos exceto vírgula, ponto e sinal
                        valor_str = str(value_raw).strip()
                        valor_str = valor_str.replace('R$', '').replace(' ', '').replace('\xa0', '')

                        # Formato brasileiro: 1.234,56 ou 1234,56
                        # Remove pontos (separador de milhar) e substitui vírgula por ponto
                        if ',' in valor_str:
                            valor_str = valor_str.replace('.', '').replace(',', '.')
                        # Se não tem vírgula mas tem ponto, assume formato americano ou já está correto

                        try:
                            valor_num = Decimal(valor_str)
                        except (ValueError, decimal.InvalidOperation) as e:
                            errors.append(f'Linha {idx}: formato de valor inválido: {value_raw} (erro: {str(e)})')
                            continue

                    # Quantiza para 2 casas decimais para garantir precisão
                    valor_num = valor_num.quantize(Decimal('0.01'))

                    # Determina tipo (Entrada ou Saída) baseado no sinal
                    if valor_num > 0:
                        tipo = 'E'  # Entrada
                        valor = valor_num  # Não precisa de abs() pois já é positivo
                    else:
                        tipo = 'S'  # Saída
                        valor = abs(valor_num)

                    # Extrai descrição/observação
                    observacao = ''
                    if desc_col is not None and desc_col < len(row):
                        desc_value = row[desc_col]
                        if desc_value:
                            observacao = str(desc_value).strip()

                    # ============================================
                    # VERIFICAÇÃO DE DUPLICATAS
                    # ============================================
                    # Busca pagamentos existentes com mesma data e valor (em QUALQUER banco)
                    existing_payments = Payment.objects.filter(
                        company=request.user.company,
                        data_pagamento=data_pagamento,
                        valor=valor,
                        tipo=tipo
                    )

                    # Verifica se existe duplicata exata (incluindo observação)
                    duplicata_exata = existing_payments.filter(observacao=observacao).exists()

                    if duplicata_exata:
                        # Duplicata exata encontrada - pular este pagamento
                        skipped_count += 1
                        continue

                    # Verifica se existe duplicata potencial (data + valor, mas observação diferente)
                    duplicata_potencial = existing_payments.exclude(observacao=observacao).first()

                    if duplicata_potencial:
                        # Se NÃO confirmado, apenas adiciona à lista (não importa nada ainda)
                        if not confirmed:
                            potential_duplicates.append({
                                'line_index': idx,
                                'new_payment': {
                                    'data': data_pagamento.strftime('%Y-%m-%d'),
                                    'valor': str(valor),
                                    'tipo': tipo,
                                    'observacao': observacao or '',
                                    'banco': conta_bancaria.nome
                                },
                                'existing_payment': {
                                    'id': duplicata_potencial.id,
                                    'data': duplicata_potencial.data_pagamento.strftime('%Y-%m-%d'),
                                    'valor': str(duplicata_potencial.valor),
                                    'tipo': duplicata_potencial.tipo,
                                    'observacao': duplicata_potencial.observacao or '',
                                    'banco': duplicata_potencial.conta_bancaria.nome
                                }
                            })
                            continue  # Pula e não adiciona à lista de criação

                        # Se confirmed=true, verifica se usuário escolheu importar esta linha
                        if idx not in force_import_lines:
                            # Usuário escolheu não importar esta duplicata
                            skipped_count += 1
                            continue

                    # Adiciona à lista de pagamentos a serem criados
                    payments_to_create.append({
                        'company': request.user.company,
                        'conta_bancaria': conta_bancaria,
                        'tipo': tipo,
                        'valor': valor,
                        'data_pagamento': data_pagamento,
                        'observacao': observacao
                    })

                except Exception as e:
                    errors.append(f'Linha {idx}: {str(e)}')

            # =====================================================
            # VERIFICAÇÃO FINAL: Se houver duplicatas potenciais na primeira passagem,
            # retorna SEM criar nenhum pagamento
            # =====================================================
            if potential_duplicates and not confirmed:
                response_data = {
                    'success': False,
                    'requires_confirmation': True,
                    'potential_duplicates': potential_duplicates,
                    'message': f'Encontradas {len(potential_duplicates)} possível(is) duplicata(s) que requerem confirmação.'
                }
                return Response(response_data, status=status.HTTP_200_OK)

            # =====================================================
            # CRIAÇÃO DOS PAGAMENTOS
            # Se chegou aqui, pode criar os pagamentos
            # =====================================================
            for payment_data in payments_to_create:
                try:
                    payment = Payment.objects.create(**payment_data)

                    # Atualiza saldo da conta usando F() para operação atômica
                    if payment_data['tipo'] == 'E':
                        ContaBancaria.objects.filter(pk=conta_bancaria.pk).update(
                            saldo_atual=F('saldo_atual') + payment_data['valor']
                        )
                        total_entradas += payment_data['valor']
                    else:
                        ContaBancaria.objects.filter(pk=conta_bancaria.pk).update(
                            saldo_atual=F('saldo_atual') - payment_data['valor']
                        )
                        total_saidas += payment_data['valor']

                    created_count += 1
                except Exception as e:
                    errors.append(f'Erro ao criar pagamento: {str(e)}')

            # Recarrega a conta para obter o saldo final atualizado
            conta_bancaria.refresh_from_db()
            saldo_final = conta_bancaria.saldo_atual

            # Calcula a diferença esperada vs real
            saldo_esperado = saldo_inicial + total_entradas - total_saidas
            diferenca = saldo_final - saldo_esperado

            # Se houver duplicatas potenciais (não deveria chegar aqui na primeira passagem)
            if potential_duplicates:
                response_data = {
                    'success': False,
                    'requires_confirmation': True,
                    'potential_duplicates': potential_duplicates,
                    'message': f'Encontradas {len(potential_duplicates)} possível(is) duplicata(s) que requerem confirmação.'
                }
                return Response(response_data, status=status.HTTP_200_OK)

            # Caso contrário, retorna sucesso normal
            response_data = {
                'success': True,
                'created_count': created_count,
                'skipped_count': skipped_count,  # Quantidade de pagamentos ignorados (duplicatas exatas)
                'conta_bancaria': conta_bancaria.nome,
                'saldo_inicial': str(saldo_inicial),
                'saldo_final': str(saldo_final),
                'total_entradas': str(total_entradas),
                'total_saidas': str(total_saidas),
                'saldo_esperado': str(saldo_esperado),
                'diferenca': str(diferenca)
            }

            if errors:
                response_data['errors'] = errors[:10]  # Limita a 10 erros
                response_data['total_errors'] = len(errors)

            return Response(response_data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'Erro ao processar arquivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='conciliar-bancario')
    def conciliar_bancario(self, request):
        """
        Concilia pagamentos sem alocação com receitas/despesas/custódias em aberto
        do mês especificado, fazendo match por valor e nome da contraparte na observação.

        Espera:
        - mes: int (1-12)
        - ano: int (ex: 2026)
        """
        import unicodedata
        import re

        def normalizar_string(texto):
            """Remove acentos e converte para lowercase para comparação"""
            if not texto:
                return ''
            # Normaliza unicode e remove acentos
            texto_nfd = unicodedata.normalize('NFD', str(texto))
            texto_sem_acentos = ''.join(
                char for char in texto_nfd
                if unicodedata.category(char) != 'Mn'
            )
            return texto_sem_acentos.lower().strip()

        def extrair_palavras_significativas(texto):
            """
            Extrai palavras significativas de um texto, removendo:
            - Preposições comuns (de, da, do, dos, das, para, pra, em, no, na, etc.)
            - Palavras muito curtas (< 3 caracteres)
            - Palavras bancárias comuns (pix, ted, transferencia, recebido, enviado, etc.)
            """
            if not texto:
                return set()

            # Normaliza o texto
            texto_norm = normalizar_string(texto)

            # Remove pontuação e divide em palavras
            palavras = re.findall(r'\b\w+\b', texto_norm)

            # Stop words - palavras a ignorar
            stop_words = {
                'de', 'da', 'do', 'dos', 'das', 'para', 'pra', 'em', 'no', 'na', 'nos', 'nas',
                'com', 'por', 'a', 'o', 'e', 'ou', 'um', 'uma', 'ao', 'aos', 'as',
                'pix', 'ted', 'transferencia', 'recebido', 'enviado', 'pagamento', 'recebimento',
                'valor', 'ref', 'referente', 'ltda', 'me', 'sa', 'eireli'
            }

            # Filtra palavras significativas
            palavras_significativas = {
                palavra for palavra in palavras
                if len(palavra) >= 3 and palavra not in stop_words
            }

            return palavras_significativas

        def match_por_palavras_comuns(observacao, *textos_para_comparar):
            """
            Verifica se há pelo menos 2 palavras significativas em comum entre
            a observação e qualquer um dos textos fornecidos (nome cliente, nome despesa, etc.)

            Retorna True se encontrar 2+ palavras em comum com qualquer texto.
            """
            if not observacao:
                return False

            palavras_obs = extrair_palavras_significativas(observacao)

            if len(palavras_obs) < 2:
                return False

            for texto in textos_para_comparar:
                if not texto:
                    continue

                palavras_texto = extrair_palavras_significativas(texto)
                palavras_em_comum = palavras_obs.intersection(palavras_texto)

                # Match válido se há 2 ou mais palavras significativas em comum
                if len(palavras_em_comum) >= 2:
                    return True

            return False

        def nome_em_observacao(observacao, *nomes):
            """
            Verifica se algum dos nomes está contido na observação.
            Agora usa duas estratégias:
            1. Match exato (nome completo na observação) - mais confiável
            2. Match por palavras comuns (2+ palavras) - mais flexível
            """
            if not observacao:
                return False

            # Estratégia 1: Match exato (nome completo aparece na observação)
            obs_norm = normalizar_string(observacao)
            for nome in nomes:
                if nome:
                    nome_norm = normalizar_string(nome)
                    if nome_norm and nome_norm in obs_norm:
                        return True

            # Estratégia 2: Match por palavras comuns (pelo menos 2 palavras)
            if match_por_palavras_comuns(observacao, *nomes):
                return True

            return False

        mes = request.data.get('mes')
        ano = request.data.get('ano')

        if not mes or not ano:
            return Response(
                {'error': 'Parâmetros mes e ano são obrigatórios'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            mes = int(mes)
            ano = int(ano)

            if mes < 1 or mes > 12:
                return Response(
                    {'error': 'Mês deve estar entre 1 e 12'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except ValueError:
            return Response(
                {'error': 'Mes e ano devem ser números inteiros'},
                status=status.HTTP_400_BAD_REQUEST
            )

        company = request.user.company

        # Busca payments sem alocação no mês especificado
        payments_sem_alocacao = Payment.objects.filter(
            company=company,
            data_pagamento__year=ano,
            data_pagamento__month=mes
        ).annotate(
            num_allocations=Count('allocations')
        ).filter(
            num_allocations=0
        ).order_by('data_pagamento')

        # Busca receitas em aberto ou vencidas no mês (não pagas)
        # Pre-calcula total alocado para evitar N+1 queries
        receitas_abertas = Receita.objects.filter(
            company=company,
            data_vencimento__year=ano,
            data_vencimento__month=mes,
            situacao__in=['A', 'V']  # Em Aberto ou Vencida
        ).annotate(
            total_alocado=Coalesce(Sum('allocations__valor'), Decimal('0.00'))
        ).order_by('data_vencimento')

        # Debug: total de receitas no mês (qualquer situação)
        total_receitas_mes = Receita.objects.filter(
            company=company,
            data_vencimento__year=ano,
            data_vencimento__month=mes
        ).count()

        # Busca despesas em aberto ou vencidas no mês (não pagas)
        # Pre-calcula total alocado para evitar N+1 queries
        despesas_abertas = Despesa.objects.filter(
            company=company,
            data_vencimento__year=ano,
            data_vencimento__month=mes,
            situacao__in=['A', 'V']  # Em Aberto ou Vencida
        ).annotate(
            total_alocado=Coalesce(Sum('allocations__valor'), Decimal('0.00'))
        ).order_by('data_vencimento')

        # Debug: total de despesas no mês (qualquer situação)
        total_despesas_mes = Despesa.objects.filter(
            company=company,
            data_vencimento__year=ano,
            data_vencimento__month=mes
        ).count()

        # Busca custódias em aberto
        # Pre-calcula totais de entradas e saídas para evitar N+1 queries
        custodias_abertas = Custodia.objects.filter(
            company=company,
            status='A'
        ).annotate(
            valor_restante=F('valor_total') - F('valor_liquidado'),
            total_entradas=Coalesce(
                Sum('allocations__valor', filter=Q(allocations__payment__tipo='E')),
                Decimal('0.00')
            ),
            total_saidas=Coalesce(
                Sum('allocations__valor', filter=Q(allocations__payment__tipo='S')),
                Decimal('0.00')
            )
        ).filter(
            valor_restante__gt=0
        )

        # Estatísticas
        matches_receitas = 0
        matches_despesas = 0
        matches_custodias = 0
        erros = []

        # Lista para armazenar sugestões de matches (apenas por valor, sem nome)
        sugestoes = []

        # Dicionários para rastrear alocações em tempo real durante o loop
        # Chave: ID da entidade, Valor: total adicional alocado neste loop
        receitas_alocadas_no_loop = {}
        despesas_alocadas_no_loop = {}
        custodias_alocadas_no_loop = {}

        # Processa cada payment sem alocação
        for payment in payments_sem_alocacao:
            match_found = False

            if payment.tipo == 'E':
                # Entrada: busca receitas com VALOR EXATO E NOME NA OBSERVAÇÃO (ambas condições obrigatórias)
                for receita in receitas_abertas:
                    # Calcula total alocado = annotation inicial + alocações feitas neste loop
                    total_alocado_atual = receita.total_alocado + receitas_alocadas_no_loop.get(receita.id, Decimal('0.00'))
                    valor_nao_alocado = receita.valor - total_alocado_atual

                    # Verifica se há saldo disponível e se o valor do payment é compatível
                    if valor_nao_alocado >= payment.valor and receita.valor == payment.valor:
                        # Verifica se o nome do cliente ou nome da receita está na observação
                        nome_encontrado = nome_em_observacao(
                            payment.observacao,
                            receita.cliente.nome if receita.cliente else None,
                            receita.nome
                        )

                        # APENAS faz match se AMBAS as condições forem verdadeiras
                        if nome_encontrado:
                            try:
                                Allocation.objects.create(
                                    company=company,
                                    payment=payment,
                                    receita=receita,
                                    valor=payment.valor
                                )
                                receita.atualizar_status()

                                # Atualiza o dicionário de alocações em tempo real
                                receitas_alocadas_no_loop[receita.id] = receitas_alocadas_no_loop.get(receita.id, Decimal('0.00')) + payment.valor

                                matches_receitas += 1
                                match_found = True
                                break  # Encontrou match válido, para de procurar
                            except Exception as e:
                                erros.append(f'Erro ao alocar payment {payment.id} para receita {receita.id}: {str(e)}')

                # Se não encontrou receita, tenta custodia tipo Ativo (VALOR E NOME obrigatórios)
                if not match_found:
                    for custodia in custodias_abertas:
                        if custodia.tipo == 'A':  # Ativo - a receber
                            # Calcula totais considerando alocações feitas neste loop
                            alocacoes_loop = custodias_alocadas_no_loop.get(custodia.id, {'entradas': Decimal('0.00'), 'saidas': Decimal('0.00')})
                            total_entradas = custodia.total_entradas + alocacoes_loop['entradas']
                            total_saidas = custodia.total_saidas + alocacoes_loop['saidas']

                            valor_liquidado = min(total_saidas, total_entradas)
                            valor_restante = custodia.valor_total - valor_liquidado

                            # Verifica se há saldo disponível E se o valor total da custódia é exato
                            if valor_restante >= payment.valor and custodia.valor_total == payment.valor:
                                # Verifica se nome do cliente/funcionário ou nome da custódia está na observação
                                nome_encontrado = nome_em_observacao(
                                    payment.observacao,
                                    custodia.cliente.nome if custodia.cliente else None,
                                    custodia.funcionario.nome if custodia.funcionario else None,
                                    custodia.nome
                                )

                                # APENAS faz match se AMBAS as condições forem verdadeiras
                                if nome_encontrado:
                                    try:
                                        Allocation.objects.create(
                                            company=company,
                                            payment=payment,
                                            custodia=custodia,
                                            valor=payment.valor
                                        )
                                        custodia.atualizar_status()

                                        # Atualiza o dicionário de alocações em tempo real (entrada para ativo)
                                        if custodia.id not in custodias_alocadas_no_loop:
                                            custodias_alocadas_no_loop[custodia.id] = {'entradas': Decimal('0.00'), 'saidas': Decimal('0.00')}
                                        custodias_alocadas_no_loop[custodia.id]['entradas'] += payment.valor

                                        matches_custodias += 1
                                        match_found = True
                                        break  # Encontrou match válido, para de procurar
                                    except Exception as e:
                                        erros.append(f'Erro ao alocar payment {payment.id} para custódia {custodia.id}: {str(e)}')

            elif payment.tipo == 'S':
                # Saída: busca despesas com VALOR EXATO E NOME NA OBSERVAÇÃO (ambas condições obrigatórias)
                for despesa in despesas_abertas:
                    # Calcula total alocado = annotation inicial + alocações feitas neste loop
                    total_alocado_atual = despesa.total_alocado + despesas_alocadas_no_loop.get(despesa.id, Decimal('0.00'))
                    valor_nao_alocado = despesa.valor - total_alocado_atual

                    # Verifica se há saldo disponível e se o valor do payment é compatível
                    if valor_nao_alocado >= payment.valor and despesa.valor == payment.valor:
                        # Verifica se o nome do responsável ou nome da despesa está na observação
                        nome_encontrado = nome_em_observacao(
                            payment.observacao,
                            despesa.responsavel.nome if despesa.responsavel else None,
                            despesa.nome
                        )

                        # APENAS faz match se AMBAS as condições forem verdadeiras
                        if nome_encontrado:
                            try:
                                Allocation.objects.create(
                                    company=company,
                                    payment=payment,
                                    despesa=despesa,
                                    valor=payment.valor
                                )
                                despesa.atualizar_status()

                                # Atualiza o dicionário de alocações em tempo real
                                despesas_alocadas_no_loop[despesa.id] = despesas_alocadas_no_loop.get(despesa.id, Decimal('0.00')) + payment.valor

                                matches_despesas += 1
                                match_found = True
                                break  # Encontrou match válido, para de procurar
                            except Exception as e:
                                erros.append(f'Erro ao alocar payment {payment.id} para despesa {despesa.id}: {str(e)}')

                # Se não encontrou despesa, tenta custodia tipo Passivo (VALOR E NOME obrigatórios)
                if not match_found:
                    for custodia in custodias_abertas:
                        if custodia.tipo == 'P':  # Passivo - a pagar
                            # Calcula totais considerando alocações feitas neste loop
                            alocacoes_loop = custodias_alocadas_no_loop.get(custodia.id, {'entradas': Decimal('0.00'), 'saidas': Decimal('0.00')})
                            total_entradas = custodia.total_entradas + alocacoes_loop['entradas']
                            total_saidas = custodia.total_saidas + alocacoes_loop['saidas']
                            valor_liquidado = min(total_entradas, total_saidas)
                            valor_restante = custodia.valor_total - valor_liquidado

                            # Verifica se há saldo disponível E se o valor total da custódia é exato
                            if valor_restante >= payment.valor and custodia.valor_total == payment.valor:
                                # Verifica se nome do cliente/funcionário ou nome da custódia está na observação
                                nome_encontrado = nome_em_observacao(
                                    payment.observacao,
                                    custodia.cliente.nome if custodia.cliente else None,
                                    custodia.funcionario.nome if custodia.funcionario else None,
                                    custodia.nome
                                )

                                # APENAS faz match se AMBAS as condições forem verdadeiras
                                if nome_encontrado:
                                    try:
                                        Allocation.objects.create(
                                            company=company,
                                            payment=payment,
                                            custodia=custodia,
                                            valor=payment.valor
                                        )
                                        custodia.atualizar_status()

                                        # Atualiza o dicionário de alocações em tempo real (saída para passivo)
                                        if custodia.id not in custodias_alocadas_no_loop:
                                            custodias_alocadas_no_loop[custodia.id] = {'entradas': Decimal('0.00'), 'saidas': Decimal('0.00')}
                                        custodias_alocadas_no_loop[custodia.id]['saidas'] += payment.valor

                                        matches_custodias += 1
                                        match_found = True
                                        break  # Encontrou match válido, para de procurar
                                    except Exception as e:
                                        erros.append(f'Erro ao alocar payment {payment.id} para custódia {custodia.id}: {str(e)}')

            # Se não houve match automático, coleta sugestões apenas por valor
            if not match_found:
                sugestoes_payment = []

                if payment.tipo == 'E':
                    # Sugestões de receitas com mesmo valor E saldo disponível
                    for receita in receitas_abertas:
                        # Calcula o valor não alocado
                        total_alocado = receita.allocations.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
                        valor_nao_alocado = receita.valor - total_alocado

                        # Só sugere se há saldo disponível suficiente
                        if valor_nao_alocado >= payment.valor and receita.valor == payment.valor:
                            sugestoes_payment.append({
                                'tipo': 'receita',
                                'entidade_id': receita.id,
                                'entidade_nome': receita.nome,
                                'entidade_cliente': receita.cliente.nome if receita.cliente else None,
                                'entidade_valor': str(receita.valor),
                                'entidade_vencimento': receita.data_vencimento.isoformat()
                            })

                    # Sugestões de custódias Ativo com mesmo valor E saldo disponível
                    for custodia in custodias_abertas:
                        if custodia.tipo == 'A':
                            # Calcula valor restante baseado nas allocations atuais
                            total_entradas = custodia.allocations.filter(
                                payment__tipo='E'
                            ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

                            total_saidas = custodia.allocations.filter(
                                payment__tipo='S'
                            ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

                            valor_liquidado = min(total_saidas, total_entradas)
                            valor_restante = custodia.valor_total - valor_liquidado

                            # Só sugere se há saldo disponível suficiente E valor é exatamente igual
                            if valor_restante >= payment.valor and valor_restante == payment.valor:
                                contraparte = None
                                if custodia.cliente:
                                    contraparte = custodia.cliente.nome
                                elif custodia.funcionario:
                                    contraparte = custodia.funcionario.nome

                                sugestoes_payment.append({
                                    'tipo': 'custodia',
                                    'entidade_id': custodia.id,
                                    'entidade_nome': custodia.nome,
                                    'entidade_contraparte': contraparte,
                                    'entidade_valor': str(valor_restante),
                                    'entidade_tipo': 'Ativo'
                                })

                elif payment.tipo == 'S':
                    # Sugestões de despesas com mesmo valor E saldo disponível
                    for despesa in despesas_abertas:
                        # Calcula o valor não alocado
                        total_alocado = despesa.allocations.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
                        valor_nao_alocado = despesa.valor - total_alocado

                        # Só sugere se há saldo disponível suficiente
                        if valor_nao_alocado >= payment.valor and despesa.valor == payment.valor:
                            sugestoes_payment.append({
                                'tipo': 'despesa',
                                'entidade_id': despesa.id,
                                'entidade_nome': despesa.nome,
                                'entidade_responsavel': despesa.responsavel.nome if despesa.responsavel else None,
                                'entidade_valor': str(despesa.valor),
                                'entidade_vencimento': despesa.data_vencimento.isoformat()
                            })

                    # Sugestões de custódias Passivo com mesmo valor E saldo disponível
                    for custodia in custodias_abertas:
                        if custodia.tipo == 'P':
                            # Calcula valor restante baseado nas allocations atuais
                            total_entradas = custodia.allocations.filter(
                                payment__tipo='E'
                            ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

                            total_saidas = custodia.allocations.filter(
                                payment__tipo='S'
                            ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

                            valor_liquidado = min(total_entradas, total_saidas)
                            valor_restante = custodia.valor_total - valor_liquidado

                            # Só sugere se há saldo disponível suficiente E valor é exatamente igual
                            if valor_restante >= payment.valor and valor_restante == payment.valor:
                                contraparte = None
                                if custodia.cliente:
                                    contraparte = custodia.cliente.nome
                                elif custodia.funcionario:
                                    contraparte = custodia.funcionario.nome

                                sugestoes_payment.append({
                                    'tipo': 'custodia',
                                    'entidade_id': custodia.id,
                                    'entidade_nome': custodia.nome,
                                    'entidade_contraparte': contraparte,
                                    'entidade_valor': str(valor_restante),
                                    'entidade_tipo': 'Passivo'
                                })

                # Se há sugestões, adiciona à lista
                if sugestoes_payment:
                    sugestoes.append({
                        'payment_id': payment.id,
                        'payment_tipo': payment.get_tipo_display(),
                        'payment_valor': str(payment.valor),
                        'payment_data': payment.data_pagamento.isoformat(),
                        'payment_observacao': payment.observacao or '',
                        'payment_conta': payment.conta_bancaria.nome if payment.conta_bancaria else None,
                        'opcoes': sugestoes_payment
                    })

        return Response({
            'success': True,
            'mes': mes,
            'ano': ano,
            'total_payments_processados': payments_sem_alocacao.count(),
            'matches': {
                'receitas': matches_receitas,
                'despesas': matches_despesas,
                'custodias': matches_custodias,
                'total': matches_receitas + matches_despesas + matches_custodias
            },
            'sugestoes': sugestoes,
            'total_sugestoes': len(sugestoes),
            'debug': {
                'total_receitas_abertas': receitas_abertas.count(),
                'total_despesas_abertas': despesas_abertas.count(),
                'total_custodias_abertas': custodias_abertas.count(),
                'payments_entrada': payments_sem_alocacao.filter(tipo='E').count(),
                'payments_saida': payments_sem_alocacao.filter(tipo='S').count(),
            },
            'erros': erros
        })

    @action(detail=False, methods=['post'], url_path='confirmar-sugestao')
    def confirmar_sugestao(self, request):
        """
        Confirma uma sugestão de match manual, criando a alocação.

        Espera:
        - payment_id: ID do pagamento
        - tipo: 'receita', 'despesa' ou 'custodia'
        - entidade_id: ID da entidade (receita/despesa/custodia)
        """
        payment_id = request.data.get('payment_id')
        tipo = request.data.get('tipo')
        entidade_id = request.data.get('entidade_id')

        if not all([payment_id, tipo, entidade_id]):
            return Response(
                {'error': 'Parâmetros payment_id, tipo e entidade_id são obrigatórios'},
                status=status.HTTP_400_BAD_REQUEST
            )

        company = request.user.company

        # Busca o payment
        try:
            payment = Payment.objects.get(id=payment_id, company=company)
        except Payment.DoesNotExist:
            return Response(
                {'error': 'Pagamento não encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Verifica se o payment já tem alocação
        if payment.allocations.exists():
            return Response(
                {'error': 'Este pagamento já possui alocação'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Cria a alocação baseado no tipo
        try:
            if tipo == 'receita':
                receita = Receita.objects.get(id=entidade_id, company=company)
                if receita.valor != payment.valor:
                    return Response(
                        {'error': 'Valor da receita não corresponde ao valor do pagamento'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                Allocation.objects.create(
                    company=company,
                    payment=payment,
                    receita=receita,
                    valor=payment.valor
                )
                receita.atualizar_status()
                entidade_nome = receita.nome

            elif tipo == 'despesa':
                despesa = Despesa.objects.get(id=entidade_id, company=company)
                if despesa.valor != payment.valor:
                    return Response(
                        {'error': 'Valor da despesa não corresponde ao valor do pagamento'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                Allocation.objects.create(
                    company=company,
                    payment=payment,
                    despesa=despesa,
                    valor=payment.valor
                )
                despesa.atualizar_status()
                entidade_nome = despesa.nome

            elif tipo == 'custodia':
                custodia = Custodia.objects.get(id=entidade_id, company=company)
                valor_restante = custodia.valor_total - custodia.valor_liquidado
                if valor_restante != payment.valor:
                    return Response(
                        {'error': 'Valor restante da custódia não corresponde ao valor do pagamento'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                Allocation.objects.create(
                    company=company,
                    payment=payment,
                    custodia=custodia,
                    valor=payment.valor
                )
                custodia.atualizar_status()
                entidade_nome = custodia.nome

            else:
                return Response(
                    {'error': 'Tipo inválido. Use: receita, despesa ou custodia'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            return Response({
                'success': True,
                'message': f'Pagamento vinculado com sucesso a {tipo}: {entidade_nome}',
                'payment_id': payment.id,
                'tipo': tipo,
                'entidade_id': entidade_id
            })

        except (Receita.DoesNotExist, Despesa.DoesNotExist, Custodia.DoesNotExist):
            return Response(
                {'error': f'{tipo.capitalize()} não encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Erro ao criar alocação: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def perform_destroy(self, instance):
        from django.db.models import F

        conta_id = instance.conta_bancaria.pk
        valor = instance.valor
        tipo = instance.tipo

        # Guarda as alocações antes de deletar
        allocations = list(instance.allocations.all())

        # Deleta o pagamento primeiro
        instance.delete()

        # Reverte o pagamento do saldo usando F() para operação atômica
        if tipo == 'E':
            # Remove entrada
            ContaBancaria.objects.filter(pk=conta_id).update(
                saldo_atual=F('saldo_atual') - valor
            )
        else:
            # Remove saída
            ContaBancaria.objects.filter(pk=conta_id).update(
                saldo_atual=F('saldo_atual') + valor
            )

        # Atualiza status de todas as contas que estavam alocadas
        for allocation in allocations:
            if allocation.receita:
                allocation.receita.atualizar_status()
            elif allocation.despesa:
                allocation.despesa.atualizar_status()
            elif allocation.custodia:
                allocation.custodia.atualizar_status()

class ContaBancariaViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """API endpoint para gerenciar contas bancárias."""
    queryset = ContaBancaria.objects.all()
    serializer_class = ContaBancariaSerializer
    pagination_class = DynamicPageSizePagination


class CustodiaViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """API endpoint para gerenciar custódias (valores de terceiros - ativos e passivos)."""
    queryset = Custodia.objects.all()
    serializer_class = CustodiaSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        queryset = super().get_queryset().select_related('cliente', 'funcionario')

        params = self.request.query_params

        # Filtro por tipo (Ativo ou Passivo)
        tipo_filter = params.get('tipo')
        if tipo_filter:
            queryset = queryset.filter(tipo=tipo_filter)

        # Filtro por status (aceita múltiplos valores)
        status_filter = params.getlist('status')
        if status_filter:
            queryset = queryset.filter(status__in=status_filter)
        elif params.get('status'):
            # Fallback para single value
            queryset = queryset.filter(status=params.get('status'))

        # Filtro por cliente
        cliente_id = params.get('cliente_id')
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)

        # Filtro por funcionário
        funcionario_id = params.get('funcionario_id')
        if funcionario_id:
            queryset = queryset.filter(funcionario_id=funcionario_id)

        # Busca global
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) |
                Q(descricao__icontains=search) |
                Q(cliente__nome__icontains=search) |
                Q(funcionario__nome__icontains=search)
            )

        return queryset.order_by('-criado_em', 'id')


class TransferViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """API endpoint para gerenciar transferências entre contas bancárias."""
    queryset = Transfer.objects.all()
    serializer_class = TransferSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        from django.db.models.functions import Coalesce

        queryset = super().get_queryset().select_related('from_bank', 'to_bank').annotate(
            valor_saida=Coalesce(
                Sum('allocations__valor', filter=Q(allocations__payment__tipo='S')),
                Decimal('0.00')
            ),
            valor_entrada=Coalesce(
                Sum('allocations__valor', filter=Q(allocations__payment__tipo='E')),
                Decimal('0.00')
            )
        )

        params = self.request.query_params

        # Filtro por banco de origem
        from_bank_id = params.get('from_bank_id')
        if from_bank_id:
            queryset = queryset.filter(from_bank_id=from_bank_id)

        # Filtro por banco de destino
        to_bank_id = params.get('to_bank_id')
        if to_bank_id:
            queryset = queryset.filter(to_bank_id=to_bank_id)

        # Filtro por status (aceita múltiplos valores)
        status_filter = params.getlist('status')
        if status_filter:
            queryset = queryset.filter(status__in=status_filter)
        elif params.get('status'):
            # Fallback para single value
            queryset = queryset.filter(status=params.get('status'))

        # Filtro por data
        data_inicio = params.get('data_inicio')
        if data_inicio:
            queryset = queryset.filter(data_transferencia__gte=data_inicio)

        data_fim = params.get('data_fim')
        if data_fim:
            queryset = queryset.filter(data_transferencia__lte=data_fim)

        # Busca global
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(descricao__icontains=search) |
                Q(from_bank__nome__icontains=search) |
                Q(to_bank__nome__icontains=search)
            )

        return queryset.order_by('-data_transferencia', '-criado_em', 'id')


class AllocationViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """API endpoint para gerenciar alocações de pagamentos."""
    queryset = Allocation.objects.all()
    serializer_class = AllocationSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'payment', 'payment__conta_bancaria',
            'receita', 'receita__cliente',
            'despesa', 'despesa__responsavel',
            'custodia', 'custodia__cliente', 'custodia__funcionario',
            'transfer', 'transfer__from_bank', 'transfer__to_bank'
        )

        params = self.request.query_params

        # Filtro por payment
        payment_id = params.get('payment_id')
        if payment_id:
            queryset = queryset.filter(payment_id=payment_id)

        # Filtro por receita
        receita_id = params.get('receita_id')
        if receita_id:
            queryset = queryset.filter(receita_id=receita_id)

        # Filtro por despesa
        despesa_id = params.get('despesa_id')
        if despesa_id:
            queryset = queryset.filter(despesa_id=despesa_id)

        # Filtro por custódia
        custodia_id = params.get('custodia_id')
        if custodia_id:
            queryset = queryset.filter(custodia_id=custodia_id)

        # Filtro por transferência
        transfer_id = params.get('transfer_id')
        if transfer_id:
            queryset = queryset.filter(transfer_id=transfer_id)

        # Filtro por tipo de conta
        tipo_conta = params.get('tipo_conta')  # 'receita', 'despesa', 'custodia', 'transfer'
        if tipo_conta == 'receita':
            queryset = queryset.filter(receita__isnull=False)
        elif tipo_conta == 'despesa':
            queryset = queryset.filter(despesa__isnull=False)
        elif tipo_conta == 'custodia':
            queryset = queryset.filter(custodia__isnull=False)
        elif tipo_conta == 'transfer':
            queryset = queryset.filter(transfer__isnull=False)

        # Busca global
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(observacao__icontains=search) |
                Q(valor__icontains=search) |
                Q(payment__observacao__icontains=search) |
                Q(receita__nome__icontains=search) |
                Q(despesa__nome__icontains=search) |
                Q(custodia__nome__icontains=search)
            )

        return queryset.order_by('-criado_em', 'id')

    def perform_create(self, serializer):
        allocation = serializer.save(company=self.request.user.company)

        # Atualizar status da conta após criar alocação
        if allocation.receita:
            allocation.receita.atualizar_status()
        elif allocation.despesa:
            allocation.despesa.atualizar_status()
        elif allocation.custodia:
            allocation.custodia.atualizar_status()
        elif allocation.transfer:
            allocation.transfer.atualizar_status()

    def perform_update(self, serializer):
        # Guarda referências antigas antes de atualizar
        old_allocation = Allocation.objects.get(pk=serializer.instance.pk)
        old_receita = old_allocation.receita
        old_despesa = old_allocation.despesa
        old_custodia = old_allocation.custodia
        old_transfer = old_allocation.transfer

        # Salva a nova alocação
        allocation = serializer.save()

        # Atualiza status da conta antiga (se mudou)
        if old_receita and old_receita != allocation.receita:
            old_receita.atualizar_status()
        if old_despesa and old_despesa != allocation.despesa:
            old_despesa.atualizar_status()
        if old_custodia and old_custodia != allocation.custodia:
            old_custodia.atualizar_status()
        if old_transfer and old_transfer != allocation.transfer:
            old_transfer.atualizar_status()

        # Atualiza status da conta nova
        if allocation.receita:
            allocation.receita.atualizar_status()
        elif allocation.despesa:
            allocation.despesa.atualizar_status()
        elif allocation.custodia:
            allocation.custodia.atualizar_status()
        elif allocation.transfer:
            allocation.transfer.atualizar_status()

    def perform_destroy(self, instance):
        receita = instance.receita
        despesa = instance.despesa
        custodia = instance.custodia
        transfer = instance.transfer

        # Deleta a alocação
        instance.delete()

        # Atualiza status da conta após deletar alocação
        if receita:
            receita.atualizar_status()
        if despesa:
            despesa.atualizar_status()
        if custodia:
            custodia.atualizar_status()
        if transfer:
            transfer.atualizar_status()
