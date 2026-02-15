import logging
import json
import secrets
from rest_framework import viewsets, permissions, status, generics, serializers
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.throttling import UserRateThrottle
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.conf import settings as django_settings
from django.db import transaction
from requests.exceptions import RequestException, HTTPError

logger = logging.getLogger(__name__)
from django.db.models import Sum, Q, F, Case, When, IntegerField, Count, Prefetch, DecimalField
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.timezone import now
from datetime import date, datetime, timedelta
import decimal
from decimal import Decimal
from .pagination import DynamicPageSizePagination
from django.shortcuts import get_object_or_404
from .models import Company, CustomUser, Cliente, Funcionario, Receita, ReceitaRecorrente, Despesa, DespesaRecorrente, Payment, ContaBancaria, Custodia, Transfer, Allocation, PlanoAssinatura, AssinaturaEmpresa, WebhookLog
from .serializers import (
    CompanySerializer, CustomUserSerializer, ClienteSerializer,
    FuncionarioSerializer, ReceitaSerializer, ReceitaAbertaSerializer, ReceitaRecorrenteSerializer, DespesaSerializer, DespesaAbertaSerializer,
    DespesaRecorrenteSerializer, PaymentSerializer, ContaBancariaSerializer, CustodiaSerializer, TransferSerializer, AllocationSerializer,
    PlanoAssinaturaSerializer, AssinaturaEmpresaSerializer,
)
from .permissions import IsSubscriptionActive
from .asaas_service import criar_cliente_asaas, atualizar_cliente_asaas, criar_assinatura_cartao_asaas, atualizar_cartao_assinatura, cancelar_assinatura_asaas


def _add_one_year_safe(base_date):
    """Adds one year preserving day when possible, clamping to month end otherwise."""
    import calendar

    target_year = base_date.year + 1
    last_day = calendar.monthrange(target_year, base_date.month)[1]
    return base_date.replace(year=target_year, day=min(base_date.day, last_day))


def _add_one_month_safe(base_date):
    """Adds one month preserving day when possible, clamping to month end otherwise."""
    import calendar

    month = base_date.month % 12 + 1
    year = base_date.year + (base_date.month // 12)
    last_day = calendar.monthrange(year, month)[1]
    return base_date.replace(year=year, month=month, day=min(base_date.day, last_day))


def _normalize_digits(value: str) -> str:
    return ''.join(ch for ch in str(value or '') if ch.isdigit())


def _is_valid_cpf(cpf: str) -> bool:
    cpf = _normalize_digits(cpf)
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False

    total = sum(int(cpf[i]) * (10 - i) for i in range(9))
    first_digit = ((total * 10) % 11) % 10
    if first_digit != int(cpf[9]):
        return False

    total = sum(int(cpf[i]) * (11 - i) for i in range(10))
    second_digit = ((total * 10) % 11) % 10
    return second_digit == int(cpf[10])


def _is_valid_cnpj(cnpj: str) -> bool:
    cnpj = _normalize_digits(cnpj)
    if len(cnpj) != 14 or cnpj == cnpj[0] * 14:
        return False

    weights_1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    total = sum(int(cnpj[i]) * weights_1[i] for i in range(12))
    remainder = total % 11
    first_digit = 0 if remainder < 2 else 11 - remainder
    if first_digit != int(cnpj[12]):
        return False

    weights_2 = [6] + weights_1
    total = sum(int(cnpj[i]) * weights_2[i] for i in range(13))
    remainder = total % 11
    second_digit = 0 if remainder < 2 else 11 - remainder
    return second_digit == int(cnpj[13])


def _is_valid_cpf_cnpj(value: str) -> bool:
    digits = _normalize_digits(value)
    if len(digits) == 11:
        return _is_valid_cpf(digits)
    if len(digits) == 14:
        return _is_valid_cnpj(digits)
    return False


class PaymentRateThrottle(UserRateThrottle):
    scope = 'payment'


# --- Base ViewSet for Company context ---
class CompanyScopedViewSetMixin:
    """Mixin to scope querysets and creation to the user's company."""
    permission_classes = [permissions.IsAuthenticated, IsSubscriptionActive]

    def get_queryset(self):
        """Filter queryset to only include objects belonging to the user's company."""
        user = self.request.user
        if user.is_superuser:
            # Superusers can see all companies' data (adjust if needed)
            return self.queryset.all()
        if hasattr(user, 'company') and user.company:
            return self.queryset.filter(company=user.company)
        # If user has no company, they see nothing (or handle as error)
        return self.queryset.none()

    def perform_create(self, serializer):
        """Automatically assign the user's company during creation."""
        user = self.request.user
        if hasattr(user, 'company') and user.company:
            serializer.save(company=user.company)
        elif self.request.data.get('company_id') and user.is_superuser:
             # Allow superuser to specify company
             company = Company.objects.get(pk=self.request.data.get('company_id'))
             serializer.save(company=company)
        else:
            # Handle cases where company cannot be determined (e.g., raise validation error)
            # This depends on specific requirements for users without companies
            # For now, assume authenticated users must have a company unless superuser
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("User must belong to a company to create this object.")

    def get_serializer_context(self):
        """Add request to the serializer context."""
        context = super().get_serializer_context()
        context.update({"request": self.request})
        return context

# --- ViewSets ---

class CompanyViewSet(viewsets.ModelViewSet):
    """API endpoint for Companies. Accessible only by superusers for management."""
    queryset = Company.objects.all()
    serializer_class = CompanySerializer
    # Typically, only superusers should manage companies
    permission_classes = [permissions.IsAdminUser]

    @action(detail=False, methods=['get', 'patch'], permission_classes=[permissions.IsAuthenticated])
    def me(self, request):
        """
        GET /api/companies/me/ - Returns the authenticated user's company
        PATCH /api/companies/me/ - Updates the authenticated user's company
        """
        user = request.user

        if not hasattr(user, 'company') or not user.company:
            return Response(
                {"detail": "User does not belong to a company."},
                status=status.HTTP_404_NOT_FOUND
            )

        if request.method == 'GET':
            serializer = self.get_serializer(user.company)
            return Response(serializer.data)

        elif request.method == 'PATCH':
            serializer = self.get_serializer(
                user.company,
                data=request.data,
                partial=True
            )
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data) 

class CustomUserViewSet(viewsets.ModelViewSet):
    """API endpoint for Users. Allows creation and management within a company context."""
    queryset = CustomUser.objects.all()
    serializer_class = CustomUserSerializer
    permission_classes = [permissions.IsAuthenticated] # Start with authenticated, refine later if needed

    def get_queryset(self):
        """Users can see other users in their own company. Superusers see all."""
        user = self.request.user
        if user.is_superuser:
            return CustomUser.objects.all()
        if hasattr(user, 'company') and user.company:
            # Users can see/manage others in the same company
            return CustomUser.objects.filter(company=user.company)
        # Users without a company or not superuser might only see themselves
        # return CustomUser.objects.filter(pk=user.pk) # Or return none()
        return CustomUser.objects.none()

    def perform_create(self, serializer):
        """Assign company based on creating user, allow superuser override."""
        user = self.request.user
        company_id = self.request.data.get('company_id')
        
        target_company = None
        if user.is_superuser and company_id:
            try:
                target_company = Company.objects.get(pk=company_id)
            except Company.DoesNotExist:
                 raise serializers.ValidationError({"company_id": "Invalid company specified."}) 
        elif hasattr(user, 'company') and user.company:
            target_company = user.company
        
        # We must set the company before validating the serializer if it relies on it
        # Or pass it in context. Here, we save with the determined company.
        if target_company:
             serializer.save(company=target_company)
        else:
             # Non-superuser without a company cannot create users for other companies
             # Or maybe allow creating users without company? Depends on rules.
             # For now, assume users must belong to a company if created by non-superuser
             if not user.is_superuser:
                 from rest_framework.exceptions import PermissionDenied
                 raise PermissionDenied("You must belong to a company to create users.")
             else:
                 # Superuser creating user without company
                 serializer.save(company=None)

class ClienteViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """API endpoint for Clientes, scoped by company."""
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    pagination_class = DynamicPageSizePagination
    # CompanyScopedViewSetMixin handles permissions and queryset filtering

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related('formas_cobranca', 'comissoes__funcionario')

        # Search filter
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) |
                Q(cpf__icontains=search) |
                Q(email__icontains=search) |
                Q(telefone__icontains=search)
            )

        return queryset

    @action(detail=False, methods=['post'], url_path='gerar-comissoes')
    def gerar_comissoes(self, request):
        """
        Gera despesas de comissão para o mês/ano especificado.

        POST /api/clientes/gerar-comissoes/
        Body: {
            "mes": 1-12,
            "ano": 2024
        }

        Retorna:
        {
            "comissionados": [
                {"id": 1, "nome": "João", "valor": 1000.00},
                ...
            ],
            "total": 5000.00,
            "mes": 1,
            "ano": 2024
        }
        """
        from datetime import date
        import calendar
        from django.db.models import Sum
        from core.models import Payment, Despesa, Funcionario

        # Validar parâmetros
        mes = request.data.get('mes')
        ano = request.data.get('ano')

        if not mes or not ano:
            return Response(
                {'erro': 'Parâmetros "mes" e "ano" são obrigatórios'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            mes = int(mes)
            ano = int(ano)
            if not (1 <= mes <= 12):
                raise ValueError()
        except ValueError:
            return Response(
                {'erro': 'Mês deve ser um número entre 1 e 12'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Calcular data de vencimento (último dia do mês)
        ultimo_dia_mes = calendar.monthrange(ano, mes)[1]
        data_vencimento = date(ano, mes, ultimo_dia_mes)

        # Buscar todas as alocações de pagamentos do mês/ano para receitas
        from core.models import Allocation

        allocations = Allocation.objects.filter(
            company=request.user.company,
            receita__isnull=False,
            payment__data_pagamento__month=mes,
            payment__data_pagamento__year=ano
        ).prefetch_related(
            'receita__comissoes__funcionario',
            'receita__cliente__comissoes__funcionario'
        ).select_related('payment', 'receita__cliente')

        # Percentual padrão da company (usado somente se a regra não definir o seu)
        percentual_default = request.user.company.percentual_comissao or Decimal('20.00')

        # Agrupar por comissionado
        comissionados_dict = {}
        for allocation in allocations:
            # Regras efetivas: da receita se existirem, senão do cliente
            regras = list(allocation.receita.comissoes.all())
            if not regras:
                regras = list(allocation.receita.cliente.comissoes.all())

            for regra in regras:
                func = regra.funcionario
                percentual = regra.percentual / Decimal('100.00')
                valor_comissao_alloc = allocation.valor * percentual

                if func.id not in comissionados_dict:
                    comissionados_dict[func.id] = {
                        'comissionado': func,
                        'valor_comissao': Decimal('0.00')
                    }
                comissionados_dict[func.id]['valor_comissao'] += valor_comissao_alloc

        # Deletar despesas de comissão do mês que não aparecem mais no cálculo
        ids_comissionados_ativos = list(comissionados_dict.keys())
        Despesa.objects.filter(
            company=request.user.company,
            tipo='C',
            data_vencimento=data_vencimento,
        ).exclude(responsavel_id__in=ids_comissionados_ativos).delete()

        comissionados_resultado = []
        total_comissoes = Decimal('0.00')

        for data in comissionados_dict.values():
            comissionado = data['comissionado']
            valor_comissao = data['valor_comissao']

            if valor_comissao > 0:
                # Atualiza se existe, cria se não existe
                despesa, created = Despesa.objects.update_or_create(
                    company=request.user.company,
                    responsavel=comissionado,
                    tipo='C',
                    data_vencimento=data_vencimento,
                    defaults={
                        'nome': f'Comissão {mes}/{ano} - {comissionado.nome}',
                        'descricao': f'Comissão referente aos pagamentos de {mes}/{ano}',
                        'valor': valor_comissao,
                        'situacao': 'A'
                    }
                )

                comissionados_resultado.append({
                    'id': comissionado.id,
                    'nome': comissionado.nome,
                    'valor': float(valor_comissao)
                })
                total_comissoes += valor_comissao

        if not comissionados_resultado:
            return Response(
                {'mensagem': f'Nenhuma comissão gerada para {mes}/{ano}'},
                status=status.HTTP_200_OK
            )

        return Response({
            'comissionados': comissionados_resultado,
            'total': float(total_comissoes),
            'mes': mes,
            'ano': ano
        })

class FuncionarioViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """API endpoint for Funcionarios, scoped by company."""
    queryset = Funcionario.objects.all()
    serializer_class = FuncionarioSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        queryset = super().get_queryset().filter(tipo__in=['F', 'P'])

        # Search filter
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) |
                Q(cpf__icontains=search) |
                Q(email__icontains=search) |
                Q(telefone__icontains=search)
            )

        return queryset
    # CompanyScopedViewSetMixin handles permissions and queryset filtering

class FornecedorViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Funcionario.objects.all()
    serializer_class = FuncionarioSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        queryset = super().get_queryset().filter(tipo='O')

        # Search filter
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) |
                Q(cpf__icontains=search) |
                Q(email__icontains=search) |
                Q(telefone__icontains=search)
            )

        return queryset

class FavorecidoViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Funcionario.objects.all()
    serializer_class = FuncionarioSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        return super().get_queryset().filter(tipo__in=['F', 'P', 'O'])


from django.db.models import Q
from django.utils.timezone import now

class ReceitaViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Receita.objects.all()
    serializer_class = ReceitaSerializer
    pagination_class = DynamicPageSizePagination

    def get_serializer_class(self):
        situacoes = self.request.query_params.getlist("situacao")

        # 🔹 Receitas em aberto → serializer com saldo
        if situacoes and set(situacoes).issubset({"A", "V"}):
            return ReceitaAbertaSerializer

        return ReceitaSerializer

    def _atualizar_vencidas(self):
        """Atualiza automaticamente receitas vencidas (on-the-fly)."""
        hoje = timezone.now().date()
        Receita.objects.filter(
            company=self.request.user.company,
            situacao='A',
            data_vencimento__lt=hoje
        ).update(situacao='V')

    def get_queryset(self):
        # 🔄 Atualiza vencidas antes de retornar o queryset
        self._atualizar_vencidas()

        queryset = super().get_queryset().select_related(
            "cliente", "company"
        ).prefetch_related(
            "allocations"
        )

        params = self.request.query_params

        # 🔎 FILTRO GLOBAL
        search = params.get("search")
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) |
                Q(descricao__icontains=search) |
                Q(cliente__nome__icontains=search) |
                Q(valor__icontains=search) |
                Q(data_vencimento__icontains=search)
            )

        # 🔸 filtros
        situacoes = params.getlist("situacao")
        if situacoes:
            queryset = queryset.filter(situacao__in=situacoes)

        cliente_id = params.get("cliente_id")
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)

        start_date = params.get("start_date")
        end_date = params.get("end_date")

        if start_date:
            queryset = queryset.filter(data_vencimento__gte=start_date)
        if end_date:
            queryset = queryset.filter(data_vencimento__lte=end_date)

        # 🔥 ORDENAÇÃO (adiciona id para garantir ordenação determinística)
        if situacoes and set(situacoes).issubset({"P", "V"}):
            queryset = queryset.order_by("-data_pagamento", "-data_vencimento", "id")
        else:
            queryset = queryset.order_by("data_vencimento", "id")

        return queryset

    def perform_create(self, serializer):
        receita = serializer.save(company=self.request.user.company)

        # 💰 Handle payment creation if marked as paid
        marcar_como_pago = self.request.data.get('marcar_como_pago', False)
        if marcar_como_pago:
            data_pagamento = self.request.data.get('data_pagamento')
            conta_bancaria_id = self.request.data.get('conta_bancaria_id')
            observacao_pagamento = self.request.data.get('observacao_pagamento', '')

            if data_pagamento and conta_bancaria_id:
                from core.models import Payment, ContaBancaria, Allocation

                try:
                    conta_bancaria = ContaBancaria.objects.get(
                        id=conta_bancaria_id,
                        company=self.request.user.company
                    )

                    # Cria o payment neutro (entrada)
                    payment = Payment.objects.create(
                        company=self.request.user.company,
                        tipo='E',  # Entrada
                        conta_bancaria=conta_bancaria,
                        valor=receita.valor,
                        data_pagamento=data_pagamento,
                        observacao=observacao_pagamento
                    )

                    # Cria a alocação para a receita
                    Allocation.objects.create(
                        company=self.request.user.company,
                        payment=payment,
                        receita=receita,
                        valor=receita.valor
                    )

                    # Atualiza saldo da conta bancária (entrada de dinheiro)
                    conta_bancaria.saldo_atual += payment.valor
                    conta_bancaria.save()

                    # Atualiza status da receita
                    receita.atualizar_status()
                except ContaBancaria.DoesNotExist:
                    logger.warning(
                        f"Conta bancária {conta_bancaria_id} não encontrada ao processar receita {receita.id}. "
                        "Pagamento não criado."
                    )


class ReceitaRecorrenteViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para gerenciar receitas recorrentes.

    Endpoints:
    - GET/POST /api/receitas-recorrentes/
    - GET/PUT/PATCH/DELETE /api/receitas-recorrentes/{id}/
    - POST /api/receitas-recorrentes/gerar-mes/
    """

    queryset = ReceitaRecorrente.objects.all()
    serializer_class = ReceitaRecorrenteSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'cliente', 'company'
        )

        params = self.request.query_params

        # Busca
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) |
                Q(descricao__icontains=search) |
                Q(cliente__nome__icontains=search)
            )

        # Filtros
        status_filter = params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        cliente_id = params.get('cliente_id')
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)

        tipo = params.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo=tipo)

        return queryset.order_by('nome', 'id')

    @action(detail=False, methods=['post'], url_path='gerar-mes')
    def gerar_mes(self, request):
        """
        Gera receitas individuais para o mês atual baseado nas recorrentes ativas.

        POST /api/receitas-recorrentes/gerar-mes/
        Body: {
            "mes": "2024-01" (opcional, default: mês atual)
        }

        Retorna:
        {
            "criadas": 5,
            "ignoradas": 2,
            "detalhes": [...]
        }
        """
        from datetime import date
        import calendar
        from django.db.models import Q

        # Pega mês da requisição ou usa mês atual
        mes_str = request.data.get('mes')
        if mes_str:
            try:
                ano, mes = map(int, mes_str.split('-'))
                mes_referencia = date(ano, mes, 1)
            except (ValueError, IndexError) as e:
                return Response(
                    {'erro': f'Formato de mês inválido. Use YYYY-MM. Detalhe: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            hoje = timezone.now().date()
            mes_referencia = date(hoje.year, hoje.month, 1)

        # Busca todas as receitas recorrentes ativas
        # Quando o mês é especificado manualmente, não filtra por data_inicio nem data_fim
        recorrentes = ReceitaRecorrente.objects.filter(
            company=request.user.company,
            status='A'
        )

        criadas = 0
        ignoradas = 0
        detalhes = []

        for recorrente in recorrentes:
            # Verifica se já existe receita para este mês
            nome_esperado = f"{recorrente.nome} - {mes_referencia.strftime('%m/%Y')}"

            receita_existente = Receita.objects.filter(
                company=request.user.company,
                nome=nome_esperado,
                cliente=recorrente.cliente
            ).exists()

            if receita_existente:
                ignoradas += 1
                detalhes.append({
                    'nome': recorrente.nome,
                    'status': 'ignorada',
                    'motivo': 'Já gerada para este mês'
                })
                continue

            # Calcula data de vencimento
            ultimo_dia_mes = calendar.monthrange(
                mes_referencia.year,
                mes_referencia.month
            )[1]
            dia_vencimento = min(recorrente.dia_vencimento, ultimo_dia_mes)
            data_vencimento = date(
                mes_referencia.year,
                mes_referencia.month,
                dia_vencimento
            )

            # Cria receita individual
            try:
                receita = Receita.objects.create(
                    company=request.user.company,
                    cliente=recorrente.cliente,
                    nome=f"{recorrente.nome} - {mes_referencia.strftime('%m/%Y')}",
                    descricao=recorrente.descricao or '',
                    valor=recorrente.valor,
                    tipo=recorrente.tipo,
                    data_vencimento=data_vencimento,
                    situacao='A',
                    forma_pagamento=recorrente.forma_pagamento
                )

                # Copia regras de comissão da recorrente para a receita gerada
                from core.models import ReceitaComissao
                for regra in recorrente.comissoes.all():
                    ReceitaComissao.objects.create(
                        receita=receita,
                        funcionario=regra.funcionario,
                        percentual=regra.percentual
                    )

                criadas += 1
                detalhes.append({
                    'nome': recorrente.nome,
                    'status': 'criada',
                    'data_vencimento': str(data_vencimento)
                })

            except Exception as e:
                detalhes.append({
                    'nome': recorrente.nome,
                    'status': 'erro',
                    'motivo': str(e)
                })

        return Response({
            'criadas': criadas,
            'ignoradas': ignoradas,
            'mes': mes_referencia.strftime('%Y-%m'),
            'detalhes': detalhes
        })

    @action(detail=True, methods=['post'], url_path='gerar-proximos-meses')
    def gerar_proximos_meses(self, request, pk=None):
        """
        Gera receitas para os próximos X meses de uma receita recorrente específica.

        POST /api/receitas-recorrentes/{id}/gerar-proximos-meses/
        Body: {
            "quantidade_meses": 10
        }

        Retorna:
        {
            "criadas": 10,
            "ignoradas": 0,
            "detalhes": [...]
        }
        """
        from datetime import date
        import calendar

        recorrente = self.get_object()
        quantidade_meses = request.data.get('quantidade_meses', 1)

        if not isinstance(quantidade_meses, int) or quantidade_meses < 1 or quantidade_meses > 24:
            return Response(
                {'erro': 'Quantidade de meses deve ser um número entre 1 e 24'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Helper function to add months
        def add_months(source_date, months):
            month = source_date.month - 1 + months
            year = source_date.year + month // 12
            month = month % 12 + 1
            day = min(source_date.day, calendar.monthrange(year, month)[1])
            return date(year, month, day)

        # Começar sempre do mês atual
        hoje = timezone.now().date()
        mes_inicial = date(hoje.year, hoje.month, 1)

        criadas = 0
        ignoradas = 0
        detalhes = []

        for i in range(quantidade_meses):
            mes_referencia = add_months(mes_inicial, i)

            # Verifica se está dentro do período de validade (compara apenas ano/mês)
            if (recorrente.data_inicio.year > mes_referencia.year or
                (recorrente.data_inicio.year == mes_referencia.year and
                 recorrente.data_inicio.month > mes_referencia.month)):
                ignoradas += 1
                detalhes.append({
                    'mes': mes_referencia.strftime('%Y-%m'),
                    'status': 'ignorada',
                    'motivo': 'Antes da data de início'
                })
                continue

            if recorrente.data_fim:
                if (recorrente.data_fim.year < mes_referencia.year or
                    (recorrente.data_fim.year == mes_referencia.year and
                     recorrente.data_fim.month < mes_referencia.month)):
                    ignoradas += 1
                    detalhes.append({
                        'mes': mes_referencia.strftime('%Y-%m'),
                        'status': 'ignorada',
                        'motivo': 'Depois da data de fim'
                    })
                    continue

            # Verifica se já existe
            nome_esperado = f"{recorrente.nome} - {mes_referencia.strftime('%m/%Y')}"
            receita_existente = Receita.objects.filter(
                company=request.user.company,
                nome=nome_esperado,
                cliente=recorrente.cliente
            ).exists()

            if receita_existente:
                ignoradas += 1
                detalhes.append({
                    'mes': mes_referencia.strftime('%Y-%m'),
                    'status': 'ignorada',
                    'motivo': 'Já existe'
                })
                continue

            # Calcula data de vencimento
            ultimo_dia_mes = calendar.monthrange(
                mes_referencia.year,
                mes_referencia.month
            )[1]
            dia_vencimento = min(recorrente.dia_vencimento, ultimo_dia_mes)
            data_vencimento = date(
                mes_referencia.year,
                mes_referencia.month,
                dia_vencimento
            )

            # Cria receita individual
            try:
                receita = Receita.objects.create(
                    company=request.user.company,
                    cliente=recorrente.cliente,
                    nome=nome_esperado,
                    descricao=recorrente.descricao or '',
                    valor=recorrente.valor,
                    tipo=recorrente.tipo,
                    data_vencimento=data_vencimento,
                    situacao='A',
                    forma_pagamento=recorrente.forma_pagamento
                )

                criadas += 1
                detalhes.append({
                    'mes': mes_referencia.strftime('%Y-%m'),
                    'status': 'criada',
                    'data_vencimento': str(data_vencimento)
                })

            except Exception as e:
                detalhes.append({
                    'mes': mes_referencia.strftime('%Y-%m'),
                    'status': 'erro',
                    'motivo': str(e)
                })

        return Response({
            'criadas': criadas,
            'ignoradas': ignoradas,
            'detalhes': detalhes
        })

    def perform_update(self, serializer):
        instance = serializer.save()
        instance.atualizar_status()


class DespesaViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    queryset = Despesa.objects.all()
    serializer_class = DespesaSerializer
    pagination_class = DynamicPageSizePagination

    def get_serializer_class(self):
            situacoes = self.request.query_params.getlist("situacao")

            # 🔹 Despesas em aberto → serializer com saldo
            if situacoes and set(situacoes).issubset({"A", "V"}):
                return DespesaAbertaSerializer

            return DespesaSerializer

    def _atualizar_vencidas(self):
        """Atualiza automaticamente despesas vencidas (on-the-fly)."""
        hoje = timezone.now().date()
        Despesa.objects.filter(
            company=self.request.user.company,
            situacao='A',
            data_vencimento__lt=hoje
        ).update(situacao='V')

    def get_queryset(self):
        # 🔄 Atualiza vencidas antes de retornar o queryset
        self._atualizar_vencidas()

        queryset = super().get_queryset().select_related(
            "responsavel", "company"
        ).prefetch_related(
            "allocations"
        )

        params = self.request.query_params

        # 🔎 BUSCA
        search = params.get("search")
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) |
                Q(descricao__icontains=search) |
                Q(responsavel__nome__icontains=search) |
                Q(valor__icontains=search) |
                Q(data_vencimento__icontains=search)
            )

        # 🔸 filtros
        situacoes = params.getlist("situacao")
        if situacoes:
            queryset = queryset.filter(situacao__in=situacoes)

        responsavel_id = params.get("responsavel_id")
        if responsavel_id:
            queryset = queryset.filter(responsavel_id=responsavel_id)

        start_date = params.get("start_date")
        end_date = params.get("end_date")

        if start_date:
            queryset = queryset.filter(data_vencimento__gte=start_date)
        if end_date:
            queryset = queryset.filter(data_vencimento__lte=end_date)

        # 🔥 ORDENAÇÃO (adiciona id para garantir ordenação determinística)
        if situacoes and set(situacoes).issubset({"P", "V"}):
            queryset = queryset.order_by("-data_pagamento", "-data_vencimento", "id")
        else:
            queryset = queryset.order_by("data_vencimento", "id")

        return queryset

    def perform_create(self, serializer):
        despesa = serializer.save(company=self.request.user.company)

        # 💰 Handle payment creation if marked as paid
        marcar_como_pago = self.request.data.get('marcar_como_pago', False)
        if marcar_como_pago:
            data_pagamento = self.request.data.get('data_pagamento')
            conta_bancaria_id = self.request.data.get('conta_bancaria_id')
            observacao_pagamento = self.request.data.get('observacao_pagamento', '')

            if data_pagamento and conta_bancaria_id:
                from core.models import Payment, ContaBancaria, Allocation

                try:
                    conta_bancaria = ContaBancaria.objects.get(
                        id=conta_bancaria_id,
                        company=self.request.user.company
                    )

                    # Cria o payment neutro (saída)
                    payment = Payment.objects.create(
                        company=self.request.user.company,
                        tipo='S',  # Saída
                        conta_bancaria=conta_bancaria,
                        valor=despesa.valor,
                        data_pagamento=data_pagamento,
                        observacao=observacao_pagamento
                    )

                    # Cria a alocação para a despesa
                    Allocation.objects.create(
                        company=self.request.user.company,
                        payment=payment,
                        despesa=despesa,
                        valor=despesa.valor
                    )

                    # Atualiza saldo da conta bancária (saída de dinheiro)
                    conta_bancaria.saldo_atual -= payment.valor
                    conta_bancaria.save()

                    # Atualiza status da despesa
                    despesa.atualizar_status()
                except ContaBancaria.DoesNotExist:
                    logger.warning(
                        f"Conta bancária {conta_bancaria_id} não encontrada ao processar despesa {despesa.id}. "
                        "Pagamento não criado."
                    )

    def perform_update(self, serializer):
        instance = serializer.save()
        instance.atualizar_status()


class DespesaRecorrenteViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet para gerenciar despesas recorrentes.

    Endpoints:
    - GET/POST /api/despesas-recorrentes/
    - GET/PUT/PATCH/DELETE /api/despesas-recorrentes/{id}/
    - POST /api/despesas-recorrentes/gerar-mes/
    """

    queryset = DespesaRecorrente.objects.all()
    serializer_class = DespesaRecorrenteSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'responsavel', 'company'
        )

        params = self.request.query_params

        # Busca
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) |
                Q(descricao__icontains=search) |
                Q(responsavel__nome__icontains=search)
            )

        # Filtros
        status_filter = params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        responsavel_id = params.get('responsavel_id')
        if responsavel_id:
            queryset = queryset.filter(responsavel_id=responsavel_id)

        tipo = params.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo=tipo)

        return queryset.order_by('nome', 'id')

    @action(detail=False, methods=['post'], url_path='gerar-mes')
    def gerar_mes(self, request):
        """
        Gera despesas individuais para o mês atual baseado nas recorrentes ativas.

        POST /api/despesas-recorrentes/gerar-mes/
        Body: {
            "mes": "2024-01" (opcional, default: mês atual)
        }

        Retorna:
        {
            "criadas": 5,
            "ignoradas": 2,
            "detalhes": [...]
        }
        """
        from datetime import date
        import calendar
        from django.db.models import Q

        # Pega mês da requisição ou usa mês atual
        mes_str = request.data.get('mes')
        if mes_str:
            try:
                ano, mes = map(int, mes_str.split('-'))
                mes_referencia = date(ano, mes, 1)
            except (ValueError, IndexError) as e:
                return Response(
                    {'erro': f'Formato de mês inválido. Use YYYY-MM. Detalhe: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            hoje = timezone.now().date()
            mes_referencia = date(hoje.year, hoje.month, 1)

        # Busca todas as despesas recorrentes ativas
        # Quando o mês é especificado manualmente, não filtra por data_inicio nem data_fim
        recorrentes = DespesaRecorrente.objects.filter(
            company=request.user.company,
            status='A'
        )

        criadas = 0
        ignoradas = 0
        detalhes = []
        total_recorrentes = recorrentes.count()

        for recorrente in recorrentes:
            # Verifica se já existe despesa para este mês
            nome_esperado = f"{recorrente.nome} - {mes_referencia.strftime('%m/%Y')}"

            despesa_existente = Despesa.objects.filter(
                company=request.user.company,
                nome=nome_esperado,
                responsavel=recorrente.responsavel
            ).exists()

            if despesa_existente:
                ignoradas += 1
                detalhes.append({
                    'nome': recorrente.nome,
                    'status': 'ignorada',
                    'motivo': 'Já gerada para este mês'
                })
                continue

            # Calcula data de vencimento
            ultimo_dia_mes = calendar.monthrange(
                mes_referencia.year,
                mes_referencia.month
            )[1]
            dia_vencimento = min(recorrente.dia_vencimento, ultimo_dia_mes)
            data_vencimento = date(
                mes_referencia.year,
                mes_referencia.month,
                dia_vencimento
            )

            # Cria despesa individual
            try:
                Despesa.objects.create(
                    company=request.user.company,
                    responsavel=recorrente.responsavel,
                    nome=f"{recorrente.nome} - {mes_referencia.strftime('%m/%Y')}",
                    descricao=recorrente.descricao or '',
                    valor=recorrente.valor,
                    tipo=recorrente.tipo,
                    data_vencimento=data_vencimento,
                    situacao='A'
                )

                criadas += 1
                detalhes.append({
                    'nome': recorrente.nome,
                    'status': 'criada',
                    'data_vencimento': str(data_vencimento)
                })

            except Exception as e:
                detalhes.append({
                    'nome': recorrente.nome,
                    'status': 'erro',
                    'motivo': str(e)
                })

        return Response({
            'criadas': criadas,
            'ignoradas': ignoradas,
            'total_recorrentes': total_recorrentes,
            'mes': mes_referencia.strftime('%Y-%m'),
            'detalhes': detalhes
        })

    @action(detail=True, methods=['post'], url_path='gerar-proximos-meses')
    def gerar_proximos_meses(self, request, pk=None):
        """
        Gera despesas para os próximos X meses de uma despesa recorrente específica.

        POST /api/despesas-recorrentes/{id}/gerar-proximos-meses/
        Body: {
            "quantidade_meses": 10
        }

        Retorna:
        {
            "criadas": 10,
            "ignoradas": 0,
            "detalhes": [...]
        }
        """
        from datetime import date
        import calendar

        recorrente = self.get_object()
        quantidade_meses = request.data.get('quantidade_meses', 1)

        if not isinstance(quantidade_meses, int) or quantidade_meses < 1 or quantidade_meses > 24:
            return Response(
                {'erro': 'Quantidade de meses deve ser um número entre 1 e 24'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Helper function to add months
        def add_months(source_date, months):
            month = source_date.month - 1 + months
            year = source_date.year + month // 12
            month = month % 12 + 1
            day = min(source_date.day, calendar.monthrange(year, month)[1])
            return date(year, month, day)

        # Começar sempre do mês atual
        hoje = timezone.now().date()
        mes_inicial = date(hoje.year, hoje.month, 1)

        criadas = 0
        ignoradas = 0
        detalhes = []

        for i in range(quantidade_meses):
            mes_referencia = add_months(mes_inicial, i)

            # Verifica se está dentro do período de validade (compara apenas ano/mês)
            if (recorrente.data_inicio.year > mes_referencia.year or
                (recorrente.data_inicio.year == mes_referencia.year and
                 recorrente.data_inicio.month > mes_referencia.month)):
                ignoradas += 1
                detalhes.append({
                    'mes': mes_referencia.strftime('%Y-%m'),
                    'status': 'ignorada',
                    'motivo': 'Antes da data de início'
                })
                continue

            if recorrente.data_fim:
                if (recorrente.data_fim.year < mes_referencia.year or
                    (recorrente.data_fim.year == mes_referencia.year and
                     recorrente.data_fim.month < mes_referencia.month)):
                    ignoradas += 1
                    detalhes.append({
                        'mes': mes_referencia.strftime('%Y-%m'),
                        'status': 'ignorada',
                        'motivo': 'Depois da data de fim'
                    })
                    continue

            # Verifica se já existe
            nome_esperado = f"{recorrente.nome} - {mes_referencia.strftime('%m/%Y')}"
            despesa_existente = Despesa.objects.filter(
                company=request.user.company,
                nome=nome_esperado,
                responsavel=recorrente.responsavel
            ).exists()

            if despesa_existente:
                ignoradas += 1
                detalhes.append({
                    'mes': mes_referencia.strftime('%Y-%m'),
                    'status': 'ignorada',
                    'motivo': 'Já existe'
                })
                continue

            # Calcula data de vencimento
            ultimo_dia_mes = calendar.monthrange(
                mes_referencia.year,
                mes_referencia.month
            )[1]
            dia_vencimento = min(recorrente.dia_vencimento, ultimo_dia_mes)
            data_vencimento = date(
                mes_referencia.year,
                mes_referencia.month,
                dia_vencimento
            )

            # Cria despesa individual
            try:
                Despesa.objects.create(
                    company=request.user.company,
                    responsavel=recorrente.responsavel,
                    nome=nome_esperado,
                    descricao=recorrente.descricao or '',
                    valor=recorrente.valor,
                    tipo=recorrente.tipo,
                    data_vencimento=data_vencimento,
                    situacao='A'
                )

                criadas += 1
                detalhes.append({
                    'mes': mes_referencia.strftime('%Y-%m'),
                    'status': 'criada',
                    'data_vencimento': str(data_vencimento)
                })

            except Exception as e:
                detalhes.append({
                    'mes': mes_referencia.strftime('%Y-%m'),
                    'status': 'erro',
                    'motivo': str(e)
                })

        return Response({
            'criadas': criadas,
            'ignoradas': ignoradas,
            'detalhes': detalhes
        })


from django.db.models import Q
from rest_framework import viewsets

class PaymentViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """
    API endpoint para registrar pagamentos neutros (entrada/saída de caixa).
    As alocações para Receitas/Despesas/Passivos são feitas via Allocation.
    """
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'conta_bancaria'
        ).prefetch_related(
            Prefetch(
                'allocations',
                queryset=Allocation.objects.select_related(
                    'receita__cliente',
                    'despesa__responsavel',
                    'custodia__cliente',
                    'custodia__funcionario',
                    'transfer__from_bank',
                    'transfer__to_bank'
                )
            )
        )
        params = self.request.query_params

        # Filtros por data
        start_date = params.get('start_date')
        end_date = params.get('end_date')

        if start_date:
            queryset = queryset.filter(data_pagamento__gte=start_date)
        if end_date:
            queryset = queryset.filter(data_pagamento__lte=end_date)

        # Filtro por tipo (Entrada/Saída)
        tipo = params.get('tipo')  # 'E' | 'S'
        if tipo:
            queryset = queryset.filter(tipo=tipo)

        # Filtro por conta bancária
        conta_bancaria_id = params.get('conta_bancaria_id')
        if conta_bancaria_id:
            queryset = queryset.filter(conta_bancaria_id=conta_bancaria_id)

        # Filtro por situação da receita/despesa
        situacao = params.get('situacao')  # 'P' | 'A' | 'V'
        if situacao:
            # Filtra payments que têm allocations com receitas ou despesas na situação especificada
            queryset = queryset.filter(
                Q(allocations__receita__situacao=situacao) |
                Q(allocations__despesa__situacao=situacao)
            ).distinct()

        # Busca global
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(valor__icontains=search) |
                Q(observacao__icontains=search) |
                Q(data_pagamento__icontains=search) |
                # Filtro por nome do banco
                Q(conta_bancaria__nome__icontains=search) |
                # Filtro por entidades vinculadas
                Q(allocations__receita__nome__icontains=search) |
                Q(allocations__despesa__nome__icontains=search) |
                Q(allocations__custodia__nome__icontains=search) |
                Q(allocations__transfer__from_bank__nome__icontains=search) |
                Q(allocations__transfer__to_bank__nome__icontains=search)
            ).distinct()

        return queryset.order_by('-data_pagamento', '-id')

    def perform_create(self, serializer):
        from django.db.models import F
        from django.db import transaction

        with transaction.atomic():
            payment = serializer.save(company=self.request.user.company)

            # Atualiza saldo da conta bancária usando F() para operação atômica
            # select_for_update() garante que não haja race condition
            if payment.tipo == 'E':
                # Entrada de dinheiro (+)
                ContaBancaria.objects.select_for_update().filter(
                    pk=payment.conta_bancaria.pk
                ).update(saldo_atual=F('saldo_atual') + payment.valor)
            else:
                # Saída de dinheiro (-)
                ContaBancaria.objects.select_for_update().filter(
                    pk=payment.conta_bancaria.pk
                ).update(saldo_atual=F('saldo_atual') - payment.valor)

    def perform_update(self, serializer):
        from django.db.models import F
        from django.db import transaction

        with transaction.atomic():
            # Guarda informações antigas antes de atualizar
            old_payment = Payment.objects.select_for_update().get(pk=serializer.instance.pk)
            old_valor = old_payment.valor
            old_tipo = old_payment.tipo
            old_conta = old_payment.conta_bancaria

            payment = serializer.save()

            # Reverte a operação antiga usando F() para operação atômica
            if old_tipo == 'E':
                ContaBancaria.objects.select_for_update().filter(pk=old_conta.pk).update(
                    saldo_atual=F('saldo_atual') - old_valor
                )
            else:
                ContaBancaria.objects.select_for_update().filter(pk=old_conta.pk).update(
                    saldo_atual=F('saldo_atual') + old_valor
                )

            # Aplica a operação nova usando F() para operação atômica
            if payment.tipo == 'E':
                ContaBancaria.objects.select_for_update().filter(
                    pk=payment.conta_bancaria.pk
                ).update(saldo_atual=F('saldo_atual') + payment.valor)
            else:
                ContaBancaria.objects.select_for_update().filter(
                    pk=payment.conta_bancaria.pk
                ).update(saldo_atual=F('saldo_atual') - payment.valor)

            # Atualiza status de todas as contas alocadas
            for allocation in payment.allocations.all():
                if allocation.receita:
                    allocation.receita.atualizar_status()
                elif allocation.despesa:
                    allocation.despesa.atualizar_status()
                elif allocation.custodia:
                    allocation.custodia.atualizar_status()

    @action(detail=False, methods=['post'], url_path='import-extrato')
    def import_extrato(self, request):
        """
        Importa pagamentos a partir de um arquivo XLSX de extrato bancário do BTG.
        Espera:
        - file: arquivo XLSX
        - conta_bancaria_id: ID da conta bancária
        """
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Arquivo não fornecido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if 'conta_bancaria_id' not in request.data:
            return Response(
                {'error': 'conta_bancaria_id não fornecido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES['file']
        conta_bancaria_id = request.data['conta_bancaria_id']

        # Verifica se a conta bancária existe e pertence ao usuário
        try:
            conta_bancaria = ContaBancaria.objects.get(
                id=conta_bancaria_id,
                company=request.user.company
            )
        except ContaBancaria.DoesNotExist:
            return Response(
                {'error': 'Conta bancária não encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            from openpyxl import load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
            from django.db.models import F

            # Carrega o workbook
            try:
                wb = load_workbook(file, data_only=True)
            except InvalidFileException:
                return Response(
                    {'error': 'Arquivo inválido. Por favor, envie um arquivo XLSX válido.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            ws = wb.active

            # Procura a linha de cabeçalho
            # Formato BTG: "Data de lançamento | Descrição do lançamento | Entradas / Saídas (R$) | Saldo (R$)"
            header_row = None
            date_col = None
            value_col = None
            desc_col = None

            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
                if row and any(row):
                    # Procura por palavras-chave nas células
                    for col_idx, cell_value in enumerate(row):
                        if cell_value and isinstance(cell_value, str):
                            cell_lower = cell_value.lower().strip()
                            # Remove acentos para comparação
                            import unicodedata
                            cell_normalized = ''.join(
                                c for c in unicodedata.normalize('NFD', cell_lower)
                                if unicodedata.category(c) != 'Mn'
                            )

                            # Procura coluna de data (Data de lançamento, Data, etc.)
                            if date_col is None and 'data' in cell_normalized:
                                date_col = col_idx

                            # Procura coluna de valor (Entradas/Saídas, Valor, etc.)
                            # Ignora coluna "Saldo"
                            if value_col is None and 'saldo' not in cell_normalized:
                                if any(kw in cell_normalized for kw in ['entrada', 'saida', 'valor', 'movimentacao']):
                                    value_col = col_idx

                            # Procura coluna de descrição (mas não a coluna de data)
                            # Prioriza "Descrição do lançamento" (formato BTG) ou "Histórico"
                            if desc_col is None and 'data' not in cell_normalized:
                                # Verifica primeiro se é exatamente "descrição do lançamento" ou similar
                                if 'lancamento' in cell_normalized and 'descri' in cell_normalized:
                                    desc_col = col_idx
                                # Caso contrário, aceita qualquer coluna com descrição ou histórico
                                elif any(kw in cell_normalized for kw in ['descri', 'historico']):
                                    desc_col = col_idx

                    # Se encontrou data E valor, considera essa linha como cabeçalho
                    if date_col is not None and value_col is not None:
                        header_row = idx
                        break

            if header_row is None or date_col is None:
                return Response(
                    {'error': 'Não foi possível identificar a coluna de data no extrato.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if value_col is None:
                return Response(
                    {'error': 'Não foi possível identificar a coluna de valor (Entradas/Saídas) no extrato.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Processa as linhas de dados
            created_count = 0
            skipped_count = 0  # Pagamentos duplicados ignorados
            errors = []
            potential_duplicates = []  # Lista de duplicatas potenciais para confirmação do usuário
            payments_to_create = []  # Lista de pagamentos a serem criados (apenas na segunda passagem)
            total_entradas = Decimal('0.00')
            total_saidas = Decimal('0.00')

            # Verifica se há lista de linhas confirmadas para importar (segunda passagem)
            force_import_lines = request.data.get('force_import_lines', [])
            confirmed = request.data.get('confirmed', 'false').lower() == 'true'

            if isinstance(force_import_lines, str):
                import json
                try:
                    force_import_lines = json.loads(force_import_lines)
                except json.JSONDecodeError as e:
                    # Se não conseguir parsear o JSON, assume lista vazia
                    force_import_lines = []
                    print(f"Aviso: Erro ao parsear force_import_lines JSON: {e}")

            # Se confirmed=true, significa que o usuário já viu o diálogo e escolheu
            # Neste caso, apenas importar as linhas que estão em force_import_lines
            # e pular as duplicatas potenciais que não foram selecionadas

            # Recarrega a conta para garantir que temos o saldo mais recente
            conta_bancaria.refresh_from_db()
            saldo_inicial = conta_bancaria.saldo_atual

            for idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                if not row or not any(row):
                    continue

                try:
                    # Extrai data
                    date_value = row[date_col] if date_col < len(row) else None
                    if not date_value:
                        continue

                    # Converte data
                    if isinstance(date_value, datetime):
                        data_pagamento = date_value.date()
                    elif isinstance(date_value, date):
                        data_pagamento = date_value
                    else:
                        # Tenta parsear string
                        from datetime import datetime as dt
                        date_str = str(date_value).strip()

                        # Tenta diferentes formatos
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y']:
                            try:
                                data_pagamento = dt.strptime(date_str, fmt).date()
                                break
                            except ValueError:
                                continue
                        else:
                            errors.append(f'Linha {idx}: formato de data inválido: {date_value}')
                            continue

                    # Extrai valor
                    value_raw = row[value_col] if value_col < len(row) else None
                    if value_raw is None or value_raw == '' or value_raw == 0:
                        continue

                    # Converte valor
                    if isinstance(value_raw, (int, float, Decimal)):
                        # Converte para string primeiro para evitar problemas de precisão de float
                        valor_num = Decimal(str(value_raw))
                    else:
                        # Remove caracteres não numéricos exceto vírgula, ponto e sinal
                        valor_str = str(value_raw).strip()
                        valor_str = valor_str.replace('R$', '').replace(' ', '').replace('\xa0', '')

                        # Formato brasileiro: 1.234,56 ou 1234,56
                        # Remove pontos (separador de milhar) e substitui vírgula por ponto
                        if ',' in valor_str:
                            valor_str = valor_str.replace('.', '').replace(',', '.')
                        # Se não tem vírgula mas tem ponto, assume formato americano ou já está correto

                        try:
                            valor_num = Decimal(valor_str)
                        except (ValueError, decimal.InvalidOperation) as e:
                            errors.append(f'Linha {idx}: formato de valor inválido: {value_raw} (erro: {str(e)})')
                            continue

                    # Quantiza para 2 casas decimais para garantir precisão
                    valor_num = valor_num.quantize(Decimal('0.01'))

                    # Determina tipo (Entrada ou Saída) baseado no sinal
                    if valor_num > 0:
                        tipo = 'E'  # Entrada
                        valor = valor_num  # Não precisa de abs() pois já é positivo
                    else:
                        tipo = 'S'  # Saída
                        valor = abs(valor_num)

                    # Extrai descrição/observação
                    observacao = ''
                    if desc_col is not None and desc_col < len(row):
                        desc_value = row[desc_col]
                        if desc_value:
                            observacao = str(desc_value).strip()

                    # ============================================
                    # VERIFICAÇÃO DE DUPLICATAS
                    # ============================================
                    # Busca pagamentos existentes com mesma data e valor (em QUALQUER banco)
                    existing_payments = Payment.objects.filter(
                        company=request.user.company,
                        data_pagamento=data_pagamento,
                        valor=valor,
                        tipo=tipo
                    )

                    # Verifica se existe duplicata exata (incluindo observação)
                    duplicata_exata = existing_payments.filter(observacao=observacao).exists()

                    if duplicata_exata:
                        # Duplicata exata encontrada - pular este pagamento
                        skipped_count += 1
                        continue

                    # Verifica se existe duplicata potencial (data + valor, mas observação diferente)
                    duplicata_potencial = existing_payments.exclude(observacao=observacao).first()

                    if duplicata_potencial:
                        # Se NÃO confirmado, apenas adiciona à lista (não importa nada ainda)
                        if not confirmed:
                            potential_duplicates.append({
                                'line_index': idx,
                                'new_payment': {
                                    'data': data_pagamento.strftime('%Y-%m-%d'),
                                    'valor': str(valor),
                                    'tipo': tipo,
                                    'observacao': observacao or '',
                                    'banco': conta_bancaria.nome
                                },
                                'existing_payment': {
                                    'id': duplicata_potencial.id,
                                    'data': duplicata_potencial.data_pagamento.strftime('%Y-%m-%d'),
                                    'valor': str(duplicata_potencial.valor),
                                    'tipo': duplicata_potencial.tipo,
                                    'observacao': duplicata_potencial.observacao or '',
                                    'banco': duplicata_potencial.conta_bancaria.nome
                                }
                            })
                            continue  # Pula e não adiciona à lista de criação

                        # Se confirmed=true, verifica se usuário escolheu importar esta linha
                        if idx not in force_import_lines:
                            # Usuário escolheu não importar esta duplicata
                            skipped_count += 1
                            continue

                    # Adiciona à lista de pagamentos a serem criados
                    payments_to_create.append({
                        'company': request.user.company,
                        'conta_bancaria': conta_bancaria,
                        'tipo': tipo,
                        'valor': valor,
                        'data_pagamento': data_pagamento,
                        'observacao': observacao
                    })

                except Exception as e:
                    errors.append(f'Linha {idx}: {str(e)}')

            # =====================================================
            # VERIFICAÇÃO FINAL: Se houver duplicatas potenciais na primeira passagem,
            # retorna SEM criar nenhum pagamento
            # =====================================================
            if potential_duplicates and not confirmed:
                response_data = {
                    'success': False,
                    'requires_confirmation': True,
                    'potential_duplicates': potential_duplicates,
                    'message': f'Encontradas {len(potential_duplicates)} possível(is) duplicata(s) que requerem confirmação.'
                }
                return Response(response_data, status=status.HTTP_200_OK)

            # =====================================================
            # CRIAÇÃO DOS PAGAMENTOS
            # Se chegou aqui, pode criar os pagamentos
            # =====================================================
            for payment_data in payments_to_create:
                try:
                    payment = Payment.objects.create(**payment_data)

                    # Atualiza saldo da conta usando F() para operação atômica
                    if payment_data['tipo'] == 'E':
                        ContaBancaria.objects.filter(pk=conta_bancaria.pk).update(
                            saldo_atual=F('saldo_atual') + payment_data['valor']
                        )
                        total_entradas += payment_data['valor']
                    else:
                        ContaBancaria.objects.filter(pk=conta_bancaria.pk).update(
                            saldo_atual=F('saldo_atual') - payment_data['valor']
                        )
                        total_saidas += payment_data['valor']

                    created_count += 1
                except Exception as e:
                    errors.append(f'Erro ao criar pagamento: {str(e)}')

            # Recarrega a conta para obter o saldo final atualizado
            conta_bancaria.refresh_from_db()
            saldo_final = conta_bancaria.saldo_atual

            # Calcula a diferença esperada vs real
            saldo_esperado = saldo_inicial + total_entradas - total_saidas
            diferenca = saldo_final - saldo_esperado

            # Se houver duplicatas potenciais (não deveria chegar aqui na primeira passagem)
            if potential_duplicates:
                response_data = {
                    'success': False,
                    'requires_confirmation': True,
                    'potential_duplicates': potential_duplicates,
                    'message': f'Encontradas {len(potential_duplicates)} possível(is) duplicata(s) que requerem confirmação.'
                }
                return Response(response_data, status=status.HTTP_200_OK)

            # Caso contrário, retorna sucesso normal
            response_data = {
                'success': True,
                'created_count': created_count,
                'skipped_count': skipped_count,  # Quantidade de pagamentos ignorados (duplicatas exatas)
                'conta_bancaria': conta_bancaria.nome,
                'saldo_inicial': str(saldo_inicial),
                'saldo_final': str(saldo_final),
                'total_entradas': str(total_entradas),
                'total_saidas': str(total_saidas),
                'saldo_esperado': str(saldo_esperado),
                'diferenca': str(diferenca)
            }

            if errors:
                response_data['errors'] = errors[:10]  # Limita a 10 erros
                response_data['total_errors'] = len(errors)

            return Response(response_data, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'Erro ao processar arquivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='conciliar-bancario')
    def conciliar_bancario(self, request):
        """
        Concilia pagamentos sem alocação com receitas/despesas/custódias em aberto
        do mês especificado, fazendo match por valor e nome da contraparte na observação.

        Espera:
        - mes: int (1-12)
        - ano: int (ex: 2026)
        """
        import unicodedata
        import re

        def normalizar_string(texto):
            """Remove acentos e converte para lowercase para comparação"""
            if not texto:
                return ''
            # Normaliza unicode e remove acentos
            texto_nfd = unicodedata.normalize('NFD', str(texto))
            texto_sem_acentos = ''.join(
                char for char in texto_nfd
                if unicodedata.category(char) != 'Mn'
            )
            return texto_sem_acentos.lower().strip()

        def extrair_palavras_significativas(texto):
            """
            Extrai palavras significativas de um texto, removendo:
            - Preposições comuns (de, da, do, dos, das, para, pra, em, no, na, etc.)
            - Palavras muito curtas (< 3 caracteres)
            - Palavras bancárias comuns (pix, ted, transferencia, recebido, enviado, etc.)
            """
            if not texto:
                return set()

            # Normaliza o texto
            texto_norm = normalizar_string(texto)

            # Remove pontuação e divide em palavras
            palavras = re.findall(r'\b\w+\b', texto_norm)

            # Stop words - palavras a ignorar
            stop_words = {
                'de', 'da', 'do', 'dos', 'das', 'para', 'pra', 'em', 'no', 'na', 'nos', 'nas',
                'com', 'por', 'a', 'o', 'e', 'ou', 'um', 'uma', 'ao', 'aos', 'as',
                'pix', 'ted', 'transferencia', 'recebido', 'enviado', 'pagamento', 'recebimento',
                'valor', 'ref', 'referente', 'ltda', 'me', 'sa', 'eireli'
            }

            # Filtra palavras significativas
            palavras_significativas = {
                palavra for palavra in palavras
                if len(palavra) >= 3 and palavra not in stop_words
            }

            return palavras_significativas

        def match_por_palavras_comuns(observacao, *textos_para_comparar):
            """
            Verifica se há pelo menos 2 palavras significativas em comum entre
            a observação e qualquer um dos textos fornecidos (nome cliente, nome despesa, etc.)

            Retorna True se encontrar 2+ palavras em comum com qualquer texto.
            """
            if not observacao:
                return False

            palavras_obs = extrair_palavras_significativas(observacao)

            if len(palavras_obs) < 2:
                return False

            for texto in textos_para_comparar:
                if not texto:
                    continue

                palavras_texto = extrair_palavras_significativas(texto)
                palavras_em_comum = palavras_obs.intersection(palavras_texto)

                # Match válido se há 2 ou mais palavras significativas em comum
                if len(palavras_em_comum) >= 2:
                    return True

            return False

        def nome_em_observacao(observacao, *nomes):
            """
            Verifica se algum dos nomes está contido na observação.
            Agora usa duas estratégias:
            1. Match exato (nome completo na observação) - mais confiável
            2. Match por palavras comuns (2+ palavras) - mais flexível
            """
            if not observacao:
                return False

            # Estratégia 1: Match exato (nome completo aparece na observação)
            obs_norm = normalizar_string(observacao)
            for nome in nomes:
                if nome:
                    nome_norm = normalizar_string(nome)
                    if nome_norm and nome_norm in obs_norm:
                        return True

            # Estratégia 2: Match por palavras comuns (pelo menos 2 palavras)
            if match_por_palavras_comuns(observacao, *nomes):
                return True

            return False

        mes = request.data.get('mes')
        ano = request.data.get('ano')

        if not mes or not ano:
            return Response(
                {'error': 'Parâmetros mes e ano são obrigatórios'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            mes = int(mes)
            ano = int(ano)

            if mes < 1 or mes > 12:
                return Response(
                    {'error': 'Mês deve estar entre 1 e 12'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except ValueError:
            return Response(
                {'error': 'Mes e ano devem ser números inteiros'},
                status=status.HTTP_400_BAD_REQUEST
            )

        company = request.user.company

        # Busca payments sem alocação no mês especificado
        payments_sem_alocacao = Payment.objects.filter(
            company=company,
            data_pagamento__year=ano,
            data_pagamento__month=mes
        ).annotate(
            num_allocations=Count('allocations')
        ).filter(
            num_allocations=0
        ).order_by('data_pagamento')

        # Busca receitas em aberto ou vencidas no mês (não pagas)
        # Pre-calcula total alocado para evitar N+1 queries
        receitas_abertas = Receita.objects.filter(
            company=company,
            data_vencimento__year=ano,
            data_vencimento__month=mes,
            situacao__in=['A', 'V']  # Em Aberto ou Vencida
        ).annotate(
            total_alocado=Coalesce(Sum('allocations__valor'), Decimal('0.00'))
        ).order_by('data_vencimento')

        # Debug: total de receitas no mês (qualquer situação)
        total_receitas_mes = Receita.objects.filter(
            company=company,
            data_vencimento__year=ano,
            data_vencimento__month=mes
        ).count()

        # Busca despesas em aberto ou vencidas no mês (não pagas)
        # Pre-calcula total alocado para evitar N+1 queries
        despesas_abertas = Despesa.objects.filter(
            company=company,
            data_vencimento__year=ano,
            data_vencimento__month=mes,
            situacao__in=['A', 'V']  # Em Aberto ou Vencida
        ).annotate(
            total_alocado=Coalesce(Sum('allocations__valor'), Decimal('0.00'))
        ).order_by('data_vencimento')

        # Debug: total de despesas no mês (qualquer situação)
        total_despesas_mes = Despesa.objects.filter(
            company=company,
            data_vencimento__year=ano,
            data_vencimento__month=mes
        ).count()

        # Busca custódias em aberto
        # Pre-calcula totais de entradas e saídas para evitar N+1 queries
        custodias_abertas = Custodia.objects.filter(
            company=company,
            status='A'
        ).annotate(
            valor_restante=F('valor_total') - F('valor_liquidado'),
            total_entradas=Coalesce(
                Sum('allocations__valor', filter=Q(allocations__payment__tipo='E')),
                Decimal('0.00')
            ),
            total_saidas=Coalesce(
                Sum('allocations__valor', filter=Q(allocations__payment__tipo='S')),
                Decimal('0.00')
            )
        ).filter(
            valor_restante__gt=0
        )

        # Estatísticas
        matches_receitas = 0
        matches_despesas = 0
        matches_custodias = 0
        erros = []

        # Lista para armazenar sugestões de matches (apenas por valor, sem nome)
        sugestoes = []

        # Dicionários para rastrear alocações em tempo real durante o loop
        # Chave: ID da entidade, Valor: total adicional alocado neste loop
        receitas_alocadas_no_loop = {}
        despesas_alocadas_no_loop = {}
        custodias_alocadas_no_loop = {}

        # Processa cada payment sem alocação
        for payment in payments_sem_alocacao:
            match_found = False

            if payment.tipo == 'E':
                # Entrada: busca receitas com VALOR EXATO E NOME NA OBSERVAÇÃO (ambas condições obrigatórias)
                for receita in receitas_abertas:
                    # Calcula total alocado = annotation inicial + alocações feitas neste loop
                    total_alocado_atual = receita.total_alocado + receitas_alocadas_no_loop.get(receita.id, Decimal('0.00'))
                    valor_nao_alocado = receita.valor - total_alocado_atual

                    # Verifica se há saldo disponível e se o valor do payment é compatível
                    if valor_nao_alocado >= payment.valor and receita.valor == payment.valor:
                        # Verifica se o nome do cliente ou nome da receita está na observação
                        nome_encontrado = nome_em_observacao(
                            payment.observacao,
                            receita.cliente.nome if receita.cliente else None,
                            receita.nome
                        )

                        # APENAS faz match se AMBAS as condições forem verdadeiras
                        if nome_encontrado:
                            try:
                                Allocation.objects.create(
                                    company=company,
                                    payment=payment,
                                    receita=receita,
                                    valor=payment.valor
                                )
                                receita.atualizar_status()

                                # Atualiza o dicionário de alocações em tempo real
                                receitas_alocadas_no_loop[receita.id] = receitas_alocadas_no_loop.get(receita.id, Decimal('0.00')) + payment.valor

                                matches_receitas += 1
                                match_found = True
                                break  # Encontrou match válido, para de procurar
                            except Exception as e:
                                erros.append(f'Erro ao alocar payment {payment.id} para receita {receita.id}: {str(e)}')

                # Se não encontrou receita, tenta custodia tipo Ativo (VALOR E NOME obrigatórios)
                if not match_found:
                    for custodia in custodias_abertas:
                        if custodia.tipo == 'A':  # Ativo - a receber
                            # Calcula totais considerando alocações feitas neste loop
                            alocacoes_loop = custodias_alocadas_no_loop.get(custodia.id, {'entradas': Decimal('0.00'), 'saidas': Decimal('0.00')})
                            total_entradas = custodia.total_entradas + alocacoes_loop['entradas']
                            total_saidas = custodia.total_saidas + alocacoes_loop['saidas']

                            valor_liquidado = min(total_saidas, total_entradas)
                            valor_restante = custodia.valor_total - valor_liquidado

                            # Verifica se há saldo disponível E se o valor total da custódia é exato
                            if valor_restante >= payment.valor and custodia.valor_total == payment.valor:
                                # Verifica se nome do cliente/funcionário ou nome da custódia está na observação
                                nome_encontrado = nome_em_observacao(
                                    payment.observacao,
                                    custodia.cliente.nome if custodia.cliente else None,
                                    custodia.funcionario.nome if custodia.funcionario else None,
                                    custodia.nome
                                )

                                # APENAS faz match se AMBAS as condições forem verdadeiras
                                if nome_encontrado:
                                    try:
                                        Allocation.objects.create(
                                            company=company,
                                            payment=payment,
                                            custodia=custodia,
                                            valor=payment.valor
                                        )
                                        custodia.atualizar_status()

                                        # Atualiza o dicionário de alocações em tempo real (entrada para ativo)
                                        if custodia.id not in custodias_alocadas_no_loop:
                                            custodias_alocadas_no_loop[custodia.id] = {'entradas': Decimal('0.00'), 'saidas': Decimal('0.00')}
                                        custodias_alocadas_no_loop[custodia.id]['entradas'] += payment.valor

                                        matches_custodias += 1
                                        match_found = True
                                        break  # Encontrou match válido, para de procurar
                                    except Exception as e:
                                        erros.append(f'Erro ao alocar payment {payment.id} para custódia {custodia.id}: {str(e)}')

            elif payment.tipo == 'S':
                # Saída: busca despesas com VALOR EXATO E NOME NA OBSERVAÇÃO (ambas condições obrigatórias)
                for despesa in despesas_abertas:
                    # Calcula total alocado = annotation inicial + alocações feitas neste loop
                    total_alocado_atual = despesa.total_alocado + despesas_alocadas_no_loop.get(despesa.id, Decimal('0.00'))
                    valor_nao_alocado = despesa.valor - total_alocado_atual

                    # Verifica se há saldo disponível e se o valor do payment é compatível
                    if valor_nao_alocado >= payment.valor and despesa.valor == payment.valor:
                        # Verifica se o nome do responsável ou nome da despesa está na observação
                        nome_encontrado = nome_em_observacao(
                            payment.observacao,
                            despesa.responsavel.nome if despesa.responsavel else None,
                            despesa.nome
                        )

                        # APENAS faz match se AMBAS as condições forem verdadeiras
                        if nome_encontrado:
                            try:
                                Allocation.objects.create(
                                    company=company,
                                    payment=payment,
                                    despesa=despesa,
                                    valor=payment.valor
                                )
                                despesa.atualizar_status()

                                # Atualiza o dicionário de alocações em tempo real
                                despesas_alocadas_no_loop[despesa.id] = despesas_alocadas_no_loop.get(despesa.id, Decimal('0.00')) + payment.valor

                                matches_despesas += 1
                                match_found = True
                                break  # Encontrou match válido, para de procurar
                            except Exception as e:
                                erros.append(f'Erro ao alocar payment {payment.id} para despesa {despesa.id}: {str(e)}')

                # Se não encontrou despesa, tenta custodia tipo Passivo (VALOR E NOME obrigatórios)
                if not match_found:
                    for custodia in custodias_abertas:
                        if custodia.tipo == 'P':  # Passivo - a pagar
                            # Calcula totais considerando alocações feitas neste loop
                            alocacoes_loop = custodias_alocadas_no_loop.get(custodia.id, {'entradas': Decimal('0.00'), 'saidas': Decimal('0.00')})
                            total_entradas = custodia.total_entradas + alocacoes_loop['entradas']
                            total_saidas = custodia.total_saidas + alocacoes_loop['saidas']
                            valor_liquidado = min(total_entradas, total_saidas)
                            valor_restante = custodia.valor_total - valor_liquidado

                            # Verifica se há saldo disponível E se o valor total da custódia é exato
                            if valor_restante >= payment.valor and custodia.valor_total == payment.valor:
                                # Verifica se nome do cliente/funcionário ou nome da custódia está na observação
                                nome_encontrado = nome_em_observacao(
                                    payment.observacao,
                                    custodia.cliente.nome if custodia.cliente else None,
                                    custodia.funcionario.nome if custodia.funcionario else None,
                                    custodia.nome
                                )

                                # APENAS faz match se AMBAS as condições forem verdadeiras
                                if nome_encontrado:
                                    try:
                                        Allocation.objects.create(
                                            company=company,
                                            payment=payment,
                                            custodia=custodia,
                                            valor=payment.valor
                                        )
                                        custodia.atualizar_status()

                                        # Atualiza o dicionário de alocações em tempo real (saída para passivo)
                                        if custodia.id not in custodias_alocadas_no_loop:
                                            custodias_alocadas_no_loop[custodia.id] = {'entradas': Decimal('0.00'), 'saidas': Decimal('0.00')}
                                        custodias_alocadas_no_loop[custodia.id]['saidas'] += payment.valor

                                        matches_custodias += 1
                                        match_found = True
                                        break  # Encontrou match válido, para de procurar
                                    except Exception as e:
                                        erros.append(f'Erro ao alocar payment {payment.id} para custódia {custodia.id}: {str(e)}')

            # Se não houve match automático, coleta sugestões apenas por valor
            if not match_found:
                sugestoes_payment = []

                if payment.tipo == 'E':
                    # Sugestões de receitas com mesmo valor E saldo disponível
                    for receita in receitas_abertas:
                        # Calcula o valor não alocado
                        total_alocado = receita.allocations.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
                        valor_nao_alocado = receita.valor - total_alocado

                        # Só sugere se há saldo disponível suficiente
                        if valor_nao_alocado >= payment.valor and receita.valor == payment.valor:
                            sugestoes_payment.append({
                                'tipo': 'receita',
                                'entidade_id': receita.id,
                                'entidade_nome': receita.nome,
                                'entidade_cliente': receita.cliente.nome if receita.cliente else None,
                                'entidade_valor': str(receita.valor),
                                'entidade_vencimento': receita.data_vencimento.isoformat()
                            })

                    # Sugestões de custódias Ativo com mesmo valor E saldo disponível
                    for custodia in custodias_abertas:
                        if custodia.tipo == 'A':
                            # Calcula valor restante baseado nas allocations atuais
                            total_entradas = custodia.allocations.filter(
                                payment__tipo='E'
                            ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

                            total_saidas = custodia.allocations.filter(
                                payment__tipo='S'
                            ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

                            valor_liquidado = min(total_saidas, total_entradas)
                            valor_restante = custodia.valor_total - valor_liquidado

                            # Só sugere se há saldo disponível suficiente E valor é exatamente igual
                            if valor_restante >= payment.valor and valor_restante == payment.valor:
                                contraparte = None
                                if custodia.cliente:
                                    contraparte = custodia.cliente.nome
                                elif custodia.funcionario:
                                    contraparte = custodia.funcionario.nome

                                sugestoes_payment.append({
                                    'tipo': 'custodia',
                                    'entidade_id': custodia.id,
                                    'entidade_nome': custodia.nome,
                                    'entidade_contraparte': contraparte,
                                    'entidade_valor': str(valor_restante),
                                    'entidade_tipo': 'Ativo'
                                })

                elif payment.tipo == 'S':
                    # Sugestões de despesas com mesmo valor E saldo disponível
                    for despesa in despesas_abertas:
                        # Calcula o valor não alocado
                        total_alocado = despesa.allocations.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
                        valor_nao_alocado = despesa.valor - total_alocado

                        # Só sugere se há saldo disponível suficiente
                        if valor_nao_alocado >= payment.valor and despesa.valor == payment.valor:
                            sugestoes_payment.append({
                                'tipo': 'despesa',
                                'entidade_id': despesa.id,
                                'entidade_nome': despesa.nome,
                                'entidade_responsavel': despesa.responsavel.nome if despesa.responsavel else None,
                                'entidade_valor': str(despesa.valor),
                                'entidade_vencimento': despesa.data_vencimento.isoformat()
                            })

                    # Sugestões de custódias Passivo com mesmo valor E saldo disponível
                    for custodia in custodias_abertas:
                        if custodia.tipo == 'P':
                            # Calcula valor restante baseado nas allocations atuais
                            total_entradas = custodia.allocations.filter(
                                payment__tipo='E'
                            ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

                            total_saidas = custodia.allocations.filter(
                                payment__tipo='S'
                            ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

                            valor_liquidado = min(total_entradas, total_saidas)
                            valor_restante = custodia.valor_total - valor_liquidado

                            # Só sugere se há saldo disponível suficiente E valor é exatamente igual
                            if valor_restante >= payment.valor and valor_restante == payment.valor:
                                contraparte = None
                                if custodia.cliente:
                                    contraparte = custodia.cliente.nome
                                elif custodia.funcionario:
                                    contraparte = custodia.funcionario.nome

                                sugestoes_payment.append({
                                    'tipo': 'custodia',
                                    'entidade_id': custodia.id,
                                    'entidade_nome': custodia.nome,
                                    'entidade_contraparte': contraparte,
                                    'entidade_valor': str(valor_restante),
                                    'entidade_tipo': 'Passivo'
                                })

                # Se há sugestões, adiciona à lista
                if sugestoes_payment:
                    sugestoes.append({
                        'payment_id': payment.id,
                        'payment_tipo': payment.get_tipo_display(),
                        'payment_valor': str(payment.valor),
                        'payment_data': payment.data_pagamento.isoformat(),
                        'payment_observacao': payment.observacao or '',
                        'payment_conta': payment.conta_bancaria.nome if payment.conta_bancaria else None,
                        'opcoes': sugestoes_payment
                    })

        return Response({
            'success': True,
            'mes': mes,
            'ano': ano,
            'total_payments_processados': payments_sem_alocacao.count(),
            'matches': {
                'receitas': matches_receitas,
                'despesas': matches_despesas,
                'custodias': matches_custodias,
                'total': matches_receitas + matches_despesas + matches_custodias
            },
            'sugestoes': sugestoes,
            'total_sugestoes': len(sugestoes),
            'debug': {
                'total_receitas_abertas': receitas_abertas.count(),
                'total_despesas_abertas': despesas_abertas.count(),
                'total_custodias_abertas': custodias_abertas.count(),
                'payments_entrada': payments_sem_alocacao.filter(tipo='E').count(),
                'payments_saida': payments_sem_alocacao.filter(tipo='S').count(),
            },
            'erros': erros
        })

    @action(detail=False, methods=['post'], url_path='confirmar-sugestao')
    def confirmar_sugestao(self, request):
        """
        Confirma uma sugestão de match manual, criando a alocação.

        Espera:
        - payment_id: ID do pagamento
        - tipo: 'receita', 'despesa' ou 'custodia'
        - entidade_id: ID da entidade (receita/despesa/custodia)
        """
        payment_id = request.data.get('payment_id')
        tipo = request.data.get('tipo')
        entidade_id = request.data.get('entidade_id')

        if not all([payment_id, tipo, entidade_id]):
            return Response(
                {'error': 'Parâmetros payment_id, tipo e entidade_id são obrigatórios'},
                status=status.HTTP_400_BAD_REQUEST
            )

        company = request.user.company

        # Busca o payment
        try:
            payment = Payment.objects.get(id=payment_id, company=company)
        except Payment.DoesNotExist:
            return Response(
                {'error': 'Pagamento não encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Verifica se o payment já tem alocação
        if payment.allocations.exists():
            return Response(
                {'error': 'Este pagamento já possui alocação'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Cria a alocação baseado no tipo
        try:
            if tipo == 'receita':
                receita = Receita.objects.get(id=entidade_id, company=company)
                if receita.valor != payment.valor:
                    return Response(
                        {'error': 'Valor da receita não corresponde ao valor do pagamento'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                Allocation.objects.create(
                    company=company,
                    payment=payment,
                    receita=receita,
                    valor=payment.valor
                )
                receita.atualizar_status()
                entidade_nome = receita.nome

            elif tipo == 'despesa':
                despesa = Despesa.objects.get(id=entidade_id, company=company)
                if despesa.valor != payment.valor:
                    return Response(
                        {'error': 'Valor da despesa não corresponde ao valor do pagamento'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                Allocation.objects.create(
                    company=company,
                    payment=payment,
                    despesa=despesa,
                    valor=payment.valor
                )
                despesa.atualizar_status()
                entidade_nome = despesa.nome

            elif tipo == 'custodia':
                custodia = Custodia.objects.get(id=entidade_id, company=company)
                valor_restante = custodia.valor_total - custodia.valor_liquidado
                if valor_restante != payment.valor:
                    return Response(
                        {'error': 'Valor restante da custódia não corresponde ao valor do pagamento'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                Allocation.objects.create(
                    company=company,
                    payment=payment,
                    custodia=custodia,
                    valor=payment.valor
                )
                custodia.atualizar_status()
                entidade_nome = custodia.nome

            else:
                return Response(
                    {'error': 'Tipo inválido. Use: receita, despesa ou custodia'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            return Response({
                'success': True,
                'message': f'Pagamento vinculado com sucesso a {tipo}: {entidade_nome}',
                'payment_id': payment.id,
                'tipo': tipo,
                'entidade_id': entidade_id
            })

        except (Receita.DoesNotExist, Despesa.DoesNotExist, Custodia.DoesNotExist):
            return Response(
                {'error': f'{tipo.capitalize()} não encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Erro ao criar alocação: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def perform_destroy(self, instance):
        from django.db.models import F

        conta_id = instance.conta_bancaria.pk
        valor = instance.valor
        tipo = instance.tipo

        # Guarda as alocações antes de deletar
        allocations = list(instance.allocations.all())

        # Deleta o pagamento primeiro
        instance.delete()

        # Reverte o pagamento do saldo usando F() para operação atômica
        if tipo == 'E':
            # Remove entrada
            ContaBancaria.objects.filter(pk=conta_id).update(
                saldo_atual=F('saldo_atual') - valor
            )
        else:
            # Remove saída
            ContaBancaria.objects.filter(pk=conta_id).update(
                saldo_atual=F('saldo_atual') + valor
            )

        # Atualiza status de todas as contas que estavam alocadas
        for allocation in allocations:
            if allocation.receita:
                allocation.receita.atualizar_status()
            elif allocation.despesa:
                allocation.despesa.atualizar_status()
            elif allocation.custodia:
                allocation.custodia.atualizar_status()

class ContaBancariaViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """API endpoint para gerenciar contas bancárias."""
    queryset = ContaBancaria.objects.all()
    serializer_class = ContaBancariaSerializer
    pagination_class = DynamicPageSizePagination


class CustodiaViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """API endpoint para gerenciar custódias (valores de terceiros - ativos e passivos)."""
    queryset = Custodia.objects.all()
    serializer_class = CustodiaSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        queryset = super().get_queryset().select_related('cliente', 'funcionario')

        params = self.request.query_params

        # Filtro por tipo (Ativo ou Passivo)
        tipo_filter = params.get('tipo')
        if tipo_filter:
            queryset = queryset.filter(tipo=tipo_filter)

        # Filtro por status (aceita múltiplos valores)
        status_filter = params.getlist('status')
        if status_filter:
            queryset = queryset.filter(status__in=status_filter)
        elif params.get('status'):
            # Fallback para single value
            queryset = queryset.filter(status=params.get('status'))

        # Filtro por cliente
        cliente_id = params.get('cliente_id')
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)

        # Filtro por funcionário
        funcionario_id = params.get('funcionario_id')
        if funcionario_id:
            queryset = queryset.filter(funcionario_id=funcionario_id)

        # Busca global
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(nome__icontains=search) |
                Q(descricao__icontains=search) |
                Q(cliente__nome__icontains=search) |
                Q(funcionario__nome__icontains=search)
            )

        return queryset.order_by('-criado_em', 'id')


class TransferViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """API endpoint para gerenciar transferências entre contas bancárias."""
    queryset = Transfer.objects.all()
    serializer_class = TransferSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        from django.db.models.functions import Coalesce

        queryset = super().get_queryset().select_related('from_bank', 'to_bank').annotate(
            valor_saida=Coalesce(
                Sum('allocations__valor', filter=Q(allocations__payment__tipo='S')),
                Decimal('0.00')
            ),
            valor_entrada=Coalesce(
                Sum('allocations__valor', filter=Q(allocations__payment__tipo='E')),
                Decimal('0.00')
            )
        )

        params = self.request.query_params

        # Filtro por banco de origem
        from_bank_id = params.get('from_bank_id')
        if from_bank_id:
            queryset = queryset.filter(from_bank_id=from_bank_id)

        # Filtro por banco de destino
        to_bank_id = params.get('to_bank_id')
        if to_bank_id:
            queryset = queryset.filter(to_bank_id=to_bank_id)

        # Filtro por status (aceita múltiplos valores)
        status_filter = params.getlist('status')
        if status_filter:
            queryset = queryset.filter(status__in=status_filter)
        elif params.get('status'):
            # Fallback para single value
            queryset = queryset.filter(status=params.get('status'))

        # Filtro por data
        data_inicio = params.get('data_inicio')
        if data_inicio:
            queryset = queryset.filter(data_transferencia__gte=data_inicio)

        data_fim = params.get('data_fim')
        if data_fim:
            queryset = queryset.filter(data_transferencia__lte=data_fim)

        # Busca global
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(descricao__icontains=search) |
                Q(from_bank__nome__icontains=search) |
                Q(to_bank__nome__icontains=search)
            )

        return queryset.order_by('-data_transferencia', '-criado_em', 'id')


class AllocationViewSet(CompanyScopedViewSetMixin, viewsets.ModelViewSet):
    """API endpoint para gerenciar alocações de pagamentos."""
    queryset = Allocation.objects.all()
    serializer_class = AllocationSerializer
    pagination_class = DynamicPageSizePagination

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'payment', 'payment__conta_bancaria',
            'receita', 'receita__cliente',
            'despesa', 'despesa__responsavel',
            'custodia', 'custodia__cliente', 'custodia__funcionario',
            'transfer', 'transfer__from_bank', 'transfer__to_bank'
        )

        params = self.request.query_params

        # Filtro por payment
        payment_id = params.get('payment_id')
        if payment_id:
            queryset = queryset.filter(payment_id=payment_id)

        # Filtro por receita
        receita_id = params.get('receita_id')
        if receita_id:
            queryset = queryset.filter(receita_id=receita_id)

        # Filtro por despesa
        despesa_id = params.get('despesa_id')
        if despesa_id:
            queryset = queryset.filter(despesa_id=despesa_id)

        # Filtro por custódia
        custodia_id = params.get('custodia_id')
        if custodia_id:
            queryset = queryset.filter(custodia_id=custodia_id)

        # Filtro por transferência
        transfer_id = params.get('transfer_id')
        if transfer_id:
            queryset = queryset.filter(transfer_id=transfer_id)

        # Filtro por tipo de conta
        tipo_conta = params.get('tipo_conta')  # 'receita', 'despesa', 'custodia', 'transfer'
        if tipo_conta == 'receita':
            queryset = queryset.filter(receita__isnull=False)
        elif tipo_conta == 'despesa':
            queryset = queryset.filter(despesa__isnull=False)
        elif tipo_conta == 'custodia':
            queryset = queryset.filter(custodia__isnull=False)
        elif tipo_conta == 'transfer':
            queryset = queryset.filter(transfer__isnull=False)

        # Busca global
        search = params.get('search')
        if search:
            queryset = queryset.filter(
                Q(observacao__icontains=search) |
                Q(valor__icontains=search) |
                Q(payment__observacao__icontains=search) |
                Q(receita__nome__icontains=search) |
                Q(despesa__nome__icontains=search) |
                Q(custodia__nome__icontains=search)
            )

        return queryset.order_by('-criado_em', 'id')

    def perform_create(self, serializer):
        allocation = serializer.save(company=self.request.user.company)

        # Atualizar status da conta após criar alocação
        if allocation.receita:
            allocation.receita.atualizar_status()
        elif allocation.despesa:
            allocation.despesa.atualizar_status()
        elif allocation.custodia:
            allocation.custodia.atualizar_status()
        elif allocation.transfer:
            allocation.transfer.atualizar_status()

    def perform_update(self, serializer):
        # Guarda referências antigas antes de atualizar
        old_allocation = Allocation.objects.get(pk=serializer.instance.pk)
        old_receita = old_allocation.receita
        old_despesa = old_allocation.despesa
        old_custodia = old_allocation.custodia
        old_transfer = old_allocation.transfer

        # Salva a nova alocação
        allocation = serializer.save()

        # Atualiza status da conta antiga (se mudou)
        if old_receita and old_receita != allocation.receita:
            old_receita.atualizar_status()
        if old_despesa and old_despesa != allocation.despesa:
            old_despesa.atualizar_status()
        if old_custodia and old_custodia != allocation.custodia:
            old_custodia.atualizar_status()
        if old_transfer and old_transfer != allocation.transfer:
            old_transfer.atualizar_status()

        # Atualiza status da conta nova
        if allocation.receita:
            allocation.receita.atualizar_status()
        elif allocation.despesa:
            allocation.despesa.atualizar_status()
        elif allocation.custodia:
            allocation.custodia.atualizar_status()
        elif allocation.transfer:
            allocation.transfer.atualizar_status()

    def perform_destroy(self, instance):
        receita = instance.receita
        despesa = instance.despesa
        custodia = instance.custodia
        transfer = instance.transfer

        # Deleta a alocação
        instance.delete()

        # Atualiza status da conta após deletar alocação
        if receita:
            receita.atualizar_status()
        if despesa:
            despesa.atualizar_status()
        if custodia:
            custodia.atualizar_status()
        if transfer:
            transfer.atualizar_status()


# --- Report Views (Placeholder - Step 7 will detail these) ---
# These will likely be separate APIView or function-based views, not ViewSets

# Example structure (to be implemented in Step 7)
# class RelatorioClienteView(generics.ListAPIView):
#     permission_classes = [permissions.IsAuthenticated]
#     serializer_class = ReceitaSerializer # Or a custom report serializer
# 
#     def get_queryset(self):
#         user = self.request.user
#         cliente_id = self.kwargs.get('cliente_id') # Get from URL
#         # ... filter Receitas for this cliente_id and user's company ...
#         pass

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Sum, Q, F
from django.utils import timezone
from datetime import timedelta, date
from decimal import Decimal
from .models import (
    Company, CustomUser, Cliente, Funcionario, Receita, Despesa, 
    Payment, ContaBancaria
)

def _atualizar_vencidas_company(company):
    """Atualiza automaticamente receitas e despesas vencidas de uma empresa."""
    hoje = timezone.now().date()
    Receita.objects.filter(company=company, situacao='A', data_vencimento__lt=hoje).update(situacao='V')
    Despesa.objects.filter(company=company, situacao='A', data_vencimento__lt=hoje).update(situacao='V')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_view(request):
    """
    Retorna dados consolidados do dashboard para o usuário autenticado.
    (Alinhado ao modelo financeiro do sistema: Payment = fonte da verdade)
    """
    user = request.user
    company = user.company

    # 🔄 Atualiza vencidas antes de calcular o dashboard
    _atualizar_vencidas_company(company)

    hoje = timezone.now().date()
    inicio_mes = date(hoje.year, hoje.month, 1)
    
    # Data de 30 dias atrás
    data_30_dias_atras = hoje - timedelta(days=30)
    
    # ======================================================
    # 💰 FLUXO DE CAIXA REALIZADO (ÚLTIMOS 30 DIAS)
    # ======================================================

    # Entradas dos últimos 30 dias (todos os pagamentos tipo 'E', exceto transferências)
    receitas_30_dias = (
        Payment.objects.filter(
            company=company,
            tipo='E',
            data_pagamento__gte=data_30_dias_atras,
            data_pagamento__lte=hoje
        )
        .exclude(allocations__transfer__isnull=False)
        .aggregate(total=Sum('valor'))['total']
        or Decimal('0.00')
    )

    # Saídas dos últimos 30 dias (todos os pagamentos tipo 'S', exceto transferências)
    despesas_30_dias = (
        Payment.objects.filter(
            company=company,
            tipo='S',
            data_pagamento__gte=data_30_dias_atras,
            data_pagamento__lte=hoje
        )
        .exclude(allocations__transfer__isnull=False)
        .aggregate(total=Sum('valor'))['total']
        or Decimal('0.00')
    )

    # Fluxo de caixa realizado (o que entrou - o que saiu)
    fluxo_caixa_realizado = receitas_30_dias - despesas_30_dias

    # ======================================================
    # 🏦 SALDO TOTAL DAS CONTAS
    # ======================================================

    saldo_total = (
        ContaBancaria.objects.filter(company=company)
        .aggregate(total=Sum('saldo_atual'))['total']
        or Decimal('0.00')
    )
    
    # Saldo de 30 dias atrás (para comparação)
    # Calculamos: saldo_atual - fluxo_realizado
    saldo_30_dias_atras = saldo_total - fluxo_caixa_realizado

    # ======================================================
    # 📊 RECEITAS PROJETADAS (PRÓXIMOS 30 DIAS)
    # ======================================================
    
    data_limite = hoje + timedelta(days=30)
    
    receitas_projetadas = (
        Receita.objects.filter(
            company=company,
            data_vencimento__gte=hoje,
            data_vencimento__lte=data_limite,
            situacao__in=['A', 'V']  # Não paga ainda
        )
        .annotate(
            total_alocado=Coalesce(
                Sum('allocations__valor'),
                Decimal('0.00'),
                output_field=DecimalField()
            )
        )
        .aggregate(
            total=Sum(F('valor') - F('total_alocado'), output_field=DecimalField())
        )['total']
        or Decimal('0.00')
    )
    
    # ======================================================
    # 📊 DESPESAS PROJETADAS (PRÓXIMOS 30 DIAS)
    # ======================================================
    
    despesas_projetadas = (
        Despesa.objects.filter(
            company=company,
            data_vencimento__gte=hoje,
            data_vencimento__lte=data_limite,
            situacao__in=['A', 'V']  # Não paga ainda
        )
        .annotate(
            total_alocado=Coalesce(
                Sum('allocations__valor'),
                Decimal('0.00'),
                output_field=DecimalField()
            )
        )
        .aggregate(
            total=Sum(F('valor') - F('total_alocado'), output_field=DecimalField())
        )['total']
        or Decimal('0.00')
    )

    # ======================================================
    # 🎂 ANIVERSARIANTES DO DIA
    # ======================================================
    
    hoje_mes_dia = hoje.strftime('%m-%d')
    
    # Clientes aniversariantes
    clientes_aniversariantes = Cliente.objects.filter(
        company=company,
        aniversario__isnull=False,
        aniversario__month=hoje.month,      # ← Novo
        aniversario__day=hoje.day            # ← Novo
    )

    # Funcionários aniversariantes
    funcionarios_aniversariantes = Funcionario.objects.filter(
        company=company,
        aniversario__isnull=False,
        aniversario__month=hoje.month,      # ← Novo
        aniversario__day=hoje.day            # ← Novo
    )

    
    aniversariantes = {
        'clientes': [
            {
                'id': c.id,
                'nome': c.nome,
                'tipo': 'Cliente',
                'email': c.email,
                'telefone': c.telefone
            }
            for c in clientes_aniversariantes
        ],
        'funcionarios': [
            {
                'id': f.id,
                'nome': f.nome,
                'tipo': f.get_tipo_display(),
                'email': f.email,
                'telefone': f.telefone
            }
            for f in funcionarios_aniversariantes
        ]
    }

    # ======================================================
    # 🚨 ALERTAS OPERACIONAIS (VENCIDAS)
    # ======================================================

    despesas_vencidas = Despesa.objects.filter(
        company=company,
        situacao='V'
    ).count()

    receitas_vencidas = Receita.objects.filter(
        company=company,
        situacao='V'
    ).count()

    valor_despesas_vencidas = (
        Despesa.objects.filter(
            company=company,
            situacao='V'
        )
        .annotate(
            total_alocado=Coalesce(
                Sum('allocations__valor'),
                Decimal('0.00'),
                output_field=DecimalField()
            )
        )
        .aggregate(
            total=Sum(F('valor') - F('total_alocado'), output_field=DecimalField())
        )['total']
        or Decimal('0.00')
    )

    valor_receitas_vencidas = (
        Receita.objects.filter(
            company=company,
            situacao='V'
        )
        .annotate(
            total_alocado=Coalesce(
                Sum('allocations__valor'),
                Decimal('0.00'),
                output_field=DecimalField()
            )
        )
        .aggregate(
            total=Sum(F('valor') - F('total_alocado'), output_field=DecimalField())
        )['total']
        or Decimal('0.00')
    )

    # ======================================================
    # 📊 GRÁFICO RECEITA x DESPESA (ÚLTIMOS 6 MESES - BRUTO)
    # ======================================================

    meses_data = []

    for i in range(5, -1, -1):
        ref = inicio_mes - timedelta(days=30 * i)
        mes_inicio = ref.replace(day=1)
        mes_fim = (mes_inicio.replace(day=28) + timedelta(days=4)).replace(
            day=1
        ) - timedelta(days=1)

        receita = (
            Receita.objects.filter(
                company=company,
                data_vencimento__gte=mes_inicio,
                data_vencimento__lte=mes_fim
            )
            .aggregate(total=Sum('valor'))['total']
            or Decimal('0.00')
        )

        despesa = (
            Despesa.objects.filter(
                company=company,
                data_vencimento__gte=mes_inicio,
                data_vencimento__lte=mes_fim
            )
            .aggregate(total=Sum('valor'))['total']
            or Decimal('0.00')
        )

        meses_data.append({
            'mes': mes_inicio.strftime('%b'),
            'receita': float(receita),
            'despesa': float(despesa),
        })

    # ======================================================
    # 📊 GRÁFICO FLUXO DE CAIXA REALIZADO (ÚLTIMOS 6 MESES)
    # ======================================================
    
    fluxo_caixa_data = []
    
    for i in range(5, -1, -1):
        ref = inicio_mes - timedelta(days=30 * i)
        mes_inicio = ref.replace(day=1)
        mes_fim = (mes_inicio.replace(day=28) + timedelta(days=4)).replace(
            day=1
        ) - timedelta(days=1)

        receita_mes = (
            Payment.objects.filter(
                company=company,
                tipo='E',
                data_pagamento__gte=mes_inicio,
                data_pagamento__lte=mes_fim
            )
            .exclude(allocations__transfer__isnull=False)
            .aggregate(total=Sum('valor'))['total']
            or Decimal('0.00')
        )

        despesa_mes = (
            Payment.objects.filter(
                company=company,
                tipo='S',
                data_pagamento__gte=mes_inicio,
                data_pagamento__lte=mes_fim
            )
            .exclude(allocations__transfer__isnull=False)
            .aggregate(total=Sum('valor'))['total']
            or Decimal('0.00')
        )
        
        fluxo_mes = receita_mes - despesa_mes

        fluxo_caixa_data.append({
            'mes': mes_inicio.strftime('%b'),
            'fluxo': float(fluxo_mes),
            'receita': float(receita_mes),
            'despesa': float(despesa_mes),
        })

    # ======================================================
    # 🍰 RECEITA / DESPESA POR TIPO (PAGO)
    # ======================================================

    receita_por_tipo = []
    for tipo, label in Receita.TIPO_CHOICES:
        total = (
            Allocation.objects.filter(
                company=company,
                receita__tipo=tipo
            )
            .aggregate(total=Sum('valor'))['total']
            or Decimal('0.00')
        )

        if total > 0:
            receita_por_tipo.append({
                'name': label,
                'value': float(total),
            })

    despesa_por_tipo = []
    for tipo, label in Despesa.TIPO_CHOICES:
        total = (
            Allocation.objects.filter(
                company=company,
                despesa__tipo=tipo
            )
            .aggregate(total=Sum('valor'))['total']
            or Decimal('0.00')
        )

        if total > 0:
            despesa_por_tipo.append({
                'name': label,
                'value': float(total),
            })

    # ======================================================
    # ⏰ PRÓXIMOS VENCIMENTOS
    # ======================================================

    data_limite = hoje + timedelta(days=5)

    receitas_proximas = (
        Receita.objects.filter(
            company=company,
            data_vencimento__gte=hoje,
            data_vencimento__lte=data_limite,
            situacao__in=['A', 'V']
        )
        .select_related('cliente')
        .order_by('data_vencimento')[:5]
    )

    despesas_proximas = (
        Despesa.objects.filter(
            company=company,
            data_vencimento__gte=hoje,
            data_vencimento__lte=data_limite,
            situacao__in=['A', 'V']
        )
        .select_related('responsavel')
        .order_by('data_vencimento')[:5]
    )

    # ======================================================
    # 📦 RESPONSE
    # ======================================================

    return Response({
        # Saldo e Fluxo
        'saldoTotal': float(saldo_total),
        'saldo30DiasAtras': float(saldo_30_dias_atras),
        'fluxoCaixaRealizado': float(fluxo_caixa_realizado),
        
        # Projeções (próximos 30 dias)
        'receitasProjetadas': float(receitas_projetadas),
        'despesasProjetadas': float(despesas_projetadas),
        
        # Alertas
        'despesasVencidas': despesas_vencidas,
        'receitasVencidas': receitas_vencidas,
        'valorDespesasVencidas': float(valor_despesas_vencidas),
        'valorReceitasVencidas': float(valor_receitas_vencidas),
        
        # Aniversariantes
        'aniversariantes': aniversariantes,

        # Gráficos
        'receitaVsDespesaData': meses_data,
        'fluxoCaixaData': fluxo_caixa_data,
        'receitaPorTipoData': receita_por_tipo,
        'despesaPorTipoData': despesa_por_tipo,

        # Próximos vencimentos
        'receitasProximas': [
            {
                'id': r.id,
                'nome': r.nome,
                'cliente': r.cliente.nome,
                'valor': float(r.valor),
                'dataVencimento': r.data_vencimento.isoformat(),
                'situacao': r.situacao,
            }
            for r in receitas_proximas
        ],

        'despesasProximas': [
            {
                'id': d.id,
                'nome': d.nome,
                'responsavel': d.responsavel.nome,
                'valor': float(d.valor),
                'dataVencimento': d.data_vencimento.isoformat(),
                'situacao': d.situacao,
            }
            for d in despesas_proximas
        ],
    })

# --- Report Views ---
from rest_framework.views import APIView
from django.db.models.functions import TruncMonth

class BaseReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get_company_queryset(self, model):
        user = self.request.user
        if user.is_superuser:
            return model.objects.all()
        if hasattr(user, 'company') and user.company:
            return model.objects.filter(company=user.company)
        return model.objects.none()

    def get_common_filters(self):
        params = self.request.query_params
        start_date = params.get('start_date')
        end_date = params.get('end_date')
        filters = {}
        if start_date:
            filters['data_vencimento__gte'] = start_date
        if end_date:
            filters['data_vencimento__lte'] = end_date
        return filters

class RelatorioClienteView(BaseReportView):
    """
    Resumo financeiro do cliente baseado em ALLOCATIONS (fonte da verdade)
    """

    def get(self, request, cliente_id):
        # 🔹 Cliente (com escopo da empresa)
        cliente_qs = self.get_company_queryset(Cliente)
        cliente = get_object_or_404(cliente_qs, pk=cliente_id)

        # 🔹 Receitas do cliente
        receitas_qs = self.get_company_queryset(Receita).filter(
            cliente=cliente
        )

        # 🔹 Allocations ligadas às receitas do cliente
        # Apenas receitas pagas - valores que recebemos do cliente
        allocations_qs = self.get_company_queryset(Allocation).filter(
            receita__cliente=cliente
        ).select_related('payment')

        # 🔹 Filtros comuns (data inicial / final, etc)
        filters = self.get_common_filters()
        if filters:
            receitas_qs = receitas_qs.filter(**filters)
            # Para filtros de data em allocations, usar payment__data_pagamento
            if 'data_vencimento__gte' in filters:
                allocations_qs = allocations_qs.filter(
                    payment__data_pagamento__gte=filters['data_vencimento__gte']
                )
            if 'data_vencimento__lte' in filters:
                allocations_qs = allocations_qs.filter(
                    payment__data_pagamento__lte=filters['data_vencimento__lte']
                )

        # 🔹 Soma das allocations por receita
        allocations_por_receita = (
            allocations_qs
            .values("receita_id")
            .annotate(total_pago=Sum("valor"))
        )

        allocations_map = {
            item["receita_id"]: item["total_pago"] or 0
            for item in allocations_por_receita
        }

        # 🔹 Pendências reais (saldo > 0)
        pendings = []
        total_open = 0

        for receita in receitas_qs:
            total_pago = allocations_map.get(receita.id, 0)
            saldo = receita.valor - total_pago

            if saldo > 0:
                pendings.append({
                    "id": receita.id,
                    "nome": receita.nome,
                    "description": receita.descricao,
                    "valor_total": receita.valor,
                    "valor_pago": total_pago,
                    "saldo": saldo,
                    "due_date": receita.data_vencimento,
                })
                total_open += saldo

        # 🔹 Allocations realizadas (histórico)
        allocations = AllocationSerializer(
            allocations_qs.order_by("-payment__data_pagamento"),
            many=True
        ).data

        # 🔹 Total pago
        total_paid = allocations_qs.aggregate(
            total=Sum("valor")
        )["total"] or 0

        # 🔹 Custódias do cliente (Passivos - valores que devemos ao cliente)
        custodias_qs = self.get_company_queryset(Custodia).filter(
            cliente=cliente,
            tipo='P'  # Passivos - valores a pagar
        ).exclude(status='L')

        custodias_a_pagar = []
        total_custodia_pagar = 0

        for custodia in custodias_qs:
            saldo = custodia.valor_total - custodia.valor_liquidado
            if saldo > 0:
                custodias_a_pagar.append({
                    "id": custodia.id,
                    "nome": custodia.nome,
                    "descricao": custodia.descricao,
                    "valor_total": custodia.valor_total,
                    "valor_liquidado": custodia.valor_liquidado,
                    "saldo": saldo,
                    "status": custodia.status,
                })
                total_custodia_pagar += saldo

        return Response({
            "client": ClienteSerializer(cliente).data,
            "pendings": pendings,
            "allocations": allocations,
            "custodias_a_pagar": custodias_a_pagar,
            "totals": {
                "open": total_open,
                "paid": total_paid,
                "custodia_pagar": total_custodia_pagar,
            }
        }, status=status.HTTP_200_OK)

from rest_framework.response import Response
from rest_framework import status

from decimal import Decimal
from rest_framework.response import Response
from rest_framework import status

class RelatorioFuncionarioView(BaseReportView):
    """
    Resumo financeiro do funcionário baseado em ALLOCATIONS (fonte da verdade)
    """

    def get(self, request, funcionario_id):
        # 🔹 Funcionário (escopo da empresa)
        funcionario_qs = self.get_company_queryset(Funcionario)
        funcionario = get_object_or_404(funcionario_qs, pk=funcionario_id)

        # 🔹 Despesas do funcionário
        despesas_qs = self.get_company_queryset(Despesa).filter(
            responsavel=funcionario
        )

        # 🔹 Allocations ligadas às despesas e custódias do funcionário
        # Custódias Passivas com saída (tipo='P' + payment tipo='S') = repasses ao funcionário
        # Custódias Ativas com saída (tipo='A' + payment tipo='S') = pagamentos em nome do funcionário
        allocations_qs = self.get_company_queryset(Allocation).filter(
            Q(despesa__responsavel=funcionario) |
            Q(custodia__funcionario=funcionario, custodia__tipo='P', payment__tipo='S') |
            Q(custodia__funcionario=funcionario, custodia__tipo='A', payment__tipo='S')
        ).select_related('payment')

        # 🔹 Filtros comuns (data inicial / final, etc)
        filters = self.get_common_filters()
        if filters:
            despesas_qs = despesas_qs.filter(**filters)
            # Para filtros de data em allocations, usar payment__data_pagamento
            if 'data_vencimento__gte' in filters:
                allocations_qs = allocations_qs.filter(
                    payment__data_pagamento__gte=filters['data_vencimento__gte']
                )
            if 'data_vencimento__lte' in filters:
                allocations_qs = allocations_qs.filter(
                    payment__data_pagamento__lte=filters['data_vencimento__lte']
                )

        # 🔹 Soma das allocations por despesa
        allocations_por_despesa = (
            allocations_qs
            .values("despesa_id")
            .annotate(total_pago=Sum("valor"))
        )

        allocations_map = {
            item["despesa_id"]: item["total_pago"] or 0
            for item in allocations_por_despesa
        }

        # 🔹 Pendências reais (saldo > 0)
        pendings = []
        total_open = 0

        # Despesas pendentes
        for despesa in despesas_qs:
            total_pago = allocations_map.get(despesa.id, 0)
            saldo = despesa.valor - total_pago

            if saldo > 0:
                pendings.append({
                    "id": despesa.id,
                    "nome": despesa.nome,
                    "description": despesa.descricao,
                    "valor_total": despesa.valor,
                    "valor_pago": total_pago,
                    "saldo": saldo,
                    "due_date": despesa.data_vencimento,
                })
                total_open += saldo

        # Custódias Passivas pendentes (não repassadas ao funcionário)
        custodias_passivas_qs = self.get_company_queryset(Custodia).filter(
            funcionario=funcionario,
            tipo='P'
        ).exclude(status='L')

        for custodia in custodias_passivas_qs:
            saldo = custodia.valor_total - custodia.valor_liquidado
            if saldo > 0:
                pendings.append({
                    "id": f"custodia-{custodia.id}",
                    "nome": custodia.nome,
                    "description": custodia.descricao,
                    "valor_total": custodia.valor_total,
                    "valor_pago": custodia.valor_liquidado,
                    "saldo": saldo,
                    "due_date": None,
                })
                total_open += saldo

        # 🔹 Allocations realizadas (histórico)
        allocations = AllocationSerializer(
            allocations_qs.order_by("-payment__data_pagamento"),
            many=True
        ).data

        # 🔹 Total pago
        total_paid = allocations_qs.aggregate(
            total=Sum("valor")
        )["total"] or 0

        # 🔹 Custódias do funcionário
        custodias_qs = self.get_company_queryset(Custodia).filter(
            funcionario=funcionario
        )

        # Custódias a Pagar (Passivos - valores que devemos ao funcionário)
        custodias_a_pagar = []
        total_custodia_pagar = 0

        for custodia in custodias_qs.filter(tipo='P').exclude(status='L'):
            saldo = custodia.valor_total - custodia.valor_liquidado
            if saldo > 0:
                custodias_a_pagar.append({
                    "id": custodia.id,
                    "nome": custodia.nome,
                    "descricao": custodia.descricao,
                    "valor_total": custodia.valor_total,
                    "valor_liquidado": custodia.valor_liquidado,
                    "saldo": saldo,
                    "status": custodia.status,
                })
                total_custodia_pagar += saldo

        # Custódias a Receber (Ativos - valores que o funcionário nos deve)
        custodias_a_receber = []
        total_custodia_receber = 0

        for custodia in custodias_qs.filter(tipo='A').exclude(status='L'):
            saldo = custodia.valor_total - custodia.valor_liquidado
            if saldo > 0:
                custodias_a_receber.append({
                    "id": custodia.id,
                    "nome": custodia.nome,
                    "descricao": custodia.descricao,
                    "valor_total": custodia.valor_total,
                    "valor_liquidado": custodia.valor_liquidado,
                    "saldo": saldo,
                    "status": custodia.status,
                })
                total_custodia_receber += saldo

        return Response({
            "funcionario": FuncionarioSerializer(funcionario).data,
            "pendings": pendings,
            "allocations": allocations,
            "custodias_a_pagar": custodias_a_pagar,
            "custodias_a_receber": custodias_a_receber,
            "totals": {
                "open": total_open,
                "paid": total_paid,
                "custodia_pagar": total_custodia_pagar,
                "custodia_receber": total_custodia_receber,
            }
        }, status=status.HTTP_200_OK)



class RelatorioTipoPeriodoView(BaseReportView):
    """Relatório de Receitas ou Despesas por Tipo e/ou Período."""
    def get(self, request):
        params = request.query_params
        tipo_relatorio = params.get('tipo_relatorio', '').lower() # 'receita' or 'despesa'
        tipo_item = params.get('tipo_item') # Specific type like 'F', 'V', etc.
        filters = self.get_common_filters()

        if tipo_relatorio == 'receita':
            model = Receita
            serializer = ReceitaSerializer
            if tipo_item:
                filters['tipo'] = tipo_item
        elif tipo_relatorio == 'despesa':
            model = Despesa
            serializer = DespesaSerializer
            if tipo_item:
                filters['tipo'] = tipo_item
        else:
            return Response({"detail": "Parâmetro 'tipo_relatorio' (receita ou despesa) é obrigatório."}, status=status.HTTP_400_BAD_REQUEST)

        queryset = self.get_company_queryset(model).filter(**filters)
        data = serializer(queryset, many=True).data
        return Response(data)

class RelatorioResultadoFinanceiroView(BaseReportView):
    """Relatório de Resultado Financeiro (Receitas Pagas - Despesas Pagas) por Período."""
    def get(self, request):
        filters_pagamento = {}
        params = request.query_params
        start_date = params.get('start_date')
        end_date = params.get('end_date')
        if start_date:
            filters_pagamento['data_pagamento__gte'] = start_date
        if end_date:
            filters_pagamento['data_pagamento__lte'] = end_date

        receitas_qs = self.get_company_queryset(Receita).filter(situacao='P', **filters_pagamento)
        despesas_qs = self.get_company_queryset(Despesa).filter(situacao='P', **filters_pagamento)

        total_receitas_pagas = receitas_qs.aggregate(total=Sum('valor_pago'))['total'] or Decimal('0.00')
        total_despesas_pagas = despesas_qs.aggregate(total=Sum('valor_pago'))['total'] or Decimal('0.00')

        resultado = total_receitas_pagas - total_despesas_pagas

        return Response({
            "periodo_inicio": start_date or "Início",
            "periodo_fim": end_date or "Fim",
            "total_receitas_pagas": total_receitas_pagas,
            "total_despesas_pagas": total_despesas_pagas,
            "resultado_financeiro": resultado
        })

class RelatorioFolhaSalarialView(BaseReportView):
    """Relatório de Folha Salarial Mensal."""
    def get(self, request):
        params = request.query_params
        try:
            year = int(params.get('year', date.today().year))
            month = int(params.get('month', date.today().month))
            target_date = date(year, month, 1)
        except ValueError:
            return Response({"detail": "Ano e/ou mês inválidos."}, status=status.HTTP_400_BAD_REQUEST)

        # Sum of salaries for active 'Funcionário' type employees in the company
        funcionarios_ativos = self.get_company_queryset(Funcionario).filter(tipo='F', salario_mensal__isnull=False)
        total_salarios_base = funcionarios_ativos.aggregate(total=Sum('salario_mensal'))['total'] or Decimal('0.00')

        # Sum of fixed expenses ('F') linked to 'Funcionário' type employees due in the given month/year
        # This interpretation might need refinement based on exact business logic for fixed expenses
        despesas_fixas_func = self.get_company_queryset(Despesa).filter(
            responsavel__tipo='F',
            tipo='F',
            data_vencimento__year=year,
            data_vencimento__month=month
        )
        total_despesas_fixas_func = despesas_fixas_func.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
        
        # Note: This is a simplified view. Real payroll involves taxes, benefits, etc.
        # It also assumes 'salario_mensal' is the gross amount and fixed expenses are additional costs.
        # Clarification might be needed on how 'Despesa Fixa' relates to 'Salário Mensal'.
        # Assuming here they are separate concepts unless explicitly linked.

        return Response({
            "mes": month,
            "ano": year,
            "total_salarios_base": total_salarios_base,
            "total_despesas_fixas_funcionarios (vencimento no mês)": total_despesas_fixas_func,
            "custo_total_estimado_folha": total_salarios_base + total_despesas_fixas_func # Example calculation
        })

class RelatorioComissionamentoView(BaseReportView):
    """Relatório de Comissionamento por Mês por Pessoa ou Todos."""
    def get(self, request):
        params = request.query_params
        try:
            year = int(params.get('year', date.today().year))
            month = int(params.get('month', date.today().month))
        except ValueError:
            return Response({"detail": "Ano e/ou mês inválidos."}, status=status.HTTP_400_BAD_REQUEST)
        
        funcionario_id = params.get('funcionario_id')

        # Filter commission expenses ('C') paid in the given month/year
        despesas_comissao_qs = self.get_company_queryset(Despesa).filter(
            tipo='C',
            # Assuming commission is relevant based on when the *expense* is paid
            # Or should it be based on when the *receita* was paid?
            # Model currently creates commission expense when receita is paid, due immediately.
            # Let's filter by expense payment date for simplicity.
            situacao='P',
            data_pagamento__year=year,
            data_pagamento__month=month
        )

        if funcionario_id:
            despesas_comissao_qs = despesas_comissao_qs.filter(responsavel_id=funcionario_id)
            total_comissao = despesas_comissao_qs.aggregate(total=Sum('valor_pago'))['total'] or Decimal('0.00')
            try:
                funcionario = self.get_company_queryset(Funcionario).get(pk=funcionario_id)
                responsavel_nome = funcionario.nome
            except Funcionario.DoesNotExist:
                responsavel_nome = f"ID {funcionario_id} (Não encontrado)"
            
            return Response({
                "mes": month,
                "ano": year,
                "funcionario_id": funcionario_id,
                "responsavel_nome": responsavel_nome,
                "total_comissao_paga": total_comissao
            })
        else:
            # Group by responsible person (Funcionario)
            comissao_por_pessoa = despesas_comissao_qs.values('responsavel__id', 'responsavel__nome')\
                                   .annotate(total_pago=Sum('valor_pago'))\
                                   .order_by('responsavel__nome')
            
            return Response({
                "mes": month,
                "ano": year,
                "comissao_por_pessoa": list(comissao_por_pessoa) # Convert queryset to list for response
            })

class RelatorioResultadoMensalView(BaseReportView):
    """Relatório de Resultado Financeiro Mensal."""
    def get(self, request):
        params = request.query_params
        try:
            year = int(params.get('year', date.today().year))
            month = int(params.get('month', date.today().month))
        except ValueError:
            return Response({"detail": "Ano e/ou mês inválidos."}, status=status.HTTP_400_BAD_REQUEST)

        # Filter by payment date within the specific month/year
        filters_pagamento = {
            'situacao': 'P',
            'data_pagamento__year': year,
            'data_pagamento__month': month
        }

        receitas_qs = self.get_company_queryset(Receita).filter(**filters_pagamento)
        despesas_qs = self.get_company_queryset(Despesa).filter(**filters_pagamento)

        total_receitas_pagas = receitas_qs.aggregate(total=Sum('valor_pago'))['total'] or Decimal('0.00')
        total_despesas_pagas = despesas_qs.aggregate(total=Sum('valor_pago'))['total'] or Decimal('0.00')

        resultado = total_receitas_pagas - total_despesas_pagas

        return Response({
            "mes": month,
            "ano": year,
            "total_receitas_pagas": total_receitas_pagas,
            "total_despesas_pagas": total_despesas_pagas,
            "resultado_mensal": resultado
        })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dre_consolidado(request):
    """
    Retorna a DRE consolidada com Receitas e Despesas agrupadas por tipo.
    
    Query Parameters:
    - mes: Mês (1-12)
    - ano: Ano (YYYY)
    
    Retorna:
    {
        "receitas": {
            "fixas": 10000.00,
            "variaveis": 5000.00,
            "estornos": -500.00,
            "total": 14500.00
        },
        "despesas": {
            "fixas": 3000.00,
            "variaveis": 2000.00,
            "comissoes": 1000.00,
            "total": 6000.00
        },
        "resultado": 8500.00
    }
    """
    
    try:
        # 🔹 Pegar parâmetros de mês e ano
        mes = request.query_params.get('mes')
        ano = request.query_params.get('ano')
        
        # 🔹 Se não tiver mês/ano, usar mês atual
        if not mes or not ano:
            hoje = datetime.now()
            mes = hoje.month
            ano = hoje.year
        else:
            mes = int(mes)
            ano = int(ano)
        
        # 🔹 Calcular data de início e fim do mês
        data_inicio = f"{ano}-{str(mes).zfill(2)}-01"
        # Último dia do mês
        if mes == 12:
            data_fim = f"{ano + 1}-01-01"
        else:
            data_fim = f"{ano}-{str(mes + 1).zfill(2)}-01"
        data_fim = (datetime.strptime(data_fim, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        
        # 🔹 Filtrar receitas por período do mês
        receitas = Receita.objects.filter(
            company=request.user.company,
            data_vencimento__gte=data_inicio,
            data_vencimento__lte=data_fim
        )
        
        # 🔹 Filtrar despesas por período do mês
        despesas = Despesa.objects.filter(
            company=request.user.company,
            data_vencimento__gte=data_inicio,
            data_vencimento__lte=data_fim
        )
        
        # 🔹 Agrupar receitas por tipo
        receitas_fixas = receitas.filter(tipo='F').aggregate(Sum('valor'))['valor__sum'] or 0
        receitas_variaveis = receitas.filter(tipo='V').aggregate(Sum('valor'))['valor__sum'] or 0
        estornos = receitas.filter(tipo='E').aggregate(Sum('valor'))['valor__sum'] or 0
        
        total_receitas = float(receitas_fixas) + float(receitas_variaveis) + float(estornos)
        
        # 🔹 Agrupar despesas por tipo
        despesas_fixas = despesas.filter(tipo='F').aggregate(Sum('valor'))['valor__sum'] or 0
        despesas_variaveis = despesas.filter(tipo='V').aggregate(Sum('valor'))['valor__sum'] or 0
        comissoes = despesas.filter(tipo='C').aggregate(Sum('valor'))['valor__sum'] or 0
        reembolsos = despesas.filter(tipo='R').aggregate(Sum('valor'))['valor__sum'] or 0
        
        total_despesas = float(despesas_fixas) + float(despesas_variaveis) + float(comissoes) + float(reembolsos)
        
        # 🔹 Calcular resultado
        resultado = total_receitas - total_despesas
        
        # 🔹 Retornar dados formatados
        return Response({
            'receitas': {
                'fixas': float(receitas_fixas),
                'variaveis': float(receitas_variaveis),
                'estornos': float(estornos),
                'total': total_receitas
            },
            'despesas': {
                'fixas': float(despesas_fixas),
                'variaveis': float(despesas_variaveis),
                'comissoes': float(comissoes),
                'reembolsos': float(reembolsos),
                'total': total_despesas
            },
            'resultado': resultado
        }, status=status.HTTP_200_OK)
    
    except ValueError as e:
        return Response({'error': str(e)}, status=400)
    except Exception as e:
        return Response({'error': 'Erro interno'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def balanco_patrimonial(request):
    """
    Retorna o Fluxo de Caixa Realizado (Regime de Caixa) com dois agrupamentos:
    - Por banco (padrão)
    - Por tipo (Receita Fixa, Despesa Variável, Custódia, etc.)

    Considera TODOS os pagamentos (vinculados ou não a receitas/despesas),
    excluindo apenas transferências entre contas (pois se anulam).

    Query Parameters:
    - mes: Mês (1-12)
    - ano: Ano (YYYY)

    Retorna:
    {
        "entradas": {
            "por_banco": [...],
            "por_tipo": [...],
            "total": 70000.00
        },
        "saidas": {
            "por_banco": [...],
            "por_tipo": [...],
            "total": 40000.00
        },
        "resultado": 30000.00
    }
    """

    try:
        # 🔹 Pegar parâmetros de mês e ano
        mes = request.query_params.get('mes')
        ano = request.query_params.get('ano')

        # 🔹 Se não tiver mês/ano, usar mês atual
        if not mes or not ano:
            hoje = datetime.now()
            mes = hoje.month
            ano = hoje.year
        else:
            mes = int(mes)
            ano = int(ano)

        # 🔹 Calcular data de início e fim do mês
        data_inicio = f"{ano}-{str(mes).zfill(2)}-01"
        if mes == 12:
            data_fim = f"{ano + 1}-01-01"
        else:
            data_fim = f"{ano}-{str(mes + 1).zfill(2)}-01"
        data_fim = (datetime.strptime(data_fim, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")

        # 🔹 Buscar todos os pagamentos do mês, excluindo transferências
        # (Transferências se anulam: saída de uma conta = entrada em outra)
        # Primeiro, obter os IDs dos pagamentos que são transferências
        payment_ids_com_transferencia = Allocation.objects.filter(
            payment__company=request.user.company,
            payment__data_pagamento__gte=data_inicio,
            payment__data_pagamento__lte=data_fim,
            transfer__isnull=False
        ).values_list('payment_id', flat=True)

        # Agora buscar pagamentos excluindo os que são transferências
        pagamentos = Payment.objects.filter(
            company=request.user.company,
            data_pagamento__gte=data_inicio,
            data_pagamento__lte=data_fim
        ).exclude(
            id__in=payment_ids_com_transferencia
        ).select_related('conta_bancaria').prefetch_related('allocations__receita', 'allocations__despesa', 'allocations__custodia')

        # 🔹 Buscar todas as alocações dos pagamentos do mês (exceto transferências)
        allocations = Allocation.objects.filter(
            payment__company=request.user.company,
            payment__data_pagamento__gte=data_inicio,
            payment__data_pagamento__lte=data_fim,
            transfer__isnull=True
        ).select_related('payment', 'payment__conta_bancaria', 'receita', 'despesa', 'custodia')

        # 🔹 Dicionários para agrupamentos
        # Por banco
        entradas_por_banco = {}
        saidas_por_banco = {}

        # Por tipo
        entradas_por_tipo = {}
        saidas_por_tipo = {}

        # Mapas de nomes para tipos
        TIPO_RECEITA_MAP = {
            'F': 'Receita Fixa',
            'V': 'Receita Variável',
            'E': 'Estorno',
        }
        TIPO_DESPESA_MAP = {
            'F': 'Despesa Fixa',
            'V': 'Despesa Variável',
            'C': 'Comissionamento',
            'R': 'Reembolso',
        }

        # 🔹 Processar pagamentos para agrupamento por banco
        for pagamento in pagamentos:
            banco_nome = pagamento.conta_bancaria.nome
            valor = float(pagamento.valor)

            if pagamento.tipo == 'E':  # Entrada
                if banco_nome not in entradas_por_banco:
                    entradas_por_banco[banco_nome] = 0
                entradas_por_banco[banco_nome] += valor
            elif pagamento.tipo == 'S':  # Saída
                if banco_nome not in saidas_por_banco:
                    saidas_por_banco[banco_nome] = 0
                saidas_por_banco[banco_nome] += valor

        # 🔹 Processar alocações para agrupamento por tipo
        # Rastrear valor total alocado por pagamento
        valor_alocado_por_pagamento = {}  # payment_id -> total alocado

        for allocation in allocations:
            valor = float(allocation.valor)
            tipo_pagamento = allocation.payment.tipo
            pid = allocation.payment.id

            valor_alocado_por_pagamento[pid] = valor_alocado_por_pagamento.get(pid, 0) + valor

            # Determinar tipo
            if allocation.receita:
                tipo_receita = allocation.receita.tipo
                tipo_nome = TIPO_RECEITA_MAP.get(tipo_receita, 'Outro')
            elif allocation.despesa:
                tipo_despesa = allocation.despesa.tipo
                tipo_nome = TIPO_DESPESA_MAP.get(tipo_despesa, 'Outro')
            elif allocation.custodia:
                # Diferenciar custódia por tipo de pagamento (entrada ou saída)
                if tipo_pagamento == 'E':
                    tipo_nome = 'Valores Reembolsados'
                else:
                    tipo_nome = 'Valores Reembolsáveis'
            else:
                # Alocação sem vínculo definido
                tipo_nome = 'Não Alocado'

            # Agrupar por tipo
            if tipo_pagamento == 'E':
                if tipo_nome not in entradas_por_tipo:
                    entradas_por_tipo[tipo_nome] = 0
                entradas_por_tipo[tipo_nome] += valor
            elif tipo_pagamento == 'S':
                if tipo_nome not in saidas_por_tipo:
                    saidas_por_tipo[tipo_nome] = 0
                saidas_por_tipo[tipo_nome] += valor

        # 🔹 Adicionar pagamentos não alocados (total ou parcialmente)
        for pagamento in pagamentos:
            valor_total = float(pagamento.valor)
            valor_alocado = valor_alocado_por_pagamento.get(pagamento.id, 0)
            valor_nao_alocado = round(valor_total - valor_alocado, 2)

            if valor_nao_alocado > 0:
                tipo_nome = 'Não Alocado'

                if pagamento.tipo == 'E':
                    if tipo_nome not in entradas_por_tipo:
                        entradas_por_tipo[tipo_nome] = 0
                    entradas_por_tipo[tipo_nome] += valor_nao_alocado
                elif pagamento.tipo == 'S':
                    if tipo_nome not in saidas_por_tipo:
                        saidas_por_tipo[tipo_nome] = 0
                    saidas_por_tipo[tipo_nome] += valor_nao_alocado

        # 🔹 Converter dicionários em listas
        entradas_banco_list = [{"banco": banco, "valor": valor} for banco, valor in entradas_por_banco.items()]
        saidas_banco_list = [{"banco": banco, "valor": valor} for banco, valor in saidas_por_banco.items()]

        # Ordem desejada para entradas e saídas
        ORDEM_ENTRADAS = ['Receita Fixa', 'Receita Variável', 'Valores Reembolsados', 'Estorno', 'Não Alocado']
        ORDEM_SAIDAS = ['Despesa Fixa', 'Despesa Variável', 'Valores Reembolsáveis', 'Comissionamento', 'Reembolso', 'Não Alocado']

        # Ordenar entradas conforme ordem especificada
        entradas_tipo_list = []
        for tipo in ORDEM_ENTRADAS:
            if tipo in entradas_por_tipo:
                entradas_tipo_list.append({"tipo": tipo, "valor": entradas_por_tipo[tipo]})
        # Adicionar tipos não mapeados ao final
        for tipo, valor in entradas_por_tipo.items():
            if tipo not in ORDEM_ENTRADAS:
                entradas_tipo_list.append({"tipo": tipo, "valor": valor})

        # Ordenar saídas conforme ordem especificada
        saidas_tipo_list = []
        for tipo in ORDEM_SAIDAS:
            if tipo in saidas_por_tipo:
                saidas_tipo_list.append({"tipo": tipo, "valor": saidas_por_tipo[tipo]})
        # Adicionar tipos não mapeados ao final
        for tipo, valor in saidas_por_tipo.items():
            if tipo not in ORDEM_SAIDAS:
                saidas_tipo_list.append({"tipo": tipo, "valor": valor})

        # 🔹 Calcular totais
        total_entradas = sum(entradas_por_banco.values())
        total_saidas = sum(saidas_por_banco.values())
        resultado = total_entradas - total_saidas

        # 🔹 Retornar dados formatados
        return Response({
            'entradas': {
                'por_banco': entradas_banco_list,
                'por_tipo': entradas_tipo_list,
                'total': total_entradas
            },
            'saidas': {
                'por_banco': saidas_banco_list,
                'por_tipo': saidas_tipo_list,
                'total': total_saidas
            },
            'resultado': resultado
        }, status=status.HTTP_200_OK)

    except ValueError as e:
        return Response({'error': str(e)}, status=400)
    except Exception as e:
        return Response({'error': 'Erro interno'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_conciliacao_bancaria(request):
    """
    Retorna relatório completo da conciliação bancária mensal.

    Query Parameters:
    - mes: Mês (1-12)
    - ano: Ano (YYYY)
    - conta_bancaria_id: (opcional) ID da conta bancária específica

    Retorna informações detalhadas para o usuário finalizar a conciliação:
    - Resumo geral (totais, percentuais)
    - Lançamentos conciliados e não conciliados
    - Receitas, despesas e custódias vinculadas
    - Saldo inicial e final do mês
    - Diferenças e discrepâncias
    """

    try:
        from decimal import Decimal

        # 🔹 Pegar parâmetros
        mes = request.query_params.get('mes')
        ano = request.query_params.get('ano')
        conta_bancaria_id = request.query_params.get('conta_bancaria_id')

        # 🔹 Se não tiver mês/ano, usar mês atual
        if not mes or not ano:
            hoje = datetime.now()
            mes = hoje.month
            ano = hoje.year
        else:
            mes = int(mes)
            ano = int(ano)

        # 🔹 Calcular data de início e fim do mês
        data_inicio = datetime(ano, mes, 1).date()
        if mes == 12:
            data_fim = datetime(ano + 1, 1, 1).date() - timedelta(days=1)
        else:
            data_fim = datetime(ano, mes + 1, 1).date() - timedelta(days=1)

        # 🔹 Query base de pagamentos do mês
        pagamentos_query = Payment.objects.filter(
            company=request.user.company,
            data_pagamento__gte=data_inicio,
            data_pagamento__lte=data_fim
        ).select_related('conta_bancaria')

        # Filtrar por conta bancária específica se fornecida
        if conta_bancaria_id:
            pagamentos_query = pagamentos_query.filter(conta_bancaria_id=conta_bancaria_id)

        # 🔹 Anotar número de alocações e soma dos valores alocados
        from django.db.models import Sum
        pagamentos_query = pagamentos_query.annotate(
            num_allocations=Count('allocations'),
            total_alocado=Sum('allocations__valor')
        )

        pagamentos = list(pagamentos_query)

        # 🔹 Separar conciliados vs não conciliados
        # Um pagamento está completamente conciliado se a soma das alocações = valor do pagamento
        conciliados = [p for p in pagamentos if p.total_alocado is not None and abs(float(p.total_alocado) - float(p.valor)) < 0.01]
        nao_conciliados = [p for p in pagamentos if p.total_alocado is None or abs(float(p.total_alocado) - float(p.valor)) >= 0.01]

        # 🔹 Calcular totais
        total_lancamentos = len(pagamentos)
        total_conciliados = len(conciliados)
        total_nao_conciliados = len(nao_conciliados)
        percentual_conciliado = (total_conciliados / total_lancamentos * 100) if total_lancamentos > 0 else 0

        # 🔹 Valores por tipo
        valor_entradas = sum([float(p.valor) for p in pagamentos if p.tipo == 'E'])
        valor_saidas = sum([float(p.valor) for p in pagamentos if p.tipo == 'S'])
        valor_entradas_conciliadas = sum([float(p.valor) for p in conciliados if p.tipo == 'E'])
        valor_saidas_conciliadas = sum([float(p.valor) for p in conciliados if p.tipo == 'S'])
        valor_entradas_pendentes = sum([float(p.valor) for p in nao_conciliados if p.tipo == 'E'])
        valor_saidas_pendentes = sum([float(p.valor) for p in nao_conciliados if p.tipo == 'S'])

        # 🔹 Buscar alocações do período para estatísticas detalhadas
        # IMPORTANTE: Buscar alocações de TODOS os pagamentos, não só dos conciliados
        # porque lançamentos pendentes podem ter alocações parciais
        allocations = Allocation.objects.filter(
            payment__in=pagamentos
        ).select_related('payment', 'receita', 'despesa', 'custodia')

        # Estatísticas por tipo de vinculação
        receitas_vinculadas = [a for a in allocations if a.receita is not None]
        despesas_vinculadas = [a for a in allocations if a.despesa is not None]
        custodias_vinculadas = [a for a in allocations if a.custodia is not None]

        total_receitas_vinculadas = sum([float(a.valor) for a in receitas_vinculadas])
        total_despesas_vinculadas = sum([float(a.valor) for a in despesas_vinculadas])
        total_custodias_vinculadas = sum([float(a.valor) for a in custodias_vinculadas])

        # 🔹 Agrupar por conta bancária
        contas_resumo = {}
        for p in pagamentos:
            banco_nome = p.conta_bancaria.nome
            banco_id = p.conta_bancaria.id

            if banco_id not in contas_resumo:
                contas_resumo[banco_id] = {
                    'id': banco_id,
                    'nome': banco_nome,
                    'total_lancamentos': 0,
                    'conciliados': 0,
                    'pendentes': 0,
                    'entradas': 0,
                    'saidas': 0
                }

            contas_resumo[banco_id]['total_lancamentos'] += 1
            if p.num_allocations > 0:
                contas_resumo[banco_id]['conciliados'] += 1
            else:
                contas_resumo[banco_id]['pendentes'] += 1

            if p.tipo == 'E':
                contas_resumo[banco_id]['entradas'] += float(p.valor)
            else:
                contas_resumo[banco_id]['saidas'] += float(p.valor)

        # 🔹 Calcular saldo do período
        saldo_periodo = valor_entradas - valor_saidas

        # 🔹 Formatar lançamentos não conciliados para exibição
        nao_conciliados_detalhes = []
        for p in nao_conciliados[:50]:  # Limitar a 50 para não sobrecarregar
            # Calcular valor já alocado deste pagamento
            p_allocations = [a for a in allocations if a.payment_id == p.id]
            valor_alocado = sum(float(a.valor) for a in p_allocations)
            valor_nao_vinculado = float(p.valor) - valor_alocado

            nao_conciliados_detalhes.append({
                'id': p.id,
                'tipo': 'Entrada' if p.tipo == 'E' else 'Saída',
                'valor': float(p.valor),
                'valor_alocado': round(valor_alocado, 2),
                'valor_nao_vinculado': round(valor_nao_vinculado, 2),
                'data': p.data_pagamento.strftime('%d/%m/%Y'),
                'observacao': p.observacao or '',
                'conta_bancaria': p.conta_bancaria.nome
            })

        # 🔹 Formatar lançamentos conciliados para exibição (últimos 20)
        conciliados_detalhes = []
        for p in conciliados[-20:]:
            # Buscar alocações deste pagamento
            p_allocations = [a for a in allocations if a.payment_id == p.id]
            vinculos = []

            for a in p_allocations:
                if a.receita:
                    vinculos.append({
                        'tipo': 'Receita',
                        'descricao': f"{a.receita.cliente.nome} - {a.receita.descricao}" if hasattr(a.receita, 'cliente') else a.receita.descricao,
                        'valor': float(a.valor)
                    })
                elif a.despesa:
                    vinculos.append({
                        'tipo': 'Despesa',
                        'descricao': a.despesa.descricao,
                        'valor': float(a.valor)
                    })
                elif a.custodia:
                    vinculos.append({
                        'tipo': 'Custódia',
                        'descricao': a.custodia.descricao,
                        'valor': float(a.valor)
                    })

            conciliados_detalhes.append({
                'id': p.id,
                'tipo': 'Entrada' if p.tipo == 'E' else 'Saída',
                'valor': float(p.valor),
                'data': p.data_pagamento.strftime('%d/%m/%Y'),
                'observacao': p.observacao or '',
                'conta_bancaria': p.conta_bancaria.nome,
                'vinculos': vinculos
            })

        # 🔹 Status geral da conciliação
        if total_nao_conciliados == 0:
            status_geral = 'Concluída'
            status_cor = 'success'
        elif percentual_conciliado >= 80:
            status_geral = 'Quase Concluída'
            status_cor = 'warning'
        elif percentual_conciliado >= 50:
            status_geral = 'Em Andamento'
            status_cor = 'info'
        else:
            status_geral = 'Pendente'
            status_cor = 'error'

        # 🔹 Retornar dados completos
        return Response({
            'periodo': {
                'mes': mes,
                'ano': ano,
                'data_inicio': data_inicio.strftime('%d/%m/%Y'),
                'data_fim': data_fim.strftime('%d/%m/%Y')
            },
            'resumo': {
                'total_lancamentos': total_lancamentos,
                'total_conciliados': total_conciliados,
                'total_nao_conciliados': total_nao_conciliados,
                'percentual_conciliado': round(percentual_conciliado, 2),
                'status_geral': status_geral,
                'status_cor': status_cor
            },
            'valores': {
                'total_entradas': round(valor_entradas, 2),
                'total_saidas': round(valor_saidas, 2),
                'saldo_periodo': round(saldo_periodo, 2),
                'entradas_conciliadas': round(valor_entradas_conciliadas, 2),
                'saidas_conciliadas': round(valor_saidas_conciliadas, 2),
                'entradas_pendentes': round(valor_entradas_pendentes, 2),
                'saidas_pendentes': round(valor_saidas_pendentes, 2)
            },
            'vinculacoes': {
                'receitas': {
                    'quantidade': len(receitas_vinculadas),
                    'valor_total': round(total_receitas_vinculadas, 2)
                },
                'despesas': {
                    'quantidade': len(despesas_vinculadas),
                    'valor_total': round(total_despesas_vinculadas, 2)
                },
                'custodias': {
                    'quantidade': len(custodias_vinculadas),
                    'valor_total': round(total_custodias_vinculadas, 2)
                }
            },
            'por_conta': list(contas_resumo.values()),
            'lancamentos_pendentes': nao_conciliados_detalhes,
            'lancamentos_conciliados_recentes': conciliados_detalhes,
            'total_pendentes_exibidos': len(nao_conciliados_detalhes),
            'total_pendentes': total_nao_conciliados
        }, status=status.HTTP_200_OK)

    except ValueError as e:
        return Response({'error': str(e)}, status=400)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': f'Erro interno: {str(e)}'}, status=500)


# ──────────────────────────────────────────────
# Subscription ViewSets
# ──────────────────────────────────────────────

class PlanoAssinaturaViewSet(viewsets.ReadOnlyModelViewSet):
    """Public endpoint — lists available subscription plans. No auth required."""
    queryset = PlanoAssinatura.objects.filter(ativo=True).order_by('ordem')
    serializer_class = PlanoAssinaturaSerializer
    permission_classes = []
    pagination_class = None


class AssinaturaViewSet(viewsets.GenericViewSet):
    """Subscription management for the authenticated user's company."""
    serializer_class = AssinaturaEmpresaSerializer
    permission_classes = [permissions.IsAuthenticated]
    # Intentionally does NOT include IsSubscriptionActive — users need access here even after expiry

    def _get_assinatura(self):
        company = self.request.user.company
        return AssinaturaEmpresa.objects.get(company=company)

    @action(detail=False, methods=['get'])
    def status_assinatura(self, request):
        """GET /api/assinatura/status/ — Returns current subscription status."""
        try:
            assinatura = self._get_assinatura()
        except AssinaturaEmpresa.DoesNotExist:
            return Response({'detail': 'Assinatura não encontrada.'}, status=404)
        serializer = self.get_serializer(assinatura)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], throttle_classes=[PaymentRateThrottle])
    def assinar(self, request):
        """
        POST /api/assinatura/assinar/
        Único método suportado: Cartão de Crédito (recorrente, cobrado imediatamente).
        Body: { "plano_slug": "profissional", "ciclo": "MONTHLY",
                "billing_type": "CREDIT_CARD",
                "credit_card": { "holder_name", "number", "expiry_month", "expiry_year", "ccv" },
                "holder_info": { "name", "email", "cpf_cnpj", "phone", "postal_code", "address_number" } }
        """
        plano_slug = request.data.get('plano_slug')
        ciclo = request.data.get('ciclo', 'MONTHLY').upper()
        billing_type = request.data.get('billing_type', '').upper()

        if ciclo not in ('MONTHLY', 'YEARLY'):
            return Response({'detail': 'ciclo deve ser MONTHLY ou YEARLY.'}, status=400)
        if billing_type != 'CREDIT_CARD':
            return Response({'detail': 'Apenas pagamento via Cartão de Crédito é suportado.'}, status=400)

        try:
            plano = PlanoAssinatura.objects.get(slug=plano_slug, ativo=True)
        except PlanoAssinatura.DoesNotExist:
            return Response({'detail': 'Plano não encontrado.'}, status=400)

        company = request.user.company

        cpf_cnpj = (company.cnpj or company.cpf or '').strip()
        if not cpf_cnpj:
            return Response({'detail': 'CPF_CNPJ_MISSING'}, status=400)
        if not _is_valid_cpf_cnpj(cpf_cnpj):
            return Response({'detail': 'CPF/CNPJ da empresa inválido.'}, status=400)

        try:
            with transaction.atomic():
                try:
                    assinatura = AssinaturaEmpresa.objects.select_for_update().get(company=company)
                except AssinaturaEmpresa.DoesNotExist:
                    assinatura = AssinaturaEmpresa.objects.create(
                        company=company,
                        trial_fim=timezone.now(),
                        status='trial',
                    )
                    assinatura = AssinaturaEmpresa.objects.select_for_update().get(pk=assinatura.pk)

                # Ensure Asaas customer exists
                if not assinatura.asaas_customer_id:
                    asaas_customer_id = criar_cliente_asaas(company)
                    assinatura.asaas_customer_id = asaas_customer_id
                    assinatura.save(update_fields=['asaas_customer_id'])
                else:
                    try:
                        atualizar_cliente_asaas(assinatura.asaas_customer_id, company)
                    except HTTPError as e:
                        if e.response is not None and e.response.status_code == 404:
                            logger.warning(f'Customer {assinatura.asaas_customer_id} not found in Asaas, creating new.')
                            new_id = criar_cliente_asaas(company)
                            assinatura.asaas_customer_id = new_id
                            assinatura.save(update_fields=['asaas_customer_id'])
                        else:
                            raise

                # Cancel any existing pending/overdue Asaas subscription before creating a new one
                if assinatura.asaas_subscription_id and assinatura.status != 'active':
                    old_sub_id = assinatura.asaas_subscription_id
                    try:
                        cancelar_assinatura_asaas(old_sub_id)
                        logger.info(f'Cancelled stale subscription {old_sub_id} before re-subscribing.')
                    except Exception as cancel_err:
                        logger.warning(f'Could not cancel old subscription {old_sub_id}: {cancel_err}')

                    # Preserve the old ID in audit trail regardless of cancellation outcome.
                    ids_anteriores = list(assinatura.asaas_subscription_ids_anteriores or [])
                    if old_sub_id not in ids_anteriores:
                        ids_anteriores.append(old_sub_id)

                    assinatura.asaas_subscription_id = None
                    assinatura.asaas_subscription_ids_anteriores = ids_anteriores
                    assinatura.pending_plano = None
                    assinatura.pending_ciclo = None
                    assinatura.save(update_fields=[
                        'asaas_subscription_id',
                        'asaas_subscription_ids_anteriores',
                        'pending_plano',
                        'pending_ciclo',
                    ])

                credit_card = request.data.get('credit_card')
                holder_info = request.data.get('holder_info')

                if not credit_card or not holder_info:
                    return Response({'detail': 'Dados do cartão incompletos.'}, status=400)

                required_card = ['holder_name', 'number', 'expiry_month', 'expiry_year', 'ccv']
                if not all(credit_card.get(f) for f in required_card):
                    return Response({'detail': 'Dados do cartão incompletos.'}, status=400)

                required_holder = ['name', 'cpf_cnpj']
                if not all(holder_info.get(f) for f in required_holder):
                    return Response({'detail': 'Dados do titular incompletos.'}, status=400)
                if not _is_valid_cpf_cnpj(holder_info.get('cpf_cnpj')):
                    return Response({'detail': 'CPF/CNPJ do titular inválido.'}, status=400)

                # Use company email/phone as fallback for holder info
                holder_info.setdefault('email', company.email or '')
                holder_info.setdefault('phone', company.telefone or '')

                try:
                    result = criar_assinatura_cartao_asaas(
                        assinatura.asaas_customer_id, plano, ciclo, credit_card, holder_info
                    )
                except HTTPError as e:
                    # Asaas rejects charges for removed customers — recreate and retry once
                    if e.response is not None and e.response.status_code == 400:
                        body = e.response.text.lower()
                        if 'removido' in body or 'removed' in body:
                            logger.warning(f'Customer {assinatura.asaas_customer_id} is removed in Asaas, recreating.')
                            new_id = criar_cliente_asaas(company)
                            assinatura.asaas_customer_id = new_id
                            assinatura.save(update_fields=['asaas_customer_id'])
                            result = criar_assinatura_cartao_asaas(
                                new_id, plano, ciclo, credit_card, holder_info
                            )
                        else:
                            raise
                    else:
                        raise

                # Activate immediately — Asaas charges the card synchronously
                today = timezone.localdate()
                assinatura.asaas_subscription_id = result['id']
                assinatura.plano = plano
                assinatura.ciclo = ciclo
                assinatura.status = 'active'
                assinatura.pending_plano = None
                assinatura.pending_ciclo = None
                # Next billing: ~1 month/year from today (matches nextDueDate sent to Asaas)
                if ciclo == 'YEARLY':
                    assinatura.proxima_cobranca = _add_one_year_safe(today)
                else:
                    assinatura.proxima_cobranca = _add_one_month_safe(today)
                # Store card summary returned by Asaas
                card_info = result.get('creditCard') or {}
                if card_info.get('creditCardNumber'):
                    assinatura.card_last_four = card_info.get('creditCardNumber')
                    assinatura.card_brand = card_info.get('creditCardBrand') or None
                assinatura.save()

            return Response({'success': True, 'asaas_subscription_id': result['id']})

        except HTTPError as e:
            logger.warning(f'Erro HTTP no gateway Asaas ao criar assinatura: {e}')
            if e.response is not None and 400 <= e.response.status_code < 500:
                try:
                    asaas_errors = e.response.json().get('errors', [])
                    if asaas_errors:
                        msg = asaas_errors[0].get('description', 'Erro no gateway de pagamento.')
                        return Response({'detail': msg}, status=400)
                except Exception:
                    pass
            return Response({'detail': 'Serviço de pagamento temporariamente indisponível. Tente novamente.'}, status=503)
        except RequestException as e:
            logger.warning(f'Falha de comunicação com Asaas ao criar assinatura: {e}')
            return Response({'detail': 'Serviço de pagamento temporariamente indisponível. Tente novamente.'}, status=503)
        except Exception as e:
            logger.error(f'Erro inesperado ao criar assinatura: {e}')
            return Response({'detail': 'Erro ao processar assinatura. Verifique os dados e tente novamente.'}, status=500)

    @action(detail=False, methods=['get'])
    def link_pagamento(self, request):
        """GET /api/assinatura/link_pagamento/ — Returns the pending payment URL for an overdue subscription."""
        try:
            assinatura = self._get_assinatura()
        except AssinaturaEmpresa.DoesNotExist:
            return Response({'detail': 'Assinatura não encontrada.'}, status=404)

        if assinatura.status != 'overdue' and not assinatura.pending_plano:
            return Response({'detail': 'Assinatura não está em atraso.'}, status=400)

        if not assinatura.asaas_subscription_id:
            return Response({'detail': 'ID de assinatura não disponível.'}, status=400)

        try:
            from .asaas_service import obter_url_pagamento_assinatura
            url = obter_url_pagamento_assinatura(assinatura.asaas_subscription_id)
            if not url:
                return Response({'detail': 'Link de pagamento não disponível.'}, status=404)
            return Response({'payment_url': url})
        except RequestException as e:
            logger.warning(f'Falha de comunicação com Asaas ao obter link de pagamento: {e}')
            return Response({'detail': 'Serviço de pagamento temporariamente indisponível. Tente novamente.'}, status=503)
        except Exception as e:
            logger.error(f'Erro ao obter link de pagamento: {e}')
            return Response({'detail': 'Erro ao buscar link de pagamento.'}, status=500)

    @action(detail=False, methods=['get'])
    def pagamentos(self, request):
        """GET /api/assinatura/pagamentos/ — Returns payment history from Asaas."""
        try:
            assinatura = self._get_assinatura()
        except AssinaturaEmpresa.DoesNotExist:
            return Response({'detail': 'Assinatura não encontrada.'}, status=404)

        all_ids = []
        if assinatura.asaas_subscription_id:
            all_ids.append(assinatura.asaas_subscription_id)
        for prev_id in (assinatura.asaas_subscription_ids_anteriores or []):
            if prev_id not in all_ids:
                all_ids.append(prev_id)

        if not all_ids:
            return Response([])

        try:
            from .asaas_service import listar_pagamentos_assinatura
            all_pagamentos = []
            for sub_id in all_ids:
                try:
                    pagamentos = listar_pagamentos_assinatura(sub_id)
                    all_pagamentos.extend(pagamentos)
                except RequestException as sub_e:
                    logger.warning(f'Erro de comunicação ao buscar pagamentos da assinatura {sub_id}: {sub_e}')
                except Exception as sub_e:
                    logger.warning(f'Erro ao buscar pagamentos da assinatura {sub_id}: {sub_e}')
            all_pagamentos.sort(key=lambda p: p.get('dueDate', ''), reverse=True)
            return Response(all_pagamentos[:20])
        except RequestException as e:
            logger.warning(f'Falha de comunicação com Asaas ao buscar histórico de pagamentos: {e}')
            return Response({'detail': 'Serviço de pagamento temporariamente indisponível. Tente novamente.'}, status=503)
        except Exception as e:
            logger.error(f'Erro ao buscar histórico de pagamentos: {e}')
            return Response({'detail': 'Erro ao buscar histórico de pagamentos.'}, status=500)

    @action(detail=False, methods=['post'])
    def cancelar(self, request):
        """
        POST /api/assinatura/cancelar/ — Cancels the subscription in Asaas immediately,
        then marks it locally as cancelled. Access remains until proxima_cobranca.
        """
        try:
            assinatura = self._get_assinatura()
        except AssinaturaEmpresa.DoesNotExist:
            return Response({'detail': 'Assinatura não encontrada.'}, status=404)

        if assinatura.status not in ('active', 'overdue'):
            return Response({'detail': 'Sem assinatura ativa para cancelar.'}, status=400)

        if assinatura.asaas_subscription_id:
            try:
                cancelar_assinatura_asaas(assinatura.asaas_subscription_id)
            except RequestException as e:
                logger.warning(f'Falha de comunicação com Asaas ao cancelar assinatura: {e}')
                return Response(
                    {'detail': 'Serviço de pagamento temporariamente indisponível. Tente novamente.'},
                    status=503,
                )
            except HTTPError as e:
                logger.warning(f'Erro HTTP no Asaas ao cancelar assinatura: {e}')
                return Response(
                    {'detail': 'Não foi possível cancelar a assinatura no gateway de pagamento. Tente novamente.'},
                    status=502,
                )

        assinatura.status = 'cancelled'
        assinatura.save(update_fields=['status'])
        return Response({'detail': 'Assinatura cancelada. Seu acesso continua até o fim do período pago.'})

    @action(detail=False, methods=['post'])
    def reativar(self, request):
        """
        POST /api/assinatura/reativar/
        Reactivates a cancelled subscription that still has access (proxima_cobranca in the future).
        Note: cancelar() now cancels the subscription in Asaas. Reactivation here only restores the
        local status; the user will need to subscribe again when proxima_cobranca expires.
        """
        from datetime import date
        try:
            assinatura = self._get_assinatura()
        except AssinaturaEmpresa.DoesNotExist:
            return Response({'detail': 'Assinatura não encontrada.'}, status=404)

        if assinatura.status != 'cancelled':
            return Response({'detail': 'A assinatura não está cancelada.'}, status=400)

        if not assinatura.proxima_cobranca or assinatura.proxima_cobranca < date.today():
            return Response({'detail': 'O período pago já expirou. Assine um novo plano.'}, status=400)

        assinatura.status = 'active'
        assinatura.save(update_fields=['status'])
        serializer = self.get_serializer(assinatura)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], throttle_classes=[PaymentRateThrottle])
    def atualizar_cartao(self, request):
        """
        POST /api/assinatura/atualizar_cartao/
        Updates the credit card used for the active subscription.
        Body: { credit_card: {...}, holder_info: {...} }
        """
        try:
            assinatura = self._get_assinatura()
        except AssinaturaEmpresa.DoesNotExist:
            return Response({'detail': 'Assinatura não encontrada.'}, status=404)

        if assinatura.status not in ('active', 'overdue'):
            return Response({'detail': 'Sem assinatura ativa para atualizar o cartão.'}, status=400)

        if not assinatura.asaas_subscription_id:
            return Response({'detail': 'ID de assinatura não disponível.'}, status=400)

        credit_card = request.data.get('credit_card')
        holder_info = request.data.get('holder_info')

        if not credit_card or not holder_info:
            return Response({'detail': 'Dados do cartão incompletos.'}, status=400)

        required_card = ['holder_name', 'number', 'expiry_month', 'expiry_year', 'ccv']
        if not all(credit_card.get(f) for f in required_card):
            return Response({'detail': 'Dados do cartão incompletos.'}, status=400)

        required_holder = ['name', 'cpf_cnpj']
        if not all(holder_info.get(f) for f in required_holder):
            return Response({'detail': 'Dados do titular incompletos.'}, status=400)
        if not _is_valid_cpf_cnpj(holder_info.get('cpf_cnpj')):
            return Response({'detail': 'CPF/CNPJ do titular inválido.'}, status=400)

        try:
            card_data = atualizar_cartao_assinatura(
                assinatura.asaas_subscription_id, credit_card, holder_info
            )
            card_info = card_data.get('creditCard') or card_data
            if card_info.get('creditCardNumber'):
                assinatura.card_last_four = card_info.get('creditCardNumber')
                assinatura.card_brand = card_info.get('creditCardBrand') or None
                assinatura.save(update_fields=['card_last_four', 'card_brand'])
            return Response({'success': True})
        except HTTPError as e:
            logger.warning(f'Erro HTTP no gateway Asaas ao atualizar cartão: {e}')
            if e.response is not None and 400 <= e.response.status_code < 500:
                try:
                    asaas_errors = e.response.json().get('errors', [])
                    if asaas_errors:
                        msg = asaas_errors[0].get('description', 'Erro no gateway de pagamento.')
                        return Response({'detail': msg}, status=400)
                except Exception:
                    pass
            return Response({'detail': 'Serviço de pagamento temporariamente indisponível. Tente novamente.'}, status=503)
        except RequestException as e:
            logger.warning(f'Falha de comunicação com Asaas ao atualizar cartão: {e}')
            return Response({'detail': 'Serviço de pagamento temporariamente indisponível. Tente novamente.'}, status=503)
        except Exception as e:
            logger.error(f'Erro inesperado ao atualizar cartão: {e}')
            return Response({'detail': 'Erro ao atualizar cartão. Verifique os dados e tente novamente.'}, status=500)


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny


@api_view(['POST'])
@permission_classes([AllowAny])
def register_view(request):
    """
    POST /api/register/
    Cria uma nova empresa + usuário administrador e retorna tokens JWT.
    Body: { nome_empresa, cpf_cnpj, username, email, senha, nome? }
    """
    from rest_framework_simplejwt.tokens import RefreshToken

    nome_empresa = (request.data.get('nome_empresa') or '').strip()
    cpf_cnpj     = (request.data.get('cpf_cnpj') or '').strip()
    username     = (request.data.get('username') or '').strip()
    email        = (request.data.get('email') or '').strip()
    senha        = (request.data.get('senha') or '').strip()
    nome         = (request.data.get('nome') or '').strip()

    errors = {}
    if not nome_empresa:
        errors['nome_empresa'] = 'Nome do escritório é obrigatório.'
    cpf_cnpj_digits = _normalize_digits(cpf_cnpj)
    if not cpf_cnpj_digits:
        errors['cpf_cnpj'] = 'CPF ou CNPJ é obrigatório.'
    elif not _is_valid_cpf_cnpj(cpf_cnpj_digits):
        errors['cpf_cnpj'] = 'CPF/CNPJ inválido.'
    if not username:
        errors['username'] = 'Nome de usuário é obrigatório.'
    elif CustomUser.objects.filter(username=username).exists():
        errors['username'] = 'Este nome de usuário já está em uso.'
    if not email:
        errors['email'] = 'E-mail é obrigatório.'
    elif CustomUser.objects.filter(email=email).exists():
        errors['email'] = 'Este e-mail já está cadastrado.'
    if not senha:
        errors['senha'] = 'Senha é obrigatória.'
    elif len(senha) < 8:
        errors['senha'] = 'A senha deve ter pelo menos 8 caracteres.'
    if errors:
        return Response(errors, status=400)

    try:
        company_data = {'name': nome_empresa}
        if len(cpf_cnpj_digits) == 14:
            company_data['cnpj'] = cpf_cnpj_digits
        else:
            company_data['cpf'] = cpf_cnpj_digits
        company = Company.objects.create(**company_data)

        first_name = nome.split()[0] if nome else ''
        last_name  = ' '.join(nome.split()[1:]) if nome and len(nome.split()) > 1 else ''

        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password=senha,
            company=company,
            first_name=first_name,
            last_name=last_name,
        )

        refresh = RefreshToken.for_user(user)
        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
        }, status=201)

    except Exception as e:
        logger.error(f'Erro ao registrar usuário: {e}')
        return Response({'detail': 'Erro ao criar conta. Tente novamente.'}, status=500)


@csrf_exempt
@require_POST
def asaas_webhook(request):
    """
    POST /api/asaas/webhook/
    Receives Asaas webhook events and updates subscription status accordingly.
    """
    token = (
        request.headers.get('asaas-access-token')
        or request.headers.get('x-asaas-access-token')
        or request.GET.get('token', '')
    )
    configured_token = (django_settings.ASAAS_WEBHOOK_TOKEN or '').strip()
    # Fail closed when webhook secret is not configured.
    if not configured_token:
        logger.error('Asaas webhook rejected: ASAAS_WEBHOOK_TOKEN not configured.')
        return JsonResponse({'detail': 'Unauthorized'}, status=401)
    if not token or not secrets.compare_digest(str(token), configured_token):
        return JsonResponse({'detail': 'Unauthorized'}, status=401)

    try:
        payload = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'detail': 'Invalid JSON'}, status=400)

    event_type = payload.get('event', '')
    subscription_id = (
        payload.get('payment', {}).get('subscription')
        or payload.get('subscription', {}).get('id')
        or ''
    )
    payment_id = payload.get('payment', {}).get('id', '')

    try:
        with transaction.atomic():
            already_processed = WebhookLog.objects.filter(
                event_type=event_type,
                asaas_payment_id=payment_id,
                processed=True,
            ).exists()
            if already_processed:
                return JsonResponse({'detail': 'already processed'}, status=200)

            log = WebhookLog.objects.create(
                event_type=event_type,
                asaas_subscription_id=subscription_id,
                asaas_payment_id=payment_id,
                payload=payload,
            )

            if subscription_id:
                assinatura = AssinaturaEmpresa.objects.select_for_update().filter(
                    asaas_subscription_id=subscription_id
                ).first()

                if assinatura:
                    if event_type == 'PAYMENT_RECEIVED':
                        import datetime as _dt
                        # Never restore access to a subscription the user explicitly cancelled.
                        if assinatura.status == 'cancelled':
                            logger.warning(
                                f'PAYMENT_RECEIVED for cancelled subscription '
                                f'{subscription_id} (company {assinatura.company_id}). '
                                f'Ignoring to preserve cancellation intent.'
                            )
                        else:
                            assinatura.status = 'active'
                            # Compute next billing date from current payment's dueDate
                            due_date_str = payload.get('payment', {}).get('dueDate')
                            if due_date_str:
                                try:
                                    due = _dt.date.fromisoformat(due_date_str)
                                    ciclo = (assinatura.pending_ciclo or assinatura.ciclo or 'MONTHLY')
                                    if ciclo == 'YEARLY':
                                        assinatura.proxima_cobranca = _add_one_year_safe(due)
                                    else:
                                        assinatura.proxima_cobranca = _add_one_month_safe(due)
                                except Exception:
                                    pass
                            if assinatura.pending_plano:
                                assinatura.plano = assinatura.pending_plano
                                assinatura.ciclo = assinatura.pending_ciclo or 'MONTHLY'
                                assinatura.pending_plano = None
                                assinatura.pending_ciclo = None
                                assinatura.save(update_fields=['status', 'plano', 'ciclo', 'pending_plano', 'pending_ciclo', 'proxima_cobranca'])
                            else:
                                assinatura.save(update_fields=['status', 'proxima_cobranca'])
                    elif event_type == 'PAYMENT_OVERDUE':
                        # Don't overwrite an intentional cancellation.
                        if assinatura.status != 'cancelled':
                            assinatura.status = 'overdue'
                            assinatura.save(update_fields=['status'])
                        else:
                            logger.info(
                                f'PAYMENT_OVERDUE ignored for already-cancelled '
                                f'subscription {subscription_id}.'
                            )
                    elif event_type == 'PAYMENT_REFUSED':
                        if assinatura.status != 'cancelled':
                            assinatura.status = 'payment_failed'
                            assinatura.save(update_fields=['status'])
                    elif event_type in ('SUBSCRIPTION_CANCELLED', 'SUBSCRIPTION_DELETED'):
                        assinatura.status = 'cancelled'
                        assinatura.asaas_subscription_id = None
                        assinatura.save(update_fields=['status', 'asaas_subscription_id'])

            log.processed = True
            log.save(update_fields=['processed'])
    except Exception as e:
        logger.error(f'Asaas webhook processing error: {e}')
        try:
            WebhookLog.objects.create(
                event_type=event_type,
                asaas_subscription_id=subscription_id,
                payload=payload,
                error=str(e),
            )
        except Exception:
            pass

    return JsonResponse({'received': True})
